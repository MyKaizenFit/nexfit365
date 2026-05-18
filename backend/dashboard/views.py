from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count, Avg, Sum, Max, Min
from django.utils import timezone
from django.core.cache import cache
from datetime import timedelta, datetime
import ast
import re
import calendar

from .models import (
    DashboardData,
    UserStats,
    WellnessTip,
    CoachingPlan,
    CoachingInquiry,
    DefaultPlanConfiguration,
    HelpSettings,
    ProblemReport,
)
from .serializers import (
    DashboardDataSerializer, DashboardTodaySerializer, 
    DashboardWeeklySerializer, DashboardMonthlySerializer,
    DashboardStatsSerializer, WellnessTipSerializer,
    CoachingPlanSerializer, CoachingInquirySerializer, CoachingInquiryCreateSerializer, CoachingInquiryAdminUpdateSerializer,
    DefaultPlanConfigurationSerializer, DefaultPlanConfigurationCreateUpdateSerializer,
    HelpSettingsSerializer, ProblemReportSerializer, ProblemReportCreateSerializer
)
from .permissions import DashboardPermission, DashboardDataPermission
from accounts.models import CustomUser


class DashboardViewSet(viewsets.ModelViewSet):
    """
    ViewSet para dashboard del usuario
    """
    serializer_class = DashboardDataSerializer
    permission_classes = [DashboardPermission]
    
    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return DashboardData.objects.none()
        user_id = self.kwargs.get("user_id")
        if user_id:
            return DashboardData.objects.filter(user_id=user_id)
        return DashboardData.objects.filter(user=self.request.user)
    
    @action(detail=False, methods=["get"])
    def today(self, request, user_id=None):
        """Obtener dashboard del día actual"""
        user_id = user_id or request.user.id
        user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar cache
        cache_key = f"dashboard_today_{user_id}_{timezone.now().date()}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        today = timezone.now().date()
        
        # Datos de nutrición del día
        from nutrition.models import MealLog, NutritionPlan
        active_plan = NutritionPlan.objects.filter(
            user=user, is_active=True
        ).select_related('user').prefetch_related('meals').first()
        
        if active_plan:
            meals_planned = active_plan.meals.count()
            meals_completed = MealLog.objects.filter(
                user=user, date=today, completed=True
            ).select_related('user', 'recipe').count()
            
            # Calorías consumidas (de logs del día)
            calories_consumed = MealLog.objects.filter(
                user=user, date=today
            ).select_related('user', 'recipe').aggregate(
                total_calories=Sum("calories")
            )["total_calories"] or 0
            
            calories_target = active_plan.daily_calories or 0
            nutrition_adherence = (meals_completed / meals_planned * 100) if meals_planned > 0 else 0
        else:
            meals_planned = meals_completed = calories_consumed = calories_target = 0
            nutrition_adherence = 0
        
        # Datos de entrenamiento del día
        from workouts.models import WorkoutLog, WorkoutProgram
        active_program = WorkoutProgram.objects.filter(user=user, is_active=True).first()
        
        if active_program:
            # Encontrar el día de entrenamiento para hoy
            today_name = today.strftime("%A").lower()
            workout_day = active_program.days.filter(
                Q(day__iexact=today_name) | Q(day__iexact=f"day {today.weekday() + 1}")
            ).first()
            
            workout_planned = workout_day is not None and not workout_day.is_rest_day
            workout_log = WorkoutLog.objects.select_related(
                'user', 'workout_program'
            ).filter(
                user=user, date=today, workout_day=workout_day
            ).first() if workout_day else None
            
            workout_completed = workout_log.completed if workout_log else False
            workout_duration = workout_log.duration_minutes if workout_log else None
            workout_rating = workout_log.rating if workout_log else None
        else:
            workout_planned = workout_completed = False
            workout_duration = workout_rating = None
        
        # Datos de progreso del día
        from progress.models import WeightEntry
        current_weight = WeightEntry.objects.filter(
            user=user
        ).order_by("-date").values_list("weight", flat=True).first()
        
        # Cambio de peso desde ayer
        yesterday_weight = WeightEntry.objects.select_related('user').filter(
            user=user, date=today - timedelta(days=1)
        ).values_list("weight", flat=True).first()
        
        weight_change_today = None
        if current_weight and yesterday_weight:
            weight_change_today = current_weight - yesterday_weight
        
        # Logros del día
        from achievements.models import UserAchievement
        achievements_unlocked_today = UserAchievement.objects.filter(
            user=user, unlocked_at__date=today
        ).count()
        
        points_earned_today = UserAchievement.objects.filter(
            user=user, unlocked_at__date=today
        ).aggregate(
            total_points=Sum("achievement__points")
        )["total_points"] or 0
        
        data = {
            "date": today,
            "meals_planned": meals_planned,
            "meals_completed": meals_completed,
            "calories_consumed": calories_consumed,
            "calories_target": calories_target,
            "nutrition_adherence": round(nutrition_adherence, 2),
            "workout_planned": workout_planned,
            "workout_completed": workout_completed,
            "workout_duration": workout_duration,
            "workout_rating": workout_rating,
            "current_weight": current_weight,
            "weight_change_today": weight_change_today,
            "achievements_unlocked_today": achievements_unlocked_today,
            "points_earned_today": points_earned_today,
        }
        
        # Cache por 5 minutos
        cache.set(cache_key, data, 300)
        
        serializer = DashboardTodaySerializer(data)
        return Response(serializer.data)
    
    @action(detail=False, methods=["get"])
    def weekly(self, request, user_id=None):
        """Obtener dashboard semanal"""
        user_id = user_id or request.user.id
        user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar cache
        cache_key = f"dashboard_weekly_{user_id}_{timezone.now().date()}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        # Datos de nutrición semanal
        from nutrition.models import MealLog, NutritionPlan
        active_plan = NutritionPlan.objects.filter(
            user=user, is_active=True
        ).select_related('user').prefetch_related('meals').first()
        
        if active_plan:
            meals_planned_week = active_plan.meals.count() * 7
            meals_completed_week = MealLog.objects.filter(
                user=user, date__range=[week_start, week_end], completed=True
            ).count()
            
            total_calories_week = MealLog.objects.filter(
                user=user, date__range=[week_start, week_end]
            ).aggregate(
                total_calories=Sum("meal__calories")
            )["total_calories"] or 0
            
            average_calories_day = total_calories_week / 7 if total_calories_week > 0 else 0
            nutrition_adherence_week = (meals_completed_week / meals_planned_week * 100) if meals_planned_week > 0 else 0
        else:
            meals_planned_week = meals_completed_week = total_calories_week = 0
            average_calories_day = nutrition_adherence_week = 0
        
        # Datos de entrenamiento semanal
        from workouts.models import WorkoutLog, WorkoutProgram
        active_program = WorkoutProgram.objects.filter(user=user, is_active=True).first()
        
        if active_program:
            workouts_planned_week = active_program.days.filter(is_rest_day=False).count()
            workouts_completed_week = WorkoutLog.objects.filter(
                user=user, date__range=[week_start, week_end], completed=True
            ).count()
            
            total_workout_time = WorkoutLog.objects.filter(
                user=user, date__range=[week_start, week_end]
            ).aggregate(
                total_time=Sum("duration_minutes")
            )["total_time"] or 0
            
            average_workout_rating = WorkoutLog.objects.filter(
                user=user, date__range=[week_start, week_end], rating__isnull=False
            ).aggregate(
                avg_rating=Avg("rating")
            )["avg_rating"] or 0
        else:
            workouts_planned_week = workouts_completed_week = total_workout_time = 0
            average_workout_rating = 0
        
        # Datos de progreso semanal
        from progress.models import WeightEntry, ProgressPhoto
        weight_change_week = None
        weight_change_percentage = None
        
        start_weight = WeightEntry.objects.filter(
            user=user, date__range=[week_start, week_end]
        ).order_by("date").values_list("weight", flat=True).first()
        
        end_weight = WeightEntry.objects.filter(
            user=user, date__range=[week_start, week_end]
        ).order_by("-date").values_list("weight", flat=True).first()
        
        if start_weight and end_weight:
            weight_change_week = end_weight - start_weight
            weight_change_percentage = (weight_change_week / start_weight) * 100
        
        progress_photos_week = ProgressPhoto.objects.filter(
            user=user, date__range=[week_start, week_end]
        ).count()
        
        # Logros semanales
        from achievements.models import UserAchievement
        achievements_unlocked_week = UserAchievement.objects.filter(
            user=user, unlocked_at__date__range=[week_start, week_end]
        ).count()
        
        points_earned_week = UserAchievement.objects.filter(
            user=user, unlocked_at__date__range=[week_start, week_end]
        ).aggregate(
            total_points=Sum("achievement__points")
        )["total_points"] or 0
        
        # Tendencias (simplificado)
        trend_nutrition = "stable"
        if nutrition_adherence_week > 80:
            trend_nutrition = "improving"
        elif nutrition_adherence_week < 50:
            trend_nutrition = "declining"
        
        trend_workouts = "stable"
        if workouts_completed_week > workouts_planned_week * 0.8:
            trend_workouts = "improving"
        elif workouts_completed_week < workouts_planned_week * 0.5:
            trend_workouts = "declining"
        
        trend_progress = "stable"
        if weight_change_week and weight_change_week < 0:  # Pérdida de peso
            trend_progress = "improving"
        elif weight_change_week and weight_change_week > 0:  # Ganancia de peso
            trend_progress = "declining"
        
        data = {
            "week_start": week_start,
            "week_end": week_end,
            "nutrition_adherence_week": round(nutrition_adherence_week, 2),
            "total_calories_week": total_calories_week,
            "average_calories_day": round(average_calories_day, 2),
            "meals_completed_week": meals_completed_week,
            "meals_planned_week": meals_planned_week,
            "workouts_completed_week": workouts_completed_week,
            "workouts_planned_week": workouts_planned_week,
            "total_workout_time": total_workout_time,
            "average_workout_rating": round(average_workout_rating, 2),
            "weight_change_week": weight_change_week,
            "weight_change_percentage": round(weight_change_percentage, 2) if weight_change_percentage else None,
            "progress_photos_week": progress_photos_week,
            "achievements_unlocked_week": achievements_unlocked_week,
            "points_earned_week": points_earned_week,
            "trend_nutrition": trend_nutrition,
            "trend_workouts": trend_workouts,
            "trend_progress": trend_progress,
        }
        
        # Cache por 15 minutos
        cache.set(cache_key, data, 900)
        
        serializer = DashboardWeeklySerializer(data)
        return Response(serializer.data)
    
    @action(detail=False, methods=["get"])
    def monthly(self, request, user_id=None):
        """Obtener dashboard mensual"""
        user_id = user_id or request.user.id
        user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar cache
        cache_key = f"dashboard_monthly_{user_id}_{timezone.now().strftime('%Y-%m')}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        today = timezone.now().date()
        month_start = today.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Resumen mensual
        from workouts.models import WorkoutLog
        from nutrition.models import MealLog
        from progress.models import ProgressPhoto
        from achievements.models import UserAchievement
        
        total_workouts = WorkoutLog.objects.filter(
            user=user, date__range=[month_start, month_end]
        ).count()
        
        total_meals_logged = MealLog.objects.filter(
            user=user, date__range=[month_start, month_end]
        ).count()
        
        total_progress_photos = ProgressPhoto.objects.filter(
            user=user, date__range=[month_start, month_end]
        ).count()
        
        total_achievements = UserAchievement.objects.filter(
            user=user, unlocked_at__date__range=[month_start, month_end]
        ).count()
        
        # Promedios mensuales
        average_workout_rating = WorkoutLog.objects.filter(
            user=user, date__range=[month_start, month_end], rating__isnull=False
        ).aggregate(
            avg_rating=Avg("rating")
        )["avg_rating"] or 0
        
        # Adherencia nutricional mensual
        active_plan = user.nutrition_plans.filter(is_active=True).first()
        if active_plan:
            meals_planned_month = active_plan.meals.count() * month_end.day
            meals_completed_month = MealLog.objects.filter(
                user=user, date__range=[month_start, month_end], completed=True
            ).count()
            average_nutrition_adherence = (meals_completed_month / meals_planned_month * 100) if meals_planned_month > 0 else 0
        else:
            average_nutrition_adherence = 0
        
        # Calorías promedio diarias
        total_calories_month = MealLog.objects.filter(
            user=user, date__range=[month_start, month_end]
        ).aggregate(
            total_calories=Sum("calories")
        )["total_calories"] or 0
        
        average_calories_day = total_calories_month / month_end.day if total_calories_month > 0 else 0
        
        # Objetivos mensuales (simplificado)
        goals_achieved = 0
        goals_total = 3  # Ejemplo: 3 objetivos básicos
        
        if total_workouts >= 20:  # Objetivo: 20 entrenamientos
            goals_achieved += 1
        if average_nutrition_adherence >= 70:  # Objetivo: 70% adherencia
            goals_achieved += 1
        if total_progress_photos >= 4:  # Objetivo: 4 fotos de progreso
            goals_achieved += 1
        
        goals_completion_rate = (goals_achieved / goals_total) * 100
        
        data = {
            "month": today.strftime("%Y-%m"),
            "year": today.year,
            "total_workouts": total_workouts,
            "total_meals_logged": total_meals_logged,
            "total_progress_photos": total_progress_photos,
            "total_achievements": total_achievements,
            "average_workout_rating": round(average_workout_rating, 2),
            "average_nutrition_adherence": round(average_nutrition_adherence, 2),
            "average_calories_day": round(average_calories_day, 2),
            "goals_achieved": goals_achieved,
            "goals_total": goals_total,
            "goals_completion_rate": round(goals_completion_rate, 2),
        }
        
        # Cache por 1 hora
        cache.set(cache_key, data, 3600)
        
        serializer = DashboardMonthlySerializer(data)
        return Response(serializer.data)


    @action(detail=False, methods=["get"])
    def stats(self, request, user_id=None):
        """Obtener estadísticas generales del dashboard"""
        user_id = user_id or request.user.id
        user = get_object_or_404(CustomUser, id=user_id)
        
        # Verificar cache
        cache_key = f"dashboard_stats_{user_id}"
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        # Días activos (días con al menos un log)
        from workouts.models import WorkoutLog
        from nutrition.models import MealLog
        from progress.models import WeightEntry
        
        active_dates = set()
        
        # Fechas con entrenamientos
        workout_dates = WorkoutLog.objects.filter(user=user).values_list("date", flat=True)
        active_dates.update(workout_dates)
        
        # Fechas con comidas
        meal_dates = MealLog.objects.filter(user=user).values_list("date", flat=True)
        active_dates.update(meal_dates)
        
        # Fechas con peso
        weight_dates = WeightEntry.objects.filter(user=user).values_list("date", flat=True)
        active_dates.update(weight_dates)
        
        total_days_active = len(active_dates)
        
        # Racha actual (días consecutivos)
        current_streak = 0
        longest_streak = 0
        current_streak_temp = 0
        
        if active_dates:
            sorted_dates = sorted(active_dates, reverse=True)
            today = timezone.now().date()
            
            for i, date in enumerate(sorted_dates):
                if i == 0 and date == today:
                    current_streak_temp = 1
                elif i > 0 and (sorted_dates[i-1] - date).days == 1:
                    current_streak_temp += 1
                else:
                    if current_streak_temp > longest_streak:
                        longest_streak = current_streak_temp
                    current_streak_temp = 1
            
            if current_streak_temp > longest_streak:
                longest_streak = current_streak_temp
            
            current_streak = current_streak_temp
        
        # Puntos totales
        from achievements.models import UserAchievement
        total_points = UserAchievement.objects.filter(user=user).aggregate(
            total_points=Sum("achievement__points")
        )["total_points"] or 0
        
        # Nivel basado en puntos (ejemplo: cada 100 puntos = 1 nivel)
        level = (total_points // 100) + 1
        next_level_points = level * 100
        progress_to_next_level = (total_points % 100) / 100 * 100
        
        data = {
            "user_id": user_id,
            "total_days_active": total_days_active,
            "current_streak": current_streak,
            "longest_streak": longest_streak,
            "total_points": total_points,
            "level": level,
            "next_level_points": next_level_points,
            "progress_to_next_level": round(progress_to_next_level, 2),
        }
        
        # Cache por 2 horas
        cache.set(cache_key, data, 7200)
        
        serializer = DashboardStatsSerializer(data)
        return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subscription_status(request):
    """Estado de prueba/suscripción del usuario autenticado."""
    user = request.user
    user.refresh_membership_state(commit=True)

    ends_at = user.trial_ends_at if user.subscription_status == 'trial' else user.subscription_ends_at

    return Response({
        'status': user.subscription_status,
        'plan': user.subscription_plan,
        'role': user.role,
        'is_active': user.has_active_membership,
        'trial_used': user.trial_started_at is not None,
        'can_start_trial': user.trial_started_at is None,
        'days_remaining': user.membership_days_remaining,
        'trial_started_at': user.trial_started_at.isoformat() if user.trial_started_at else None,
        'ends_at': ends_at.isoformat() if ends_at else None,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_free_trial(request):
    """Activar una prueba gratuita de 7 días para el usuario actual."""
    user = request.user

    try:
        ends_at = user.start_free_trial(days=7)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'message': 'Prueba gratuita activada correctamente.',
        'status': user.subscription_status,
        'plan': user.subscription_plan,
        'role': user.role,
        'days_remaining': user.membership_days_remaining,
        'trial_started_at': user.trial_started_at.isoformat() if user.trial_started_at else None,
        'ends_at': ends_at.isoformat() if ends_at else None,
    }, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def user_stats(request):
    """
    Obtener o actualizar estadísticas del usuario
    """
    user = request.user
    
    if request.method == 'GET':
        # Obtener o crear estadísticas del usuario
        stats, created = UserStats.objects.get_or_create(
            user=user,
            defaults={
                'calories_goal': 2000,
                'workouts_goal': 5,
                'transformation_start_date': timezone.now().date(),
                'next_review_date': timezone.now().date() + timedelta(days=30),
            }
        )
        
        # Calcular datos en tiempo real
        today = timezone.now().date()
        
        # Calorías del día
        from nutrition.models import MealLog
        calories_today = MealLog.objects.filter(
            user=user, date=today
        ).aggregate(
            total_calories=Sum("calories")
        )["total_calories"] or 0
        
        # Entrenamientos de la semana
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        from workouts.models import WorkoutLog
        workouts_this_week = WorkoutLog.objects.filter(
            user=user, date__range=[week_start, week_end], completed=True
        ).count()
        
        # Peso actual
        from progress.models import WeightEntry
        current_weight = WeightEntry.objects.filter(
            user=user
        ).order_by("-date").values_list("weight", flat=True).first()
        
        if current_weight and not stats.current_weight:
            stats.current_weight = current_weight
            stats.save()
        
        # Calcular días en transformación
        days_in_transformation = stats.days_in_transformation
        
        # Calcular cambio de peso
        weight_change = stats.weight_change
        
        # Próxima revisión
        next_review = "Próximamente"
        if stats.next_review_date:
            days_until_review = (stats.next_review_date - today).days
            if days_until_review > 0:
                next_review = f"{days_until_review} días"
            elif days_until_review == 0:
                next_review = "Hoy"
            else:
                next_review = "Vencida"
        
        response_data = {
            "caloriesToday": calories_today,
            "caloriesGoal": stats.calories_goal,
            "currentWeight": float(stats.current_weight) if stats.current_weight else None,
            "targetWeight": float(stats.weight_goal) if stats.weight_goal else None,
            "weightChange": weight_change,
            "workoutsThisWeek": workouts_this_week,
            "workoutsGoal": stats.workouts_goal,
            "nextReview": next_review,
            "daysInTransformation": days_in_transformation,
        }
        
        return Response(response_data)
    
    elif request.method == 'PUT':
        # Actualizar estadísticas del usuario
        data = request.data
        
        stats, created = UserStats.objects.get_or_create(
            user=user,
            defaults={
                'calories_goal': 2000,
                'workouts_goal': 5,
                'transformation_start_date': timezone.now().date(),
                'next_review_date': timezone.now().date() + timedelta(days=30),
            }
        )
        
        # Actualizar campos permitidos
        if 'calories_goal' in data:
            stats.calories_goal = data['calories_goal']
        if 'workouts_goal' in data:
            stats.workouts_goal = data['workouts_goal']
        if 'current_weight' in data:
            stats.current_weight = data['current_weight']
        if 'starting_weight' in data:
            stats.starting_weight = data['starting_weight']
        if 'weight_goal' in data:
            stats.weight_goal = data['weight_goal']
        if 'transformation_start_date' in data:
            stats.transformation_start_date = data['transformation_start_date']
        if 'next_review_date' in data:
            stats.next_review_date = data['next_review_date']
        
        stats.save()
        
        return Response({
            "message": "Estadísticas actualizadas correctamente",
            "stats": {
                "calories_goal": stats.calories_goal,
                "workouts_goal": stats.workouts_goal,
                "current_weight": float(stats.current_weight) if stats.current_weight else None,
                "starting_weight": float(stats.starting_weight) if stats.starting_weight else None,
                "weight_goal": float(stats.weight_goal) if stats.weight_goal else None,
                "transformation_start_date": stats.transformation_start_date.isoformat() if stats.transformation_start_date else None,
                "next_review_date": stats.next_review_date.isoformat() if stats.next_review_date else None,
            }
        })


class WellnessTipViewSet(viewsets.ModelViewSet):
    """
    CRUD de consejos publicados por administradores.
    Usuarios autenticados pueden consultar los consejos activos,
    mientras que solo staff/admin pueden crear o modificarlos.
    """

    serializer_class = WellnessTipSerializer
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAdminUser()]
        return [IsAuthenticated()]

    def get_queryset(self):
        queryset = WellnessTip.objects.all()
        user = self.request.user
        if not user.is_staff and not user.is_superuser:
            queryset = queryset.filter(is_active=True)

        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category=category)

        audience = self.request.query_params.get("audience")
        if audience:
            queryset = queryset.filter(audience__iexact=audience)

        highlighted = self.request.query_params.get("highlighted")
        if highlighted == "true":
            queryset = queryset.filter(is_highlighted=True)

        search_query = self.request.query_params.get("q")
        if search_query:
            queryset = queryset.filter(title__icontains=search_query)

        return queryset.order_by("-is_highlighted", "-created_at")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        limit = request.query_params.get("limit")
        if limit and limit.isdigit():
            queryset = queryset[: int(limit)]
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class CoachingPlanViewSet(viewsets.ReadOnlyModelViewSet):
    """Planes 1:1 visibles para los usuarios dentro de la app."""

    serializer_class = CoachingPlanSerializer
    permission_classes = [AllowAny]
    pagination_class = None

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return CoachingPlan.objects.none()
        CoachingPlan.ensure_defaults()
        queryset = CoachingPlan.objects.filter(is_active=True)
        if self.request.user.is_authenticated and (self.request.user.is_staff or self.request.user.is_superuser):
            queryset = CoachingPlan.objects.all()
        return queryset.order_by("sort_order", "created_at")


class CoachingInquiryViewSet(viewsets.ModelViewSet):
    """Captura de solicitudes del servicio personalizado 1:1."""

    queryset = CoachingInquiry.objects.select_related("user", "plan").all()
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_serializer_class(self):
        if self.action == "create":
            return CoachingInquiryCreateSerializer
        if self.action in ["update", "partial_update"]:
            return CoachingInquiryAdminUpdateSerializer
        return CoachingInquirySerializer

    def get_permissions(self):
        if self.action == "create":
            return [AllowAny()]
        if self.action == "mine":
            return [IsAuthenticated()]
        return [IsAdminUser()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        inquiry = serializer.save()
        response_serializer = CoachingInquirySerializer(inquiry)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"])
    def mine(self, request):
        queryset = self.get_queryset().filter(user=request.user)
        serializer = CoachingInquirySerializer(queryset, many=True)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(CoachingInquirySerializer(instance).data)


class DefaultPlanConfigurationViewSet(viewsets.ModelViewSet):
    """
    ViewSet para configuraciones de planes por defecto.
    Solo accesible para administradores.
    """
    queryset = DefaultPlanConfiguration.objects.all()
    permission_classes = [IsAdminUser]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return DefaultPlanConfigurationCreateUpdateSerializer
        return DefaultPlanConfigurationSerializer
    
    def get_queryset(self):
        queryset = DefaultPlanConfiguration.objects.select_related(
            'default_nutrition_plan',
            'default_workout_program'
        ).all()
        
        # Filtros opcionales
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        main_goal = self.request.query_params.get('main_goal')
        if main_goal:
            queryset = queryset.filter(main_goal=main_goal)
        
        return queryset.order_by('priority', 'created_at')
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        
        # Devolver con el serializer de lectura
        read_serializer = DefaultPlanConfigurationSerializer(instance)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        
        # Devolver con el serializer de lectura
        read_serializer = DefaultPlanConfigurationSerializer(instance)
        return Response(read_serializer.data)
    
    @action(detail=False, methods=['get'])
    def match_user(self, request):
        """
        Endpoint para obtener la mejor configuración para un perfil de usuario dado.
        """
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response(
                {'error': 'Usuario no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Buscar configuración que coincida, ordenada por prioridad
        configurations = DefaultPlanConfiguration.objects.filter(
            is_active=True
        ).order_by('priority')
        
        for config in configurations:
            if config.matches_user_profile(user):
                serializer = DefaultPlanConfigurationSerializer(config)
                return Response(serializer.data)
        
        return Response(
            {'message': 'No se encontró configuración que coincida'}, 
            status=status.HTTP_404_NOT_FOUND
        )

    @staticmethod
    def _priority_to_order(priority):
        try:
            safe = int(priority)
        except (TypeError, ValueError):
            safe = 100
        safe = max(1, safe)
        return (safe + 9) // 10

    @staticmethod
    def _order_to_priority(order):
        try:
            normalized = int(order)
        except (TypeError, ValueError):
            normalized = 10
        normalized = max(1, normalized)
        return normalized * 10

    @staticmethod
    def _clean_text(value):
        if value is None:
            return ''
        return str(value).strip()

    @staticmethod
    def _parse_bool(value, default=True):
        if value is None or str(value).strip() == '':
            return default
        normalized = str(value).strip().lower()
        if normalized in {'true', '1', 'yes', 'sí', 'si', 's'}:
            return True
        if normalized in {'false', '0', 'no', 'n'}:
            return False
        return default

    @staticmethod
    def _parse_int_or_none(value):
        try:
            text = str(value).strip()
            return int(text) if text else None
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_list(value):
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]

        text = str(value).strip()
        if not text:
            return []

        if text.startswith('[') and text.endswith(']'):
            try:
                parsed = ast.literal_eval(text)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if str(item).strip()]
            except (ValueError, SyntaxError):
                pass

        return [item.strip() for item in re.split(r'[;,]', text) if item.strip()]

    @staticmethod
    def _normalize_main_goal(value):
        raw = str(value or '').strip().lower()
        if raw in {'', 'none', 'null', 'all', 'any', 'todos', 'todas', '*'}:
            return None

        mapping = {
            'weight_loss': 'lose_weight',
            'lose_weight': 'lose_weight',
            'perder_peso': 'lose_weight',
            'muscle_gain': 'gain_muscle',
            'gain_muscle': 'gain_muscle',
            'ganar_musculo': 'gain_muscle',
            'ganar_músculo': 'gain_muscle',
            'maintenance': 'maintain',
            'maintain': 'maintain',
            'mantenimiento': 'maintain',
            'body_recomposition': 'body_recomposition',
            'recomposicion': 'body_recomposition',
            'recomposición': 'body_recomposition',
            'performance': 'performance',
            'rendimiento': 'performance',
        }
        return mapping.get(raw, raw)

    @staticmethod
    def _normalize_training_location(value):
        raw = str(value or '').strip().lower()
        if raw in {'', 'none', 'null', 'all', 'any', 'todos', 'todas', '*'}:
            return None

        mapping = {
            'home': 'home',
            'casa': 'home',
            'gym': 'gym',
            'gimnasio': 'gym',
            'outdoor': 'outdoor',
            'exterior': 'outdoor',
            'afuera': 'outdoor',
        }
        return mapping.get(raw, raw)

    @staticmethod
    def _normalize_activity_level(value):
        raw = str(value or '').strip().lower()
        if raw in {'', 'none', 'null', 'all', 'any', 'todos', 'todas', '*'}:
            return None

        mapping = {
            'sedentary': 'sedentary',
            'sedentario': 'sedentary',
            'light': 'light',
            'ligero': 'light',
            'beginner': 'light',
            'moderate': 'moderate',
            'moderado': 'moderate',
            'intermediate': 'moderate',
            'active': 'active',
            'activo': 'active',
            'advanced': 'active',
            'very_active': 'very_active',
            'muy_activo': 'very_active',
            'muy activo': 'very_active',
        }
        return mapping.get(raw, raw)

    @staticmethod
    def _normalize_gender(value):
        raw = str(value or '').strip().lower()
        if raw in {'', 'none', 'null', 'all', 'any', 'todos', 'todas', '*'}:
            return None

        mapping = {
            'male': 'male',
            'masculino': 'male',
            'hombre': 'male',
            'female': 'female',
            'femenino': 'female',
            'mujer': 'female',
            'other': 'other',
            'otro': 'other',
        }
        return mapping.get(raw, raw)

    @staticmethod
    def _header_aliases():
        return {
            'nombre': 'nombre', 'name': 'nombre',
            'descripcion': 'descripcion', 'descripción': 'descripcion', 'description': 'descripcion',
            'prioridad': 'prioridad', 'priority': 'prioridad',
            'orden_aplicacion': 'orden_aplicacion', 'orden de aplicacion': 'orden_aplicacion',
            'application_order': 'orden_aplicacion', 'application order': 'orden_aplicacion',
            'activo': 'activo', 'is_active': 'activo',
            'objetivo_principal': 'objetivo_principal', 'main_goal': 'objetivo_principal',
            'lugar_entrenamiento': 'lugar_entrenamiento', 'training_location': 'lugar_entrenamiento',
            'nivel_actividad': 'nivel_actividad', 'activity_level': 'nivel_actividad',
            'genero': 'genero', 'género': 'genero', 'gender': 'genero',
            'dias_min_entrenamiento': 'dias_min', 'min_training_days_per_week': 'dias_min',
            'dias_max_entrenamiento': 'dias_max', 'max_training_days_per_week': 'dias_max',
            'edad_min': 'edad_min', 'age_min': 'edad_min',
            'edad_max': 'edad_max', 'age_max': 'edad_max',
            'restricciones_alimentarias': 'restricciones', 'dietary_restrictions': 'restricciones',
            'equipamiento': 'equipamiento', 'equipment_keywords': 'equipamiento',
            'plan_nutricion': 'plan_nutricion', 'default_nutrition_plan': 'plan_nutricion',
            'plan_nutricion_id': 'plan_nutricion_id', 'default_nutrition_plan_id': 'plan_nutricion_id',
            'programa_entrenamiento': 'programa_entrenamiento', 'default_workout_program': 'programa_entrenamiento',
            'programa_entrenamiento_id': 'programa_entrenamiento_id', 'default_workout_program_id': 'programa_entrenamiento_id',
        }

    @classmethod
    def _canonicalize_row(cls, row_data):
        aliases = cls._header_aliases()
        canonical = {}
        for raw_key, raw_val in row_data.items():
            normalized_key = cls._clean_text(raw_key).lower()
            mapped = aliases.get(normalized_key)
            if mapped:
                canonical[mapped] = raw_val
        return canonical

    @staticmethod
    def _resolve_nutrition_plan(plan_id, plan_name):
        from nutrition.models import NutritionPlan

        clean_id = str(plan_id or '').strip()
        if clean_id:
            by_id = NutritionPlan.objects.filter(id=clean_id).first()
            if by_id:
                return by_id

        clean_name = str(plan_name or '').strip()
        if clean_name:
            return NutritionPlan.objects.filter(name__iexact=clean_name).first()

        return None

    @staticmethod
    def _resolve_workout_program(program_id, program_name):
        from workouts.models import WorkoutProgram

        clean_id = str(program_id or '').strip()
        if clean_id:
            by_id = WorkoutProgram.objects.filter(id=clean_id).first()
            if by_id:
                return by_id

        clean_name = str(program_name or '').strip()
        if clean_name:
            return WorkoutProgram.objects.filter(name__iexact=clean_name).first()

        return None

    def _build_fields_from_row(self, row):
        name = self._clean_text(row.get('nombre'))
        if not name:
            raise ValueError("'nombre' es requerido")

        order_value = self._parse_int_or_none(row.get('orden_aplicacion'))
        priority_value = self._parse_int_or_none(row.get('prioridad'))
        priority = self._order_to_priority(order_value) if order_value is not None else (priority_value if priority_value is not None else 100)

        nutrition_plan_name = self._clean_text(row.get('plan_nutricion'))
        nutrition_plan_id = self._clean_text(row.get('plan_nutricion_id'))
        workout_program_name = self._clean_text(row.get('programa_entrenamiento'))
        workout_program_id = self._clean_text(row.get('programa_entrenamiento_id'))

        nutrition_plan = self._resolve_nutrition_plan(nutrition_plan_id, nutrition_plan_name)
        if (nutrition_plan_id or nutrition_plan_name) and not nutrition_plan:
            raise ValueError(f"Plan nutricional no encontrado: '{nutrition_plan_id or nutrition_plan_name}'")

        workout_program = self._resolve_workout_program(workout_program_id, workout_program_name)
        if (workout_program_id or workout_program_name) and not workout_program:
            raise ValueError(f"Programa de entrenamiento no encontrado: '{workout_program_id or workout_program_name}'")

        fields = {
            'description': self._clean_text(row.get('descripcion')),
            'priority': priority,
            'is_active': self._parse_bool(row.get('activo'), default=True),
            'main_goal': self._normalize_main_goal(row.get('objetivo_principal')),
            'training_location': self._normalize_training_location(row.get('lugar_entrenamiento')),
            'activity_level': self._normalize_activity_level(row.get('nivel_actividad')),
            'gender': self._normalize_gender(row.get('genero')),
            'min_training_days_per_week': self._parse_int_or_none(row.get('dias_min')),
            'max_training_days_per_week': self._parse_int_or_none(row.get('dias_max')),
            'age_min': self._parse_int_or_none(row.get('edad_min')),
            'age_max': self._parse_int_or_none(row.get('edad_max')),
            'dietary_restrictions': self._parse_list(row.get('restricciones')),
            'equipment_keywords': self._parse_list(row.get('equipamiento')),
            'default_nutrition_plan': nutrition_plan,
            'default_workout_program': workout_program,
        }
        return name, fields

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """Exporta todas las configuraciones en formato CSV (alineado con la lógica actual)."""
        import csv
        from django.http import HttpResponse

        configs = self.get_queryset()
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="default_plan_configurations.csv"'

        fieldnames = [
            'nombre', 'descripcion', 'prioridad', 'orden_aplicacion', 'activo',
            'objetivo_principal', 'lugar_entrenamiento', 'nivel_actividad', 'genero',
            'dias_min_entrenamiento', 'dias_max_entrenamiento',
            'edad_min', 'edad_max',
            'restricciones_alimentarias', 'equipamiento',
            'plan_nutricion', 'plan_nutricion_id', 'programa_entrenamiento', 'programa_entrenamiento_id',
        ]
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()

        for config in configs:
            writer.writerow({
                'nombre': config.name,
                'descripcion': config.description or '',
                'prioridad': config.priority,
                'orden_aplicacion': self._priority_to_order(config.priority),
                'activo': 'Sí' if config.is_active else 'No',
                'objetivo_principal': config.main_goal or '',
                'lugar_entrenamiento': config.training_location or '',
                'nivel_actividad': config.activity_level or '',
                'genero': config.gender or '',
                'dias_min_entrenamiento': config.min_training_days_per_week if config.min_training_days_per_week is not None else '',
                'dias_max_entrenamiento': config.max_training_days_per_week if config.max_training_days_per_week is not None else '',
                'edad_min': config.age_min if config.age_min is not None else '',
                'edad_max': config.age_max if config.age_max is not None else '',
                'restricciones_alimentarias': ','.join(config.dietary_restrictions or []),
                'equipamiento': ','.join(config.equipment_keywords or []),
                'plan_nutricion': config.default_nutrition_plan.name if config.default_nutrition_plan else '',
                'plan_nutricion_id': str(config.default_nutrition_plan.id) if config.default_nutrition_plan else '',
                'programa_entrenamiento': config.default_workout_program.name if config.default_workout_program else '',
                'programa_entrenamiento_id': str(config.default_workout_program.id) if config.default_workout_program else '',
            })
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporta configuraciones en Excel con datos y referencias actualizadas."""
        import io
        import xlsxwriter
        from django.http import HttpResponse
        from nutrition.models import NutritionPlan
        from workouts.models import WorkoutProgram

        configs = self.get_queryset()
        nutrition_plans = NutritionPlan.objects.filter(is_active=True).order_by('name')
        workout_programs = WorkoutProgram.objects.filter(is_active=True).order_by('name')
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # ========== HOJA 1: DATOS ==========
        worksheet = workbook.add_worksheet('Configuraciones')
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})

        headers = [
            'nombre', 'descripcion', 'prioridad', 'orden_aplicacion', 'activo',
            'objetivo_principal', 'lugar_entrenamiento', 'nivel_actividad', 'genero',
            'dias_min_entrenamiento', 'dias_max_entrenamiento',
            'edad_min', 'edad_max',
            'restricciones_alimentarias', 'equipamiento',
            'plan_nutricion', 'plan_nutricion_id', 'programa_entrenamiento', 'programa_entrenamiento_id',
        ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        for row_idx, config in enumerate(configs, start=1):
            worksheet.write(row_idx, 0, config.name)
            worksheet.write(row_idx, 1, config.description or '')
            worksheet.write(row_idx, 2, config.priority)
            worksheet.write(row_idx, 3, self._priority_to_order(config.priority))
            worksheet.write(row_idx, 4, 'Sí' if config.is_active else 'No')
            worksheet.write(row_idx, 5, config.main_goal or '')
            worksheet.write(row_idx, 6, config.training_location or '')
            worksheet.write(row_idx, 7, config.activity_level or '')
            worksheet.write(row_idx, 8, config.gender or '')
            worksheet.write(row_idx, 9, config.min_training_days_per_week if config.min_training_days_per_week is not None else '')
            worksheet.write(row_idx, 10, config.max_training_days_per_week if config.max_training_days_per_week is not None else '')
            worksheet.write(row_idx, 11, config.age_min if config.age_min is not None else '')
            worksheet.write(row_idx, 12, config.age_max if config.age_max is not None else '')
            worksheet.write(row_idx, 13, ','.join(config.dietary_restrictions or []))
            worksheet.write(row_idx, 14, ','.join(config.equipment_keywords or []))
            worksheet.write(row_idx, 15, config.default_nutrition_plan.name if config.default_nutrition_plan else '')
            worksheet.write(row_idx, 16, str(config.default_nutrition_plan.id) if config.default_nutrition_plan else '')
            worksheet.write(row_idx, 17, config.default_workout_program.name if config.default_workout_program else '')
            worksheet.write(row_idx, 18, str(config.default_workout_program.id) if config.default_workout_program else '')

        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 40)
        worksheet.set_column('C:S', 22)

        # ========== HOJA 2: REFERENCIAS ==========
        ref_ws = workbook.add_worksheet('Referencias')
        ref_ws.write(0, 0, 'OBJETIVO_PRINCIPAL', header_format)
        for i, val in enumerate(['lose_weight', 'gain_muscle', 'body_recomposition', 'maintain', 'performance'], 1):
            ref_ws.write(i, 0, val)

        ref_ws.write(0, 2, 'LUGAR_ENTRENAMIENTO', header_format)
        for i, val in enumerate(['home', 'gym', 'outdoor', '(vacío = genérico)'], 1):
            ref_ws.write(i, 2, val)

        ref_ws.write(0, 4, 'NIVEL_ACTIVIDAD', header_format)
        for i, val in enumerate(['sedentary', 'light', 'moderate', 'active', 'very_active', '(vacío = genérico)'], 1):
            ref_ws.write(i, 4, val)

        ref_ws.write(0, 6, 'GENERO', header_format)
        for i, val in enumerate(['male', 'female', 'other', '(vacío = todos)'], 1):
            ref_ws.write(i, 6, val)

        ref_ws.write(0, 8, 'NUTRITION_PLAN_ID', header_format)
        ref_ws.write(0, 9, 'NUTRITION_PLAN_NOMBRE', header_format)
        for i, plan in enumerate(nutrition_plans, 1):
            ref_ws.write(i, 8, str(plan.id))
            ref_ws.write(i, 9, plan.name)

        ref_ws.write(0, 11, 'WORKOUT_PROGRAM_ID', header_format)
        ref_ws.write(0, 12, 'WORKOUT_PROGRAM_NOMBRE', header_format)
        for i, program in enumerate(workout_programs, 1):
            ref_ws.write(i, 11, str(program.id))
            ref_ws.write(i, 12, program.name)

        ref_ws.set_column('A:G', 24)
        ref_ws.set_column('H:H', 10)
        ref_ws.set_column('I:I', 40)
        ref_ws.set_column('J:J', 40)
        ref_ws.set_column('K:K', 10)
        ref_ws.set_column('L:L', 40)
        ref_ws.set_column('M:M', 40)

        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="default_plan_configurations.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import-csv', parser_classes=[MultiPartParser, FormParser])
    def import_csv(self, request):
        """Importa configuraciones desde CSV usando formato canónico y valores normalizados."""
        import csv
        from django.core.files.uploadedfile import UploadedFile

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo CSV no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = file.read().decode('utf-8')
        except UnicodeDecodeError:
            return Response({'error': 'El archivo debe estar en formato UTF-8'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(decoded.splitlines())
        created, updated, skipped = 0, 0, 0
        errors = []

        for row_num, row in enumerate(reader, start=2):
            try:
                canonical_row = self._canonicalize_row(row)
                name, fields = self._build_fields_from_row(canonical_row)

                config = DefaultPlanConfiguration.objects.filter(name=name).first()
                if config:
                    for k, v in fields.items():
                        setattr(config, k, v)
                    config.save()
                    updated += 1
                else:
                    DefaultPlanConfiguration.objects.create(name=name, **fields)
                    created += 1
            except Exception as e:
                errors.append(f"Fila {row_num}: {str(e)}")
                skipped += 1

        message = f"Importación completada. {created} creadas, {updated} actualizadas"
        if skipped:
            message += f", {skipped} omitidas"
        if errors:
            message += f". {len(errors)} error(es)"

        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'message': message,
            'errors': errors[:10] if errors else [],
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """Importa configuraciones desde Excel usando formato canónico y valores normalizados."""
        import openpyxl
        from django.core.files.uploadedfile import UploadedFile

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo Excel no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Error al leer el archivo Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        aliases = self._header_aliases()
        raw_headers = [self._clean_text(cell.value).lower() if cell.value is not None else '' for cell in ws[1]]
        canonical_headers = [aliases.get(header, header) for header in raw_headers]
        created, updated, skipped = 0, 0, 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_dict = dict(zip(canonical_headers, row))
                name, fields = self._build_fields_from_row(row_dict)

                config = DefaultPlanConfiguration.objects.filter(name=name).first()
                if config:
                    for k, v in fields.items():
                        setattr(config, k, v)
                    config.save()
                    updated += 1
                else:
                    DefaultPlanConfiguration.objects.create(name=name, **fields)
                    created += 1
            except Exception as e:
                errors.append(f"Fila {row_num}: {str(e)}")
                skipped += 1

        message = f"Importación completada. {created} creadas, {updated} actualizadas"
        if skipped:
            message += f", {skipped} omitidas"
        if errors:
            message += f". {len(errors)} error(es)"

        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'message': message,
            'errors': errors[:10] if errors else [],
        }, status=status.HTTP_200_OK)


class HelpSettingsViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar configuración de ayuda
    """
    serializer_class = HelpSettingsSerializer
    permission_classes = [IsAdminUser]
    queryset = HelpSettings.objects.all()
    
    def get_permissions(self):
        # El endpoint 'active' es público
        if self.action == 'active':
            return []
        return [IsAdminUser()]
    
    @action(detail=False, methods=['get'], permission_classes=[], url_path='active', url_name='active')
    def active(self, request):
        """Obtener configuración activa (público)"""
        try:
            settings = HelpSettings.get_active()
            serializer = self.get_serializer(settings)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': f'Error obteniendo configuración: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProblemReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar reportes de problemas
    """
    queryset = ProblemReport.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ProblemReportCreateSerializer
        return ProblemReportSerializer
    
    def get_permissions(self):
        if self.action == 'create':
            # Permitir crear reportes sin autenticación
            return []
        # Solo admin puede ver/editar reportes
        return [IsAdminUser()]
    
    def perform_create(self, serializer):
        # Asignar usuario si está autenticado
        if self.request.user.is_authenticated:
            serializer.save(user=self.request.user)
        else:
            serializer.save()
        
        # Enviar email de notificación (opcional)
        instance = serializer.instance
        help_settings = HelpSettings.get_active()
        if help_settings.report_enabled and help_settings.report_email:
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                subject = f"[NexFit365] Nuevo Reporte: {instance.subject}"
                message = f"""
Se ha recibido un nuevo reporte de problema:

Tipo: {instance.get_problem_type_display()}
Asunto: {instance.subject}
Descripción: {instance.description}

Usuario: {instance.user.email if instance.user else instance.contact_email or 'Anónimo'}
Estado: {instance.get_status_display()}

Ver detalles: {settings.FRONTEND_URL}/admin/problem-reports/{instance.id}/
"""
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [help_settings.report_email],
                    fail_silently=True,
                )
            except Exception as e:
                # No fallar si el email no se puede enviar
                pass
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def resolve(self, request, pk=None):
        """Marcar un reporte como resuelto"""
        report = self.get_object()
        report.status = ProblemReport.Status.RESOLVED
        report.resolved_by = request.user
        report.resolved_at = timezone.now()
        report.save()
        
        serializer = self.get_serializer(report)
        return Response(serializer.data)
