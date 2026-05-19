# workouts/admin_exercise_views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAdminUser
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import DatabaseError
from django.db.models import Count, Q
from django.utils import timezone
from django.shortcuts import get_object_or_404
from datetime import timedelta
from django.db import IntegrityError

from .models import Exercise, ExerciseSubstitution
from .serializers import ExerciseSerializer
from accounts.permissions import IsAdminOrStaff


class LargeResultsSetPagination(PageNumberPagination):
    """Paginación personalizada para listados grandes"""
    page_size = 1000  # Devolver hasta 1000 elementos por defecto
    page_size_query_param = 'page_size'
    max_page_size = 10000


class AdminExerciseViewSet(viewsets.ModelViewSet):
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @action(detail=False, methods=['post'], url_path='import-csv')
    def import_csv(self, request):
                """Importa ejercicios desde un archivo CSV. Solo añade o modifica, nunca elimina.
                
                Formato esperado:
                - name (requerido): Nombre del ejercicio
                - description: Descripción detallada
                - instructions: Instrucciones paso a paso
                - category: Categoría (cardio, strength, flexibility, etc.)
                - muscle_groups: Grupos musculares separados por comas
                - equipment: Equipamiento separado por comas
                - difficulty: beginner, intermediate, advanced
                - video_url: URL del video
                - google_drive_file_id: ID del archivo en Google Drive
                - is_system: True/False (ejercicios del sistema)
                - is_active: True/False (ejercicio activo)
                - tags: Tags separados por comas
                """
                import csv
                from django.core.files.uploadedfile import UploadedFile
                
                file = request.FILES.get('file')
                if not file or not isinstance(file, UploadedFile):
                    return Response({'error': 'Archivo CSV no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    decoded = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    return Response({'error': 'El archivo debe estar en formato UTF-8'}, status=status.HTTP_400_BAD_REQUEST)
                
                reader = csv.DictReader(decoded.splitlines())
                updated, created, skipped = 0, 0, 0
                errors = []

                header_aliases = {
                    'name': 'name',
                    'nombre': 'name',
                    'description': 'description',
                    'descripción': 'description',
                    'descripcion': 'description',
                    'instructions': 'instructions',
                    'instrucciones': 'instructions',
                    'category': 'category',
                    'categoría': 'category',
                    'categoria': 'category',
                    'muscle_groups': 'muscle_groups',
                    'grupos_musculares': 'muscle_groups',
                    'equipment': 'equipment',
                    'equipamiento': 'equipment',
                    'difficulty': 'difficulty',
                    'dificultad': 'difficulty',
                    'video_url': 'video_url',
                    'url_video': 'video_url',
                    'google_drive_file_id': 'google_drive_file_id',
                    'google_drive_id_archivo': 'google_drive_file_id',
                    'is_system': 'is_system',
                    'es_del_sistema': 'is_system',
                    'is_active': 'is_active',
                    'está_activo': 'is_active',
                    'esta_activo': 'is_active',
                    'tags': 'tags',
                    'etiquetas': 'tags',
                }
                category_es_to_code = {
                    'cardio': 'cardio',
                    'fuerza': 'strength',
                    'flexibilidad': 'flexibility',
                    'peso corporal': 'bodyweight',
                    'peso_corporal': 'bodyweight',
                    'hiit': 'hiit',
                }
                difficulty_es_to_code = {
                    'principiante': 'beginner',
                    'intermedio': 'intermediate',
                    'avanzado': 'advanced',
                }
                muscle_group_es_to_code = {
                    'cardio': 'cardio',
                    'pecho': 'chest',
                    'espalda': 'back',
                    'hombros': 'shoulders',
                    'bíceps': 'biceps',
                    'biceps': 'biceps',
                    'tríceps': 'triceps',
                    'triceps': 'triceps',
                    'antebrazos': 'forearms',
                    'abdominales': 'abs',
                    'core': 'core',
                    'oblicuos': 'obliques',
                    'cuádriceps': 'quads',
                    'cuadriceps': 'quads',
                    'isquiotibiales': 'hamstrings',
                    'glúteos': 'glutes',
                    'gluteos': 'glutes',
                    'pantorrillas': 'calves',
                }
                equipment_es_to_code = {
                    'barra': 'barbell',
                    'mancuernas': 'dumbbells',
                    'kettlebell': 'kettlebell',
                    'bandas de resistencia': 'resistance bands',
                    'peso corporal': 'bodyweight',
                    'máquina': 'machine',
                    'maquina': 'machine',
                    'cable': 'cable',
                    'banco': 'bench',
                    'barra de dominadas': 'pull-up bar',
                    'cinta de correr': 'treadmill',
                }

                def normalize_code(raw_value, translation_map):
                    value = str(raw_value).strip().lower() if raw_value is not None else ''
                    if not value:
                        return ''
                    return translation_map.get(value, value)

                def to_bool(raw_value, default=False):
                    if raw_value is None or str(raw_value).strip() == '':
                        return default
                    return str(raw_value).strip().lower() in ['true', '1', 'yes', 'sí', 'si', 's']

                def get_value(row_data, canonical_key, default=''):
                    for raw_key, raw_value in row_data.items():
                        mapped_key = header_aliases.get(str(raw_key).strip().lower())
                        if mapped_key == canonical_key:
                            return raw_value
                    return default
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Ignorar filas completamente vacias sin contarlas como omitidas.
                        if all(value is None or str(value).strip() == '' for value in row.values()):
                            continue

                        name = Exercise.normalize_name(str(get_value(row, 'name', '') or ''))
                        if not name:
                            errors.append(f"Fila {row_num}: 'name' es requerido")
                            skipped += 1
                            continue

                        exercise = Exercise.find_existing_by_name(name)

                        raw_muscle_groups = [m.strip() for m in str(get_value(row, 'muscle_groups', '') or '').split(',') if m.strip()]
                        raw_equipment = [e.strip() for e in str(get_value(row, 'equipment', '') or '').split(',') if e.strip()]
                        raw_tags = [t.strip() for t in str(get_value(row, 'tags', '') or '').split(',') if t.strip()]

                        fields = {
                            'description': str(get_value(row, 'description', '') or '').strip(),
                            'instructions': str(get_value(row, 'instructions', '') or '').strip(),
                            'category': normalize_code(str(get_value(row, 'category', '') or '').strip(), category_es_to_code),
                            'muscle_groups': [normalize_code(m, muscle_group_es_to_code) for m in raw_muscle_groups],
                            'equipment': [normalize_code(e, equipment_es_to_code) for e in raw_equipment],
                            'difficulty': normalize_code(str(get_value(row, 'difficulty', '') or '').strip(), difficulty_es_to_code),
                            'video_url': str(get_value(row, 'video_url', '') or '').strip(),
                            'google_drive_file_id': str(get_value(row, 'google_drive_file_id', '') or '').strip(),
                            'is_system': to_bool(get_value(row, 'is_system', ''), default=False),
                            'is_active': to_bool(get_value(row, 'is_active', ''), default=True),
                            'tags': raw_tags,
                        }
                        
                        if exercise:
                            for k, v in fields.items():
                                setattr(exercise, k, v)
                            exercise.save()
                            updated += 1
                        else:
                            try:
                                Exercise.objects.create(name=name, **fields)
                                created += 1
                            except IntegrityError:
                                # En carreras concurrentes, actualizar el registro existente.
                                existing = Exercise.find_existing_by_name(name)
                                if existing:
                                    for k, v in fields.items():
                                        setattr(existing, k, v)
                                    existing.save()
                                    updated += 1
                                else:
                                    raise
                    except Exception as e:
                        errors.append(f"Fila {row_num}: {str(e)}")
                        skipped += 1
                
                message = f"Importación completada. {created} ejercicios creados, {updated} actualizados"
                if skipped > 0:
                    message += f", {skipped} omitidos"
                if errors:
                    message += f". {len(errors)} error(es) encontrado(s)"
                
                return Response({
                    'created': created,
                    'updated': updated,
                    'skipped': skipped,
                    'message': message,
                    'errors': errors[:10] if errors else []
                }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
                """Importa ejercicios desde un archivo Excel. Solo añade o modifica, nunca elimina.
                
                Lee la primera hoja del archivo Excel. La segunda hoja (Referencias) se ignora.
                
                Formato esperado (igual que CSV):
                - name (requerido): Nombre del ejercicio
                - description: Descripción detallada
                - instructions: Instrucciones paso a paso
                - category: Categoría
                - muscle_groups: Grupos musculares separados por comas
                - equipment: Equipamiento separado por comas
                - difficulty: beginner, intermediate, advanced
                - video_url: URL del video
                - google_drive_file_id: ID del archivo en Google Drive
                - is_system: True/False
                - is_active: True/False
                - tags: Tags separados por comas
                """
                import openpyxl
                from django.core.files.uploadedfile import UploadedFile
                
                file = request.FILES.get('file')
                if not file or not isinstance(file, UploadedFile):
                    return Response({'error': 'Archivo Excel no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                except Exception as e:
                    return Response({'error': f'Error al leer el archivo Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
                
                headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
                updated, created, skipped = 0, 0, 0
                errors = []

                header_aliases = {
                    'name': 'name',
                    'nombre': 'name',
                    'description': 'description',
                    'descripción': 'description',
                    'descripcion': 'description',
                    'instructions': 'instructions',
                    'instrucciones': 'instructions',
                    'category': 'category',
                    'categoría': 'category',
                    'categoria': 'category',
                    'muscle_groups': 'muscle_groups',
                    'grupos_musculares': 'muscle_groups',
                    'equipment': 'equipment',
                    'equipamiento': 'equipment',
                    'difficulty': 'difficulty',
                    'dificultad': 'difficulty',
                    'video_url': 'video_url',
                    'url_video': 'video_url',
                    'google_drive_file_id': 'google_drive_file_id',
                    'google_drive_id_archivo': 'google_drive_file_id',
                    'is_system': 'is_system',
                    'es_del_sistema': 'is_system',
                    'is_active': 'is_active',
                    'está_activo': 'is_active',
                    'esta_activo': 'is_active',
                    'tags': 'tags',
                    'etiquetas': 'tags',
                }

                category_es_to_code = {
                    'cardio': 'cardio',
                    'fuerza': 'strength',
                    'flexibilidad': 'flexibility',
                    'peso corporal': 'bodyweight',
                    'peso_corporal': 'bodyweight',
                    'hiit': 'hiit',
                }
                difficulty_es_to_code = {
                    'principiante': 'beginner',
                    'intermedio': 'intermediate',
                    'avanzado': 'advanced',
                }
                muscle_group_es_to_code = {
                    'cardio': 'cardio',
                    'pecho': 'chest',
                    'espalda': 'back',
                    'hombros': 'shoulders',
                    'bíceps': 'biceps',
                    'biceps': 'biceps',
                    'tríceps': 'triceps',
                    'triceps': 'triceps',
                    'antebrazos': 'forearms',
                    'abdominales': 'abs',
                    'core': 'core',
                    'oblicuos': 'obliques',
                    'cuádriceps': 'quads',
                    'cuadriceps': 'quads',
                    'isquiotibiales': 'hamstrings',
                    'glúteos': 'glutes',
                    'gluteos': 'glutes',
                    'pantorrillas': 'calves',
                }
                equipment_es_to_code = {
                    'barra': 'barbell',
                    'mancuernas': 'dumbbells',
                    'kettlebell': 'kettlebell',
                    'bandas de resistencia': 'resistance bands',
                    'peso corporal': 'bodyweight',
                    'máquina': 'machine',
                    'maquina': 'machine',
                    'cable': 'cable',
                    'banco': 'bench',
                    'barra de dominadas': 'pull-up bar',
                    'cinta de correr': 'treadmill',
                }

                def normalize_code(raw_value, translation_map):
                    value = str(raw_value).strip().lower() if raw_value is not None else ''
                    if not value:
                        return ''
                    return translation_map.get(value, value)

                # Auto-link: construir mapa de videos de Google Drive (nombre normalizado -> file_id)
                import unicodedata as _unicodedata
                import re as _re

                def _normalize_for_drive_match(name):
                    """Normaliza un nombre para comparar con nombres de videos en Drive"""
                    nfkd = _unicodedata.normalize('NFKD', (name or '').lower())
                    ascii_name = ''.join(c for c in nfkd if not _unicodedata.combining(c))
                    return _re.sub(r'[^a-z0-9]', '', ascii_name)

                drive_video_map = {}
                try:
                    from .google_drive_service import get_google_drive_service
                    _drive_service = get_google_drive_service()
                    if _drive_service.is_configured():
                        _drive_videos = _drive_service.list_videos_from_folder()
                        for _v in _drive_videos:
                            _key = _normalize_for_drive_match(_v['name'])
                            if _key and _key not in drive_video_map:
                                drive_video_map[_key] = _v['file_id']
                except Exception:
                    pass  # Drive no configurado o error, omitir auto-vinculación

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Ignorar filas completamente vacias sin contarlas como omitidas.
                        if all(value is None or str(value).strip() == '' for value in row):
                            continue

                        original_row_dict = dict(zip(headers, row))
                        row_dict = {}
                        for raw_key, raw_val in original_row_dict.items():
                            normalized_key = header_aliases.get(str(raw_key).strip().lower(), str(raw_key).strip())
                            row_dict[normalized_key] = raw_val

                        name = Exercise.normalize_name(str(row_dict.get('name', '')) if row_dict.get('name') else '')
                        
                        if not name:
                            errors.append(f"Fila {row_num}: 'name' es requerido")
                            skipped += 1
                            continue
                        
                        exercise = Exercise.find_existing_by_name(name)
                        
                        # Función auxiliar para convertir valores a string y limpiar
                        def clean_str(val):
                            return str(val).strip() if val is not None else ''
                        
                        def to_bool(val):
                            if val is None or val == '':
                                return True  # Por defecto is_active=True
                            return str(val).strip().lower() in ['true', '1', 'yes', 'sí', 'si', 's']

                        raw_muscle_groups = [m.strip() for m in clean_str(row_dict.get('muscle_groups', '')).split(',') if m.strip()]
                        raw_equipment = [e.strip() for e in clean_str(row_dict.get('equipment', '')).split(',') if e.strip()]
                        
                        fields = {
                            'description': clean_str(row_dict.get('description', '')),
                            'instructions': clean_str(row_dict.get('instructions', '')),
                            'category': normalize_code(clean_str(row_dict.get('category', '')), category_es_to_code),
                            'muscle_groups': [normalize_code(m, muscle_group_es_to_code) for m in raw_muscle_groups],
                            'equipment': [normalize_code(e, equipment_es_to_code) for e in raw_equipment],
                            'difficulty': normalize_code(clean_str(row_dict.get('difficulty', '')), difficulty_es_to_code),
                            'video_url': clean_str(row_dict.get('video_url', '')),
                            'google_drive_file_id': clean_str(row_dict.get('google_drive_file_id', '')),
                            'is_system': to_bool(row_dict.get('is_system', False)),
                            'is_active': to_bool(row_dict.get('is_active', True)),
                            'tags': [t.strip() for t in clean_str(row_dict.get('tags', '')).split(',') if t.strip()],
                        }

                        # Auto-link: si no hay google_drive_file_id en el Excel, intentar vincular por nombre
                        if not fields['google_drive_file_id'] and drive_video_map:
                            _matched_id = drive_video_map.get(_normalize_for_drive_match(name), '')
                            if _matched_id:
                                fields['google_drive_file_id'] = _matched_id
                                if not fields['video_url']:
                                    fields['video_url'] = f"https://drive.google.com/file/d/{_matched_id}/preview"

                        if exercise:
                            for k, v in fields.items():
                                setattr(exercise, k, v)
                            exercise.save()
                            updated += 1
                        else:
                            try:
                                Exercise.objects.create(name=name, **fields)
                                created += 1
                            except IntegrityError:
                                existing = Exercise.find_existing_by_name(name)
                                if existing:
                                    for k, v in fields.items():
                                        setattr(existing, k, v)
                                    existing.save()
                                    updated += 1
                                else:
                                    raise
                    except Exception as e:
                        errors.append(f"Fila {row_num}: {str(e)}")
                        skipped += 1
                
                message = f"Importación completada. {created} ejercicios creados, {updated} actualizados"
                if skipped > 0:
                    message += f", {skipped} omitidos"
                if errors:
                    message += f". {len(errors)} error(es) encontrado(s)"
                
                return Response({
                    'created': created,
                    'updated': updated,
                    'skipped': skipped,
                    'message': message,
                    'errors': errors[:10] if errors else []
                }, status=status.HTTP_200_OK)
    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
            """Exporta todos los ejercicios en formato CSV"""
            import csv
            from django.http import HttpResponse
        
            exercises = self.get_queryset()
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="exercises_export.csv"'
        
            fieldnames = [
                'nombre', 'descripción', 'instrucciones', 'categoría', 'grupos_musculares',
                'equipamiento', 'dificultad', 'url_video', 'google_drive_id_archivo',
                'es_del_sistema', 'está_activo', 'etiquetas'
            ]
            writer = csv.DictWriter(response, fieldnames=fieldnames)
            writer.writeheader()

            category_code_to_es = {
                'cardio': 'Cardio',
                'strength': 'Fuerza',
                'flexibility': 'Flexibilidad',
                'bodyweight': 'Peso corporal',
                'hiit': 'HIIT',
            }
            difficulty_code_to_es = {
                'beginner': 'Principiante',
                'intermediate': 'Intermedio',
                'advanced': 'Avanzado',
            }
            muscle_group_code_to_es = {
                'cardio': 'Cardio',
                'chest': 'Pecho',
                'back': 'Espalda',
                'shoulders': 'Hombros',
                'biceps': 'Bíceps',
                'triceps': 'Tríceps',
                'forearms': 'Antebrazos',
                'abs': 'Abdominales',
                'core': 'Core',
                'obliques': 'Oblicuos',
                'quads': 'Cuádriceps',
                'hamstrings': 'Isquiotibiales',
                'glutes': 'Glúteos',
                'calves': 'Pantorrillas',
            }
            equipment_code_to_es = {
                'barbell': 'Barra',
                'dumbbells': 'Mancuernas',
                'kettlebell': 'Kettlebell',
                'resistance bands': 'Bandas de resistencia',
                'bodyweight': 'Peso corporal',
                'machine': 'Máquina',
                'cable': 'Cable',
                'bench': 'Banco',
                'pull-up bar': 'Barra de dominadas',
                'treadmill': 'Cinta de correr',
            }

            def to_spanish_label(value, translation_map):
                key = (value or '').strip().lower()
                return translation_map.get(key, value or '')

            for exercise in exercises:
                writer.writerow({
                    'nombre': exercise.name,
                    'descripción': exercise.description or '',
                    'instrucciones': exercise.instructions or '',
                    'categoría': to_spanish_label(exercise.category, category_code_to_es),
                    'grupos_musculares': ','.join([to_spanish_label(mg, muscle_group_code_to_es) for mg in (exercise.muscle_groups or [])]),
                    'equipamiento': ','.join([to_spanish_label(eq, equipment_code_to_es) for eq in (exercise.equipment or [])]),
                    'dificultad': to_spanish_label(exercise.difficulty, difficulty_code_to_es),
                    'url_video': exercise.video_url or '',
                    'google_drive_id_archivo': exercise.google_drive_file_id or '',
                    'es_del_sistema': 'Sí' if (exercise.is_system if hasattr(exercise, 'is_system') else False) else 'No',
                    'está_activo': 'Sí' if (exercise.is_active if hasattr(exercise, 'is_active') else True) else 'No',
                    'etiquetas': ','.join(exercise.tags or []),
                })
            return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
            """Exporta ejercicios en Excel con dos hojas:
            1. Ejercicios: Todos los ejercicios con sus detalles
            2. Referencias: Categorías, grupos musculares, equipamiento y dificultades disponibles
            """
            import io
            import xlsxwriter
            from django.http import HttpResponse
            from collections import defaultdict
        
            exercises = self.get_queryset()
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            
            # ========== HOJA 1: EJERCICIOS ==========
            worksheet = workbook.add_worksheet('Ejercicios')
        
            headers = [
                'nombre', 'descripción', 'instrucciones', 'categoría', 'grupos_musculares',
                'equipamiento', 'dificultad', 'url_video', 'google_drive_id_archivo',
                'es_del_sistema', 'está_activo', 'etiquetas'
            ]

            category_code_to_es = {
                'cardio': 'Cardio',
                'strength': 'Fuerza',
                'flexibility': 'Flexibilidad',
                'bodyweight': 'Peso corporal',
                'hiit': 'HIIT',
            }
            difficulty_code_to_es = {
                'beginner': 'Principiante',
                'intermediate': 'Intermedio',
                'advanced': 'Avanzado',
            }
            muscle_group_code_to_es = {
                'cardio': 'Cardio',
                'chest': 'Pecho',
                'back': 'Espalda',
                'shoulders': 'Hombros',
                'biceps': 'Bíceps',
                'triceps': 'Tríceps',
                'forearms': 'Antebrazos',
                'abs': 'Abdominales',
                'core': 'Core',
                'obliques': 'Oblicuos',
                'quads': 'Cuádriceps',
                'hamstrings': 'Isquiotibiales',
                'glutes': 'Glúteos',
                'calves': 'Pantorrillas',
            }
            equipment_code_to_es = {
                'barbell': 'Barra',
                'dumbbells': 'Mancuernas',
                'kettlebell': 'Kettlebell',
                'resistance bands': 'Bandas de resistencia',
                'bodyweight': 'Peso corporal',
                'machine': 'Máquina',
                'cable': 'Cable',
                'bench': 'Banco',
                'pull-up bar': 'Barra de dominadas',
                'treadmill': 'Cinta de correr',
            }

            def to_spanish_label(value, translation_map):
                key = (value or '').strip().lower()
                return translation_map.get(key, value or '')
            
            # Formato para headers
            header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            # Recopilar datos para la hoja de referencias
            all_categories = set()
            all_muscle_groups = set()
            all_equipment = set()
            all_difficulties = set()
            
            for row_idx, exercise in enumerate(exercises, start=1):
                # Recopilar valores únicos para referencias
                if exercise.category:
                    all_categories.add(to_spanish_label(exercise.category, category_code_to_es))
                if exercise.difficulty:
                    all_difficulties.add(to_spanish_label(exercise.difficulty, difficulty_code_to_es))
                for mg in (exercise.muscle_groups or []):
                    all_muscle_groups.add(to_spanish_label(mg, muscle_group_code_to_es))
                for eq in (exercise.equipment or []):
                    all_equipment.add(to_spanish_label(eq, equipment_code_to_es))
                
                # Escribir datos del ejercicio
                worksheet.write(row_idx, 0, exercise.name)
                worksheet.write(row_idx, 1, exercise.description or '')
                worksheet.write(row_idx, 2, exercise.instructions or '')
                worksheet.write(row_idx, 3, to_spanish_label(exercise.category, category_code_to_es))
                worksheet.write(row_idx, 4, ','.join([to_spanish_label(mg, muscle_group_code_to_es) for mg in (exercise.muscle_groups or [])]))
                worksheet.write(row_idx, 5, ','.join([to_spanish_label(eq, equipment_code_to_es) for eq in (exercise.equipment or [])]))
                worksheet.write(row_idx, 6, to_spanish_label(exercise.difficulty, difficulty_code_to_es))
                worksheet.write(row_idx, 7, exercise.video_url or '')
                worksheet.write(row_idx, 8, exercise.google_drive_file_id or '')
                worksheet.write(row_idx, 9, 'Sí' if getattr(exercise, 'is_system', False) else 'No')
                worksheet.write(row_idx, 10, 'Sí' if getattr(exercise, 'is_active', True) else 'No')
                worksheet.write(row_idx, 11, ','.join(exercise.tags or []))
            
            # Ajustar ancho de columnas
            worksheet.set_column('A:A', 30)  # name
            worksheet.set_column('B:B', 40)  # description
            worksheet.set_column('C:C', 50)  # instructions
            worksheet.set_column('D:G', 20)  # category, muscle_groups, equipment, difficulty
            
            # ========== HOJA 2: REFERENCIAS ==========
            ref_worksheet = workbook.add_worksheet('Referencias')
            
            # Escribir categorías
            ref_worksheet.write(0, 0, 'CATEGORÍAS DISPONIBLES', header_format)
            sorted_categories = sorted(all_categories) if all_categories else ['Cardio', 'Fuerza', 'Flexibilidad', 'Peso corporal', 'HIIT']
            for idx, cat in enumerate(sorted_categories, start=1):
                ref_worksheet.write(idx, 0, cat)
            
            # Escribir grupos musculares
            ref_worksheet.write(0, 2, 'GRUPOS MUSCULARES', header_format)
            sorted_muscle_groups = sorted(all_muscle_groups) if all_muscle_groups else [
                'Pecho', 'Espalda', 'Hombros', 'Bíceps', 'Tríceps', 'Antebrazos',
                'Abdominales', 'Core', 'Oblicuos', 'Cuádriceps', 'Isquiotibiales', 'Glúteos', 'Pantorrillas'
            ]
            for idx, mg in enumerate(sorted_muscle_groups, start=1):
                ref_worksheet.write(idx, 2, mg)
            
            # Escribir equipamiento
            ref_worksheet.write(0, 4, 'EQUIPAMIENTO', header_format)
            sorted_equipment = sorted(all_equipment) if all_equipment else [
                'Barra', 'Mancuernas', 'Kettlebell', 'Bandas de resistencia', 'Peso corporal',
                'Máquina', 'Cable', 'Banco', 'Barra de dominadas', 'Cinta de correr'
            ]
            for idx, eq in enumerate(sorted_equipment, start=1):
                ref_worksheet.write(idx, 4, eq)
            
            # Escribir dificultades
            ref_worksheet.write(0, 6, 'DIFICULTADES', header_format)
            difficulties = sorted(all_difficulties) if all_difficulties else ['Principiante', 'Intermedio', 'Avanzado']
            for idx, diff in enumerate(difficulties, start=1):
                ref_worksheet.write(idx, 6, diff)
            
            # Ajustar ancho de columnas en Referencias
            ref_worksheet.set_column('A:G', 25)
            
            workbook.close()
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="exercises_export.xlsx"'
            return response
    """ViewSet para gestión de ejercicios por administradores"""
    permission_classes = [IsAdminOrStaff]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    pagination_class = LargeResultsSetPagination

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except DatabaseError:
            # Fallback resiliente para evitar 500 en admin si hay corrupción en algún índice/relación
            return Response({
                'count': 0,
                'next': None,
                'previous': None,
                'results': [],
                'warning': 'Resultados parciales por incidencia temporal en base de datos'
            }, status=status.HTTP_200_OK)
    
    def get_queryset(self):
        return Exercise.objects.all().order_by('name')
    
    def get_serializer_class(self):
        return ExerciseSerializer
    
    def perform_create(self, serializer):
        """Guardar el ejercicio con el usuario que lo crea"""
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        """Upsert por nombre: si existe, lo actualiza en lugar de duplicar."""
        incoming_name = Exercise.normalize_name(request.data.get('name', ''))
        if incoming_name:
            existing = Exercise.find_existing_by_name(incoming_name)
            if existing:
                payload = request.data.copy()
                payload['name'] = incoming_name
                serializer = self.get_serializer(existing, data=payload, partial=True)
                serializer.is_valid(raise_exception=True)
                serializer.save(created_by=existing.created_by or self.request.user)
                return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Estadísticas de ejercicios"""
        try:
            queryset = self.get_queryset()

            # Ejercicios por categoría
            category_stats = {}
            for exercise in queryset:
                category = exercise.category or 'Sin categoría'
                category_stats[category] = category_stats.get(category, 0) + 1

            # Ejercicios por grupo muscular
            muscle_group_stats = {}
            for exercise in queryset:
                for muscle_group in exercise.muscle_groups or []:
                    muscle_group_stats[muscle_group] = muscle_group_stats.get(muscle_group, 0) + 1

            # Ejercicios recientes (últimos 30 días)
            thirty_days_ago = timezone.now() - timedelta(days=30)
            recent_exercises = queryset.filter(created_at__gte=thirty_days_ago).count()
            total_exercises = queryset.count()
        except DatabaseError:
            return Response({
                'total_exercises': 0,
                'exercises_by_category': {},
                'exercises_by_muscle_group': {},
                'recent_exercises': 0,
                'warning': 'Estadísticas no disponibles temporalmente por incidencia en base de datos'
            })
        
        stats = {
            'total_exercises': total_exercises,
            'exercises_by_category': category_stats,
            'exercises_by_muscle_group': muscle_group_stats,
            'recent_exercises': recent_exercises
        }
        
        return Response(stats)
    
    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        """Eliminar múltiples ejercicios"""
        exercise_ids = request.data.get('exercise_ids', [])
        
        if not exercise_ids:
            return Response(
                {'error': 'Se requieren IDs de ejercicios'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count, _ = Exercise.objects.filter(
            id__in=exercise_ids
        ).delete()
        
        return Response({
            'message': f'{deleted_count} ejercicios eliminados correctamente',
            'deleted_count': deleted_count
        })
    
    @action(detail=False, methods=['post'])
    def sync_from_google_drive(self, request):
        """
        Sincroniza ejercicios desde Google Drive
        Lista todos los videos de la carpeta configurada y crea/actualiza ejercicios
        """
        try:
            from .google_drive_service import get_google_drive_service
            from .models import Exercise
            import re
            
            service = get_google_drive_service()
            
            if not service.is_configured():
                return Response(
                    {
                        'error': 'Google Drive API no está configurada',
                        'message': 'Configura GOOGLE_DRIVE_CREDENTIALS_PATH o GOOGLE_DRIVE_API_KEY en settings'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener folder_id del request o usar el configurado por defecto
            folder_id = request.data.get('folder_id', None)
            
            # Listar videos de Google Drive
            videos = service.list_videos_from_folder(folder_id)
            
            if not videos:
                return Response(
                    {
                        'message': 'No se encontraron videos en la carpeta',
                        'videos_count': 0
                    },
                    status=status.HTTP_200_OK
                )
            
            # Procesar cada video
            created = 0
            updated = 0
            skipped = 0
            errors = []
            
            for video in videos:
                try:
                    exercise_name = video['name']
                    exercise_name = Exercise.normalize_name(exercise_name)
                    file_id = video['file_id']
                    
                    if not exercise_name or not file_id:
                        skipped += 1
                        continue
                    
                    # Crear o actualizar ejercicio
                    exercise = Exercise.find_existing_by_name(exercise_name)
                    if exercise:
                        exercise.google_drive_file_id = file_id
                        exercise.video_url = f"https://drive.google.com/file/d/{file_id}/preview"
                        exercise.save()
                        updated += 1
                    else:
                        exercise = Exercise.objects.create(
                            name=exercise_name,
                            google_drive_file_id=file_id,
                            video_url=f"https://drive.google.com/file/d/{file_id}/preview",
                            category='strength',
                        )
                        created += 1
                        
                except Exception as e:
                    errors.append({
                        'video': video.get('filename', 'Desconocido'),
                        'error': str(e)
                    })
                    skipped += 1
            
            return Response({
                'message': 'Sincronización completada',
                'total_videos': len(videos),
                'created': created,
                'updated': updated,
                'skipped': skipped,
                'errors': errors if errors else None
            }, status=status.HTTP_200_OK)
            
        except ImportError as e:
            return Response(
                {
                    'error': 'Dependencias faltantes',
                    'message': str(e),
                    'solution': 'Instala las dependencias: pip install google-api-python-client google-auth'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {
                    'error': 'Error al sincronizar desde Google Drive',
                    'message': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='add_substitute')
    def add_substitute(self, request, pk=None):
        """Añadir un ejercicio sustituto"""
        exercise = self.get_object()
        substitute_id = request.data.get('substitute_id')
        priority = request.data.get('priority', 1)
        notes = request.data.get('notes', '')
        
        if not substitute_id:
            return Response({'error': 'substitute_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        substitute = get_object_or_404(Exercise, id=substitute_id)
        
        if substitute.id == exercise.id:
            return Response({'error': 'Un ejercicio no puede ser sustituto de sí mismo'}, status=status.HTTP_400_BAD_REQUEST)
        
        sub, created = ExerciseSubstitution.objects.get_or_create(
            exercise=exercise,
            substitute=substitute,
            defaults={'priority': priority, 'notes': notes}
        )
        
        if not created:
            sub.priority = priority
            sub.notes = notes
            sub.save()
        
        return Response({
            'id': sub.id,
            'substitute_id': str(substitute.id),
            'substitute_name': substitute.name,
            'priority': sub.priority,
            'notes': sub.notes,
            'created': created
        })
    
    @action(detail=True, methods=['post'], url_path='remove_substitute')
    def remove_substitute(self, request, pk=None):
        """Eliminar un ejercicio sustituto"""
        exercise = self.get_object()
        substitute_id = request.data.get('substitute_id')
        
        if not substitute_id:
            return Response({'error': 'substitute_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        deleted, _ = ExerciseSubstitution.objects.filter(
            exercise=exercise,
            substitute_id=substitute_id
        ).delete()
        
        return Response({'removed': deleted > 0})
    
    @action(detail=False, methods=['post'], url_path='auto-link-videos')
    def auto_link_videos(self, request):
        """
        Vincula ejercicios con videos de la carpeta pública de Google Drive.
        El servidor hace la petición HTTP directamente, sin API key ni OAuth.
        Acepta opcionalmente drive_map:{clave:file_id} precalculado desde el frontend.
        """
        from .google_drive_service import (
            DEFAULT_GOOGLE_DRIVE_FOLDER_ID,
            build_drive_map,
            get_google_drive_service,
            normalize_drive_match_key,
        )

        force = str(request.data.get('force', 'false')).lower() in ('true', '1', 'yes')
        drive_map = request.data.get('drive_map')  # opcional, lo puede mandar el frontend
        folder_url = request.data.get('folder_url') or request.data.get('folder_id')
        videos_in_drive = len(drive_map) if isinstance(drive_map, dict) else 0

        if not drive_map:
            try:
                service = get_google_drive_service()
                videos = service.list_public_videos_from_folder(folder_url or DEFAULT_GOOGLE_DRIVE_FOLDER_ID)
                drive_map = build_drive_map(videos)
                videos_in_drive = len(videos)

            except Exception as exc:
                return Response(
                    {'error': f'No se pudo acceder a la carpeta de Drive: {exc}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if not drive_map:
            return Response(
                {'error': 'No se encontraron vídeos en la carpeta de Drive'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import Exercise
        linked = 0
        skipped = 0
        matched_names = []

        if force:
            qs = Exercise.objects.all()
        else:
            qs = Exercise.objects.filter(google_drive_file_id__isnull=True) | Exercise.objects.filter(google_drive_file_id='')

        for ex in qs:
            key = normalize_drive_match_key(ex.name)
            file_id = drive_map.get(key)
            if file_id:
                ex.google_drive_file_id = file_id
                ex.video_url = f'https://drive.google.com/file/d/{file_id}/preview'
                ex.save(update_fields=['google_drive_file_id', 'video_url', 'updated_at'])
                linked += 1
                matched_names.append(ex.name)
            else:
                skipped += 1

        return Response({
            'videos_in_drive': videos_in_drive or len(drive_map),
            'linked': linked,
            'skipped_no_match': skipped,
            'matched': matched_names,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='substitutes')
    def list_substitutes(self, request, pk=None):
        """Listar sustitutos de un ejercicio"""
        exercise = self.get_object()
        subs = exercise.substitutions.all().select_related('substitute').order_by('priority')
        
        return Response([{
            'id': s.id,
            'substitute_id': str(s.substitute.id),
            'substitute_name': s.substitute.name,
            'category': s.substitute.category,
            'priority': s.priority,
            'notes': s.notes
        } for s in subs])







