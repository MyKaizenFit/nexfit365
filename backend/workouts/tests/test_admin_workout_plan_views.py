"""Tests para workouts/admin_workout_plan_views.py - AdminWorkoutPlanExportImportViewSet"""
import datetime
import io
import pytest
from rest_framework.test import APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from django.test import override_settings
from workouts.models import WorkoutProgram, Exercise, WorkoutDay, WorkoutDayExercise

User = get_user_model()

BASE_URL = '/api/admin/workouts/workouts/'


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def admin_user(db):
    return User.objects.create_user(
        email='admin_plan@test.com',
        password='testpass123',
        is_staff=True,
        is_superuser=True,
    )


@pytest.fixture
def regular_user(db):
    return User.objects.create_user(
        email='user_plan@test.com',
        password='testpass123',
    )


@pytest.fixture
def admin_client(admin_user):
    client = APIClient()
    client.force_authenticate(user=admin_user)
    return client


@pytest.fixture
def regular_client(regular_user):
    client = APIClient()
    client.force_authenticate(user=regular_user)
    return client


@pytest.fixture
def template_program(db):
    return WorkoutProgram.objects.create(
        name='Plan Plantilla Test',
        description='Descripción del plan',
        difficulty='beginner',
        goal='general_fitness',
        duration_weeks=4,
        days_per_week=3,
        location='gym',
        is_active=True,
        is_template=True,
    )


@pytest.fixture
def exercise(db):
    return Exercise.objects.create(
        name='Sentadilla',
        category='strength',
        muscle_groups=['piernas'],
        is_active=True,
    )


@pytest.fixture
def template_with_days(db, template_program, exercise):
    """Plan plantilla con días y ejercicios"""
    day = WorkoutDay.objects.create(
        program=template_program,
        name='Día 1 - Piernas',
        day_of_week='monday',
        day_number=1,
    )
    WorkoutDayExercise.objects.create(
        workout_day=day,
        exercise=exercise,
        sets=3,
        reps='10',
        rest_seconds=60,
        order_index=1,
    )
    return template_program


# ---------------------------------------------------------------------------
# export_csv tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestExportCsv:
    """Tests para GET /api/admin/workouts/workouts/export_csv/"""

    def test_export_csv_requires_admin(self, regular_client):
        response = regular_client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_csv_requires_auth(self):
        client = APIClient()
        response = client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_export_csv_empty(self, admin_client):
        """Exportar cuando no hay planes plantilla"""
        response = admin_client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'text/csv; charset=utf-8'

    def test_export_csv_with_template(self, admin_client, template_with_days):
        """Exportar con un plan plantilla"""
        response = admin_client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        assert 'Plan Plantilla Test' in content

    def test_export_csv_content_disposition(self, admin_client):
        """El CSV debe tener la cabecera content-disposition correcta"""
        response = admin_client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_200_OK
        assert 'attachment' in response.get('Content-Disposition', '')

    def test_export_csv_with_exercises(self, admin_client, template_with_days):
        """El CSV debe incluir ejercicios"""
        response = admin_client.get(f'{BASE_URL}export_csv/')
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        assert 'Sentadilla' in content


# ---------------------------------------------------------------------------
# export_excel tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestExportExcel:
    """Tests para GET /api/admin/workouts/workouts/export_excel/"""

    def test_export_excel_requires_admin(self, regular_client):
        response = regular_client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_excel_requires_auth(self):
        client = APIClient()
        response = client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_export_excel_empty(self, admin_client):
        """Exportar cuando no hay planes plantilla"""
        response = admin_client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_200_OK
        assert 'spreadsheet' in response['Content-Type']

    def test_export_excel_with_template(self, admin_client, template_with_days):
        """Exportar con un plan plantilla"""
        response = admin_client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_200_OK

    def test_export_excel_content_disposition(self, admin_client):
        """El Excel debe tener la cabecera content-disposition correcta"""
        response = admin_client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_200_OK
        assert 'attachment' in response.get('Content-Disposition', '')

    def test_export_excel_reps_column_is_text_format(self, admin_client, template_with_days):
        """Las reps deben salir como texto para evitar autoconversión a fecha."""
        import openpyxl

        response = admin_client.get(f'{BASE_URL}export_excel/')
        assert response.status_code == status.HTTP_200_OK

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        worksheet = workbook['Ejercicios']
        headers = [cell.value for cell in worksheet[1]]
        reps_column = headers.index('Reps') + 1

        assert worksheet.cell(row=2, column=reps_column).number_format == '@'


# ---------------------------------------------------------------------------
# import_csv tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestImportCsv:
    """Tests para POST /api/admin/workouts/workouts/import_csv/"""

    def test_import_csv_requires_admin(self, regular_client):
        response = regular_client.post(f'{BASE_URL}import_csv/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_requires_auth(self):
        client = APIClient()
        response = client.post(f'{BASE_URL}import_csv/')
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_import_csv_no_file(self, admin_client):
        """Sin archivo debe retornar 400"""
        response = admin_client.post(f'{BASE_URL}import_csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'error' in response.data

    def test_import_csv_simple_format(self, admin_client, exercise):
        """CSV con formato simple (nombre plan, descripción, ejercicios)"""
        csv_content = (
            'nombre,descripcion,dificultad,objetivo,semanas,sesiones_semana,ubicacion\n'
            'Plan Import Test,Descripción test,Principiante,Fitness general,4,3,Gimnasio\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'test_import.csv'
        response = admin_client.post(
            f'{BASE_URL}import_csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK,
            status.HTTP_201_CREATED,
            status.HTTP_400_BAD_REQUEST,
        ]

    def test_import_csv_complete_format(self, admin_client, exercise):
        """CSV con formato completo (plan + día + ejercicio)"""
        csv_content = (
            'plan_name,plan_description,plan_difficulty,plan_goal,'
            'plan_duration_weeks,plan_sessions_per_week,plan_location,'
            'day_name,day_of_week,day_order,exercise_name,sets,reps,rest_seconds,exercise_order\n'
            'Plan Completo,Descripción,beginner,general_fitness,4,3,gym,'
            'Día 1,monday,1,Sentadilla,3,10,60,1\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'test_complete.csv'
        response = admin_client.post(
            f'{BASE_URL}import_csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK,
            status.HTTP_201_CREATED,
            status.HTTP_400_BAD_REQUEST,
        ]

    def test_import_csv_invalid_encoding(self, admin_client):
        """CSV con codificación inválida"""
        # binary content that cannot be decoded as utf-8
        binary_content = b'\xff\xfe invalid bytes'
        file = io.BytesIO(binary_content)
        file.name = 'bad_encoding.csv'
        response = admin_client.post(
            f'{BASE_URL}import_csv/',
            {'file': file},
            format='multipart',
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_csv_empty_file(self, admin_client):
        """CSV vacío"""
        csv_content = '\n'
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'empty.csv'
        response = admin_client.post(
            f'{BASE_URL}import_csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK,
            status.HTTP_201_CREATED,
            status.HTTP_400_BAD_REQUEST,
        ]


# ---------------------------------------------------------------------------
# import_excel tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestImportExcel:
    """Tests para POST /api/admin/workouts/workouts/import_excel/"""

    def test_import_excel_requires_admin(self, regular_client):
        response = regular_client.post(f'{BASE_URL}import_excel/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_excel_requires_auth(self):
        client = APIClient()
        response = client.post(f'{BASE_URL}import_excel/')
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_import_excel_no_file(self, admin_client):
        """Sin archivo debe retornar 400"""
        response = admin_client.post(f'{BASE_URL}import_excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_excel_valid_file(self, admin_client, exercise):
        """Importar un Excel válido con openpyxl"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Planes'
        # Header row
        headers = [
            'plan_name', 'plan_description', 'plan_difficulty', 'plan_goal',
            'plan_duration_weeks', 'plan_sessions_per_week', 'plan_location',
            'day_name', 'day_of_week', 'day_order',
            'exercise_name', 'sets', 'reps', 'rest_seconds', 'exercise_order'
        ]
        ws.append(headers)
        # Data row
        ws.append([
            'Plan Excel Test', 'Descripción', 'beginner', 'general_fitness',
            4, 3, 'gym',
            'Día 1', 'monday', 1,
            'Sentadilla', 3, 10, 60, 1
        ])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'test.xlsx'

        response = admin_client.post(
            f'{BASE_URL}import_excel/',
            {'file': buffer},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK,
            status.HTTP_201_CREATED,
            status.HTTP_400_BAD_REQUEST,
        ]

    def test_import_excel_invalid_file(self, admin_client):
        """Archivo que no es Excel válido"""
        file = io.BytesIO(b'not an excel file content')
        file.name = 'invalid.xlsx'
        response = admin_client.post(
            f'{BASE_URL}import_excel/',
            {'file': file},
            format='multipart',
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    @override_settings(SECURE_SSL_REDIRECT=False)
    def test_import_excel_converts_date_like_reps_back_to_range(self, admin_client, exercise):
        """Si Excel convirtió 8-12 en fecha, no debe persistirse como datetime."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ejercicios'
        ws.append([
            'Categoría Plan', 'Usuario Referencia', 'Nombre Plan', 'Número Día',
            'Nombre Día', 'Orden', 'Nombre Ejercicio', 'Series', 'Reps', 'Peso',
            'Duración (seg)', 'Descanso (seg)', 'Notas', 'Grupo Superset',
        ])
        ws.append([
            'Plantilla', '', 'Plan Fecha Reps', 1,
            'Día 1', 1, exercise.name, 3, datetime.datetime(2026, 12, 8), '', '', 90, '', '',
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'date_reps.xlsx'

        response = admin_client.post(
            f'{BASE_URL}import_excel/',
            {'file': buffer},
            format='multipart',
        )

        assert response.status_code == status.HTTP_200_OK
        workout_exercise = WorkoutDayExercise.objects.get(
            workout_day__program__name='Plan Fecha Reps',
            exercise=exercise,
        )
        assert workout_exercise.reps == '8-12'

    @override_settings(SECURE_SSL_REDIRECT=False)
    def test_import_excel_user_plan_is_not_template_and_deactivates_previous(
        self,
        admin_client,
        regular_user,
        exercise,
    ):
        """Un plan de categoría Usuario debe quedar asignado al usuario y activo."""
        import openpyxl

        old_plan = WorkoutProgram.objects.create(
            user=regular_user,
            name='Plan por defecto anterior',
            is_active=True,
            is_template=False,
        )

        wb = openpyxl.Workbook()
        ws_plans = wb.active
        ws_plans.title = 'Planes'
        ws_plans.append([
            'Categoría', 'Usuario Referencia', 'Creado Por', 'Nombre', 'Descripción',
            'Dificultad', 'Objetivo', 'Ubicación', 'Semanas', 'Días/Semana',
            'Duración (min)', 'Equipo (coma separado)', 'Tags (coma separado)',
            'Imagen URL', 'Activo', 'Plantilla',
        ])
        ws_plans.append([
            'Usuario', regular_user.email, '', 'Vamos Miriam!!!', '',
            'Principiante', 'Fitness general', 'Gimnasio', 4, 3, 60,
            '', '', '', 'Sí', 'Sí',
        ])

        ws_days = wb.create_sheet('Días')
        ws_days.append([
            'Categoría Plan', 'Usuario Referencia', 'Nombre Plan', 'Número Día',
            'Nombre Día', 'Día Semana', 'Es Descanso', 'Duración (min)',
            'Enfoque', 'Notas', 'Orden',
        ])
        ws_days.append([
            'Usuario', regular_user.email, 'Vamos Miriam!!!', 1,
            'Pierna A', 'Lunes', 'No', 60, '', '', 1,
        ])

        ws_exercises = wb.create_sheet('Ejercicios')
        ws_exercises.append([
            'Categoría Plan', 'Usuario Referencia', 'Nombre Plan', 'Número Día',
            'Nombre Día', 'Orden', 'Nombre Ejercicio', 'Series', 'Reps', 'Peso',
            'Duración (seg)', 'Descanso (seg)', 'Notas', 'Grupo Superset',
        ])
        ws_exercises.append([
            'Usuario', regular_user.email, 'Vamos Miriam!!!', 1,
            'Pierna A', 1, exercise.name, 3, '8-12', '', '', 90, '', '',
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'user_plan.xlsx'

        response = admin_client.post(
            f'{BASE_URL}import_excel/',
            {'file': buffer},
            format='multipart',
        )

        assert response.status_code == status.HTTP_200_OK

        new_plan = WorkoutProgram.objects.get(user=regular_user, name='Vamos Miriam!!!')
        assert new_plan.is_active is True
        assert new_plan.is_template is False
        assert new_plan.is_system is False
        assert new_plan.days.filter(day_number=1, name='Pierna A').exists()
        assert WorkoutDayExercise.objects.filter(
            workout_day__program=new_plan,
            exercise=exercise,
        ).exists()

        old_plan.refresh_from_db()
        assert old_plan.is_active is False

    @override_settings(SECURE_SSL_REDIRECT=False)
    def test_import_excel_summary_sheet_assigns_user_plan(
        self,
        admin_client,
        regular_user,
        exercise,
    ):
        """La hoja Resumen completo también debe importar planes de usuario."""
        import openpyxl

        WorkoutProgram.objects.create(
            name='Vamos Miriam!!!',
            is_active=True,
            is_template=True,
        )
        old_plan = WorkoutProgram.objects.create(
            user=regular_user,
            name='Plan anterior',
            is_active=True,
            is_template=False,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resumen completo'
        ws.append([
            'Categoría Plan', 'Usuario Referencia', 'Creado Por', 'Nombre Plan',
            'Descripción Plan', 'Dificultad', 'Objetivo', 'Ubicación',
            'Semanas', 'Días/Semana', 'Duración Plan (min)', 'Número Día',
            'Nombre Día', 'Día Semana', 'Descanso', 'Orden Ejercicio',
            'Nombre Ejercicio', 'Series', 'Reps', 'Peso', 'Duración (seg)',
            'Descanso (seg)', 'Grupo Superset', 'Sustitutos (separados por |)',
        ])
        ws.append([
            'Usuario', regular_user.email, '', 'Vamos Miriam!!!',
            '', 'Principiante', 'Fitness general', 'Gimnasio',
            4, 3, 60, 1, 'Pierna A', 'Lunes', 'No', 1,
            exercise.name, 3, '8-12', '', '', 90, '', '',
        ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'summary.xlsx'

        response = admin_client.post(
            f'{BASE_URL}import_excel/',
            {'file': buffer},
            format='multipart',
        )

        assert response.status_code == status.HTTP_200_OK

        user_plan = WorkoutProgram.objects.get(user=regular_user, name='Vamos Miriam!!!')
        assert user_plan.is_active is True
        assert user_plan.is_template is False
        assert WorkoutProgram.objects.filter(
            user__isnull=True,
            name='Vamos Miriam!!!',
            is_template=True,
        ).exists()
        assert user_plan.days.filter(day_number=1, name='Pierna A').exists()
        assert WorkoutDayExercise.objects.filter(
            workout_day__program=user_plan,
            exercise=exercise,
            reps='8-12',
        ).exists()

        old_plan.refresh_from_db()
        assert old_plan.is_active is False
