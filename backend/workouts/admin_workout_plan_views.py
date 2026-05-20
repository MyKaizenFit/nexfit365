# workouts/admin_workout_plan_views.py
"""
Views para export/import de planes de entrenamiento
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from django.contrib.auth import get_user_model
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import WorkoutProgram, Exercise, WorkoutDay, WorkoutDayExercise, ExerciseSubstitution


class AdminWorkoutPlanExportImportViewSet(viewsets.GenericViewSet):
    """ViewSet para export/import de planes de entrenamiento"""
    permission_classes = [IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]
    queryset = WorkoutProgram.objects.all()  # Required for router registration

    def _get_export_plans(self):
        return WorkoutProgram.objects.all().select_related('user', 'created_by').prefetch_related(
            'days__exercises__exercise'
        )

    def _get_plan_category(self, plan: WorkoutProgram):
        if getattr(plan, 'user_id', None):
            return 'Usuario'
        if getattr(plan, 'is_system', False):
            return 'Sistema'
        if getattr(plan, 'is_template', False):
            return 'Plantilla'
        return 'Otro'
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporta CSV completo: plan + día + ejercicio + sustitutos (sin IDs), en español."""
        try:
            difficulty_map = {
                'beginner': 'Principiante',
                'intermediate': 'Intermedio',
                'advanced': 'Avanzado',
            }
            goal_map = {
                'weight_loss': 'Pérdida de peso',
                'muscle_gain': 'Ganancia muscular',
                'strength': 'Fuerza',
                'endurance': 'Resistencia',
                'general_fitness': 'Fitness general',
                'body_recomposition': 'Recomposición corporal',
            }
            location_map = {
                'gym': 'Gimnasio',
                'home': 'Casa',
                'outdoor': 'Exterior',
                'any': 'Cualquier lugar',
            }
            day_map = {
                'monday': 'Lunes',
                'tuesday': 'Martes',
                'wednesday': 'Miércoles',
                'thursday': 'Jueves',
                'friday': 'Viernes',
                'saturday': 'Sábado',
                'sunday': 'Domingo',
            }

            def to_spanish(value, mapping):
                key = (value or '').strip().lower()
                return mapping.get(key, value or '')

            plans = self._get_export_plans()

            used_exercise_ids = set()
            for plan in plans:
                for day in plan.days.all():
                    for workout_exercise in day.exercises.all():
                        used_exercise_ids.add(workout_exercise.exercise_id)

            substitutions_map = {}
            substitutions = ExerciseSubstitution.objects.filter(
                exercise_id__in=used_exercise_ids
            ).select_related('substitute').order_by('exercise_id', 'priority', 'created_at')
            for substitution in substitutions:
                substitutions_map.setdefault(str(substitution.exercise_id), []).append(substitution.substitute.name)

            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=[
                'categoria_plan', 'usuario_referencia', 'creado_por',
                'nombre_plan', 'descripcion_plan', 'dificultad_plan', 'objetivo_plan', 'ubicacion_plan',
                'duracion_semanas_plan', 'dias_por_semana_plan', 'duracion_estimada_minutos_plan',
                'equipo_necesario_plan', 'etiquetas_plan', 'imagen_url_plan', 'plan_activo', 'plan_plantilla',
                'numero_dia', 'nombre_dia', 'dia_semana', 'dia_descanso', 'duracion_dia_minutos',
                'enfoque_dia', 'notas_dia', 'orden_dia',
                'orden_ejercicio', 'nombre_ejercicio', 'series', 'reps', 'peso',
                'duracion_segundos', 'descanso_segundos', 'notas_ejercicio', 'grupo_superset',
                'sustitutos'
            ])

            writer.writeheader()
            for plan in plans:
                base_row = {
                    'categoria_plan': self._get_plan_category(plan),
                    'usuario_referencia': getattr(plan.user, 'email', '') or '',
                    'creado_por': getattr(plan.created_by, 'email', '') or '',
                    'nombre_plan': plan.name or '',
                    'descripcion_plan': plan.description or '',
                    'dificultad_plan': to_spanish(plan.difficulty, difficulty_map),
                    'objetivo_plan': to_spanish(plan.goal, goal_map),
                    'ubicacion_plan': to_spanish(plan.location, location_map),
                    'duracion_semanas_plan': plan.duration_weeks or 4,
                    'dias_por_semana_plan': plan.days_per_week or 3,
                    'duracion_estimada_minutos_plan': plan.estimated_duration_minutes or 60,
                    'equipo_necesario_plan': ', '.join(plan.equipment_needed or []),
                    'etiquetas_plan': ', '.join(plan.tags or []),
                    'imagen_url_plan': plan.image_url or '',
                    'plan_activo': 'Sí' if plan.is_active else 'No',
                    'plan_plantilla': 'Sí' if plan.is_template else 'No',
                }

                days = list(plan.days.all().order_by('order_index', 'day_number'))
                if not days:
                    writer.writerow(base_row)
                    continue

                for day in days:
                    day_row = {
                        **base_row,
                        'numero_dia': day.day_number,
                        'nombre_dia': day.name or '',
                        'dia_semana': to_spanish(day.day_of_week, day_map),
                        'dia_descanso': 'Sí' if day.is_rest_day else 'No',
                        'duracion_dia_minutos': day.duration_minutes or '',
                        'enfoque_dia': day.focus or '',
                        'notas_dia': day.notes or '',
                        'orden_dia': day.order_index,
                    }

                    day_exercises = list(day.exercises.all().select_related('exercise').order_by('order_index'))
                    if not day_exercises:
                        writer.writerow(day_row)
                        continue

                    for workout_exercise in day_exercises:
                        substitutes = substitutions_map.get(str(workout_exercise.exercise_id), [])
                        writer.writerow({
                            **day_row,
                            'orden_ejercicio': workout_exercise.order_index,
                            'nombre_ejercicio': workout_exercise.exercise.name,
                            'series': workout_exercise.sets or 3,
                            'reps': workout_exercise.reps or '10',
                            'peso': workout_exercise.weight or '',
                            'duracion_segundos': workout_exercise.duration_seconds or '',
                            'descanso_segundos': workout_exercise.rest_seconds or 60,
                            'notas_ejercicio': workout_exercise.notes or '',
                            'grupo_superset': workout_exercise.superset_group or '',
                            'sustitutos': ' | '.join(substitutes),
                        })

            response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="workout_plans_export.csv"'
            response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporta planes completos a Excel sin IDs (editable por negocio)."""
        try:
            def parse_bool_label(value):
                return 'Sí' if value else 'No'

            difficulty_map = {
                'beginner': 'Principiante',
                'intermediate': 'Intermedio',
                'advanced': 'Avanzado',
            }
            goal_map = {
                'weight_loss': 'Pérdida de peso',
                'muscle_gain': 'Ganancia muscular',
                'strength': 'Fuerza',
                'endurance': 'Resistencia',
                'general_fitness': 'Fitness general',
                'body_recomposition': 'Recomposición corporal',
            }
            location_map = {
                'gym': 'Gimnasio',
                'home': 'Casa',
                'outdoor': 'Exterior',
                'any': 'Cualquier lugar',
            }
            day_map = {
                'monday': 'Lunes',
                'tuesday': 'Martes',
                'wednesday': 'Miércoles',
                'thursday': 'Jueves',
                'friday': 'Viernes',
                'saturday': 'Sábado',
                'sunday': 'Domingo',
            }

            def to_spanish(value, mapping):
                key = (value or '').strip().lower()
                return mapping.get(key, value or '')

            def apply_header_style(ws):
                """Aplica estilo a la primera fila"""
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            def auto_adjust_columns(ws):
                """Ajusta el ancho de columnas automáticamente"""
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Resumen completo"
            ws_summary.append([
                'Categoría Plan', 'Usuario Referencia', 'Creado Por', 'Nombre Plan', 'Descripción Plan', 'Dificultad', 'Objetivo', 'Ubicación',
                'Semanas', 'Días/Semana', 'Duración Plan (min)',
                'Número Día', 'Nombre Día', 'Día Semana', 'Descanso',
                'Orden Ejercicio', 'Nombre Ejercicio', 'Series', 'Reps',
                'Peso', 'Duración (seg)', 'Descanso (seg)', 'Grupo Superset',
                'Sustitutos (separados por |)'
            ])
            apply_header_style(ws_summary)
            
            # === HOJA 1: PLANES ===
            ws_plans = wb.create_sheet("Planes")
            ws_plans.append([
                'Categoría', 'Usuario Referencia', 'Creado Por', 'Nombre', 'Descripción', 'Dificultad', 'Objetivo', 'Ubicación',
                'Semanas', 'Días/Semana', 'Duración (min)', 'Equipo (coma separado)',
                'Tags (coma separado)', 'Imagen URL', 'Activo', 'Plantilla'
            ])
            apply_header_style(ws_plans)
            
            plans = list(self._get_export_plans())
            base_plans = [plan for plan in plans if not getattr(plan, 'user_id', None)]
            user_plans = [plan for plan in plans if getattr(plan, 'user_id', None)]
            for plan in plans:
                ws_plans.append([
                    self._get_plan_category(plan),
                    getattr(plan.user, 'email', '') or '',
                    getattr(plan.created_by, 'email', '') or '',
                    plan.name or '',
                    plan.description or '',
                    to_spanish(plan.difficulty or 'beginner', difficulty_map),
                    to_spanish(plan.goal or '', goal_map),
                    to_spanish(plan.location or '', location_map),
                    plan.duration_weeks or 4,
                    plan.days_per_week or 3,
                    plan.estimated_duration_minutes or 60,
                    ', '.join(plan.equipment_needed or []),
                    ', '.join(plan.tags or []),
                    plan.image_url or '',
                    parse_bool_label(plan.is_active),
                    parse_bool_label(plan.is_template),
                ])
            auto_adjust_columns(ws_plans)

            # === HOJA 2: DÍAS ===
            ws_days = wb.create_sheet("Días")
            ws_days.append([
                'Categoría Plan', 'Usuario Referencia', 'Nombre Plan', 'Número Día', 'Nombre Día', 'Día Semana', 'Es Descanso',
                'Duración (min)', 'Enfoque', 'Notas', 'Orden'
            ])
            apply_header_style(ws_days)
            
            for plan in plans:
                for day in plan.days.all().order_by('order_index', 'day_number'):
                    ws_days.append([
                        self._get_plan_category(plan),
                        getattr(plan.user, 'email', '') or '',
                        plan.name,
                        day.day_number,
                        day.name or '',
                        to_spanish(day.day_of_week or '', day_map),
                        parse_bool_label(day.is_rest_day),
                        day.duration_minutes or '',
                        day.focus or '',
                        day.notes or '',
                        day.order_index,
                    ])
            auto_adjust_columns(ws_days)

            # === HOJA 3: EJERCICIOS ===
            ws_exercises = wb.create_sheet("Ejercicios")
            ws_exercises.append([
                'Categoría Plan', 'Usuario Referencia', 'Nombre Plan', 'Número Día', 'Nombre Día', 'Orden', 'Nombre Ejercicio',
                'Series', 'Reps', 'Peso', 'Duración (seg)', 'Descanso (seg)',
                'Notas', 'Grupo Superset'
            ])
            apply_header_style(ws_exercises)
            
            for plan in plans:
                for day in plan.days.all().order_by('order_index', 'day_number'):
                    for exercise in day.exercises.all().select_related('exercise').order_by('order_index'):
                        ws_exercises.append([
                            self._get_plan_category(plan),
                            getattr(plan.user, 'email', '') or '',
                            plan.name,
                            day.day_number,
                            day.name,
                            exercise.order_index,
                            exercise.exercise.name,
                            exercise.sets or 3,
                            exercise.reps or '10',
                            exercise.weight or '',
                            exercise.duration_seconds or '',
                            exercise.rest_seconds or 60,
                            exercise.notes or '',
                            exercise.superset_group or '',
                        ])
            auto_adjust_columns(ws_exercises)

            ws_user_plans = wb.create_sheet("Planes_Usuario")
            ws_user_plans.append([
                'Usuario Referencia', 'Creado Por', 'Nombre', 'Descripción', 'Dificultad', 'Objetivo', 'Ubicación',
                'Semanas', 'Días/Semana', 'Duración (min)', 'Activo'
            ])
            apply_header_style(ws_user_plans)
            for plan in user_plans:
                ws_user_plans.append([
                    getattr(plan.user, 'email', '') or '',
                    getattr(plan.created_by, 'email', '') or '',
                    plan.name or '',
                    plan.description or '',
                    to_spanish(plan.difficulty or 'beginner', difficulty_map),
                    to_spanish(plan.goal or '', goal_map),
                    to_spanish(plan.location or '', location_map),
                    plan.duration_weeks or 4,
                    plan.days_per_week or 3,
                    plan.estimated_duration_minutes or 60,
                    parse_bool_label(plan.is_active),
                ])
            auto_adjust_columns(ws_user_plans)

            ws_base_plans = wb.create_sheet("Planes_Base")
            ws_base_plans.append([
                'Categoría', 'Nombre', 'Descripción', 'Dificultad', 'Objetivo', 'Ubicación',
                'Semanas', 'Días/Semana', 'Duración (min)', 'Activo', 'Plantilla'
            ])
            apply_header_style(ws_base_plans)
            for plan in base_plans:
                ws_base_plans.append([
                    self._get_plan_category(plan),
                    plan.name or '',
                    plan.description or '',
                    to_spanish(plan.difficulty or 'beginner', difficulty_map),
                    to_spanish(plan.goal or '', goal_map),
                    to_spanish(plan.location or '', location_map),
                    plan.duration_weeks or 4,
                    plan.days_per_week or 3,
                    plan.estimated_duration_minutes or 60,
                    parse_bool_label(plan.is_active),
                    parse_bool_label(plan.is_template),
                ])
            auto_adjust_columns(ws_base_plans)

            # === HOJA 4: SUSTITUTOS ===
            ws_subs = wb.create_sheet("Sustitutos")
            ws_subs.append(['Nombre Ejercicio', 'Nombre Sustituto', 'Prioridad', 'Notas'])
            apply_header_style(ws_subs)
            
            # Obtener todos los ejercicios usados en los planes
            used_exercise_ids = set()
            for plan in plans:
                for day in plan.days.all():
                    for ex in day.exercises.all():
                        used_exercise_ids.add(ex.exercise.id)
            
            # Exportar sustitutos de los ejercicios usados
            substitutions = ExerciseSubstitution.objects.filter(
                exercise_id__in=used_exercise_ids
            ).select_related('exercise', 'substitute').order_by('exercise__name', 'priority')

            summary_substitutions_map = {}
            for sub in substitutions:
                summary_substitutions_map.setdefault(str(sub.exercise_id), []).append(sub.substitute.name)
            
            for sub in substitutions:
                ws_subs.append([
                    sub.exercise.name,
                    sub.substitute.name,
                    sub.priority,
                    sub.notes or '',
                ])
            auto_adjust_columns(ws_subs)

            # Completar hoja principal de resumen completo
            for plan in plans:
                days = list(plan.days.all().order_by('order_index', 'day_number'))
                if not days:
                    ws_summary.append([
                        self._get_plan_category(plan),
                        getattr(plan.user, 'email', '') or '',
                        getattr(plan.created_by, 'email', '') or '',
                        plan.name or '',
                        plan.description or '',
                        to_spanish(plan.difficulty or '', difficulty_map),
                        to_spanish(plan.goal or '', goal_map),
                        to_spanish(plan.location or '', location_map),
                        plan.duration_weeks or 4,
                        plan.days_per_week or 3,
                        plan.estimated_duration_minutes or 60,
                        '', '', '', '', '', '', '', '', '', '', '', '', ''
                    ])
                    continue

                for day in days:
                    day_exercises = list(day.exercises.all().select_related('exercise').order_by('order_index'))
                    if not day_exercises:
                        ws_summary.append([
                            self._get_plan_category(plan),
                            getattr(plan.user, 'email', '') or '',
                            getattr(plan.created_by, 'email', '') or '',
                            plan.name or '',
                            plan.description or '',
                            to_spanish(plan.difficulty or '', difficulty_map),
                            to_spanish(plan.goal or '', goal_map),
                            to_spanish(plan.location or '', location_map),
                            plan.duration_weeks or 4,
                            plan.days_per_week or 3,
                            plan.estimated_duration_minutes or 60,
                            day.day_number,
                            day.name or '',
                            to_spanish(day.day_of_week or '', day_map),
                            parse_bool_label(day.is_rest_day),
                            '', '', '', '', '', '', '', '', ''
                        ])
                        continue

                    for exercise in day_exercises:
                        substitutes = summary_substitutions_map.get(str(exercise.exercise_id), [])
                        ws_summary.append([
                            self._get_plan_category(plan),
                            getattr(plan.user, 'email', '') or '',
                            getattr(plan.created_by, 'email', '') or '',
                            plan.name or '',
                            plan.description or '',
                            to_spanish(plan.difficulty or '', difficulty_map),
                            to_spanish(plan.goal or '', goal_map),
                            to_spanish(plan.location or '', location_map),
                            plan.duration_weeks or 4,
                            plan.days_per_week or 3,
                            plan.estimated_duration_minutes or 60,
                            day.day_number,
                            day.name or '',
                            to_spanish(day.day_of_week or '', day_map),
                            parse_bool_label(day.is_rest_day),
                            exercise.order_index,
                            exercise.exercise.name,
                            exercise.sets or 3,
                            exercise.reps or '10',
                            exercise.weight or '',
                            exercise.duration_seconds or '',
                            exercise.rest_seconds or 60,
                            exercise.superset_group or '',
                            ' | '.join(substitutes),
                        ])
            auto_adjust_columns(ws_summary)

            # === HOJA 5: REFERENCIAS ===
            ws_ref = wb.create_sheet("Referencias")
            ws_ref.append(['CAMPO', 'VALORES VÁLIDOS', 'DESCRIPCIÓN'])
            apply_header_style(ws_ref)
            
            ws_ref.append(['Dificultad', 'Principiante, Intermedio, Avanzado', 'Nivel de dificultad'])
            ws_ref.append(['Objetivo', 'Pérdida de peso, Ganancia muscular, Fuerza, Resistencia, Fitness general, Recomposición corporal', 'Objetivo del plan'])
            ws_ref.append(['Ubicación', 'Gimnasio, Casa, Exterior, Cualquier lugar', 'Lugar de entrenamiento'])
            ws_ref.append(['Día Semana', 'Lunes, Martes, Miércoles, Jueves, Viernes, Sábado, Domingo', 'Día de la semana'])
            ws_ref.append(['Es Descanso', 'Sí, No', 'Indica si es día de descanso'])
            ws_ref.append(['Activo', 'Sí, No', 'Plan activo o inactivo'])
            ws_ref.append(['Plantilla', 'Sí, No', 'Es una plantilla reutilizable'])
            ws_ref.append(['Regla de actualización', 'Por Nombre del plan', 'El import actualiza solo por nombre del plan'])
            auto_adjust_columns(ws_ref)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="workout_plans_complete.xlsx"'
            response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            return response
        except Exception as e:
            import traceback
            return Response({'error': str(e), 'traceback': traceback.format_exc()}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Importa CSV de planes (formato simple o formato completo) sin IDs."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Archivo no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = file.read().decode('utf-8')
        except UnicodeDecodeError:
            return Response({'error': 'El archivo debe estar en formato UTF-8'}, status=status.HTTP_400_BAD_REQUEST)

        def parse_bool(value, default=False):
            if value is None or str(value).strip() == '':
                return default
            return str(value).strip().lower() in ['sí', 'si', 'true', '1', 'yes', 'y']

        def parse_list(value):
            if value is None:
                return []
            if isinstance(value, list):
                return value
            return [item.strip() for item in str(value).split(',') if item and item.strip()]

        def parse_int(value, default=None):
            if value is None or str(value).strip() == '':
                return default
            return int(value)

        difficulty_map = {
            'principiante': 'beginner',
            'intermedio': 'intermediate',
            'avanzado': 'advanced',
        }
        goal_map = {
            'pérdida de peso': 'weight_loss',
            'perdida de peso': 'weight_loss',
            'ganancia muscular': 'muscle_gain',
            'fuerza': 'strength',
            'resistencia': 'endurance',
            'fitness general': 'general_fitness',
            'recomposición corporal': 'body_recomposition',
            'recomposicion corporal': 'body_recomposition',
        }
        location_map = {
            'gimnasio': 'gym',
            'casa': 'home',
            'exterior': 'outdoor',
            'cualquier lugar': 'any',
        }
        day_map = {
            'lunes': 'monday',
            'martes': 'tuesday',
            'miércoles': 'wednesday',
            'miercoles': 'wednesday',
            'jueves': 'thursday',
            'viernes': 'friday',
            'sábado': 'saturday',
            'sabado': 'saturday',
            'domingo': 'sunday',
        }

        complete_aliases = {
            'plan_name': ['plan_name', 'nombre_plan'],
            'plan_description': ['plan_description', 'descripcion_plan', 'descripción_plan'],
            'plan_difficulty': ['plan_difficulty', 'dificultad_plan'],
            'plan_goal': ['plan_goal', 'objetivo_plan'],
            'plan_location': ['plan_location', 'ubicacion_plan', 'ubicación_plan'],
            'plan_duration_weeks': ['plan_duration_weeks', 'duracion_semanas_plan', 'duración_semanas_plan'],
            'plan_days_per_week': ['plan_days_per_week', 'dias_por_semana_plan', 'días_por_semana_plan'],
            'plan_estimated_duration_minutes': ['plan_estimated_duration_minutes', 'duracion_estimada_minutos_plan', 'duración_estimada_minutos_plan'],
            'plan_equipment_needed': ['plan_equipment_needed', 'equipo_necesario_plan'],
            'plan_tags': ['plan_tags', 'etiquetas_plan'],
            'plan_image_url': ['plan_image_url', 'imagen_url_plan'],
            'plan_is_active': ['plan_is_active', 'plan_activo'],
            'plan_is_template': ['plan_is_template', 'plan_plantilla'],
            'day_number': ['day_number', 'numero_dia', 'número_dia'],
            'day_name': ['day_name', 'nombre_dia', 'nombre_día'],
            'day_of_week': ['day_of_week', 'dia_semana', 'día_semana'],
            'day_is_rest_day': ['day_is_rest_day', 'dia_descanso', 'día_descanso'],
            'day_duration_minutes': ['day_duration_minutes', 'duracion_dia_minutos', 'duración_dia_minutos'],
            'day_focus': ['day_focus', 'enfoque_dia', 'enfoque_día'],
            'day_notes': ['day_notes', 'notas_dia', 'notas_día'],
            'day_order': ['day_order', 'orden_dia', 'orden_día'],
            'exercise_order': ['exercise_order', 'orden_ejercicio'],
            'exercise_name': ['exercise_name', 'nombre_ejercicio'],
            'sets': ['sets', 'series'],
            'reps': ['reps'],
            'weight': ['weight', 'peso'],
            'duration_seconds': ['duration_seconds', 'duracion_segundos', 'duración_segundos'],
            'rest_seconds': ['rest_seconds', 'descanso_segundos'],
            'exercise_notes': ['exercise_notes', 'notas_ejercicio'],
            'superset_group': ['superset_group', 'grupo_superset'],
            'substitutes': ['substitutes', 'sustitutos'],
        }

        simple_aliases = {
            'name': ['name', 'nombre'],
            'description': ['description', 'descripcion', 'descripción'],
            'difficulty': ['difficulty', 'dificultad'],
            'goal': ['goal', 'objetivo'],
            'location': ['location', 'ubicacion', 'ubicación'],
            'duration_weeks': ['duration_weeks', 'duracion_semanas', 'duración_semanas'],
            'days_per_week': ['days_per_week', 'dias_por_semana', 'días_por_semana'],
            'estimated_duration_minutes': ['estimated_duration_minutes', 'duracion_estimada_minutos', 'duración_estimada_minutos'],
            'equipment_needed': ['equipment_needed', 'equipo_necesario'],
            'tags': ['tags', 'etiquetas'],
            'image_url': ['image_url', 'imagen_url'],
            'is_active': ['is_active', 'activo'],
            'is_template': ['is_template', 'plantilla'],
        }

        def normalize_choice(value, translation_map):
            normalized = (value or '').strip().lower()
            return translation_map.get(normalized, (value or '').strip())

        def get_value(row, aliases, default=''):
            for key in aliases:
                if key in row and row.get(key) not in [None, '']:
                    return row.get(key)
            return default

        def row_has_content_dict(row):
            for value in row.values():
                if value is None:
                    continue
                if isinstance(value, str):
                    if value.strip() != '':
                        return True
                else:
                    return True
            return False

        reader = csv.DictReader(decoded.splitlines())
        created, updated, skipped = 0, 0, 0
        errors = []

        fieldnames = {str(name).strip().lower() for name in (reader.fieldnames or [])}
        is_complete_format = 'plan_name' in fieldnames or 'nombre_plan' in fieldnames

        if is_complete_format:
            from django.db import transaction

            stats = {
                'plans': {'created': 0, 'updated': 0, 'skipped': 0},
                'days': {'created': 0, 'updated': 0, 'skipped': 0},
                'exercises': {'created': 0, 'updated': 0, 'skipped': 0},
                'substitutes': {'created': 0, 'updated': 0, 'skipped': 0},
            }

            with transaction.atomic():
                for row_num, row in enumerate(reader, start=2):
                    try:
                        if not row_has_content_dict(row):
                            continue

                        plan_name = str(get_value(row, complete_aliases['plan_name'], '') or '').strip()
                        if not plan_name:
                            errors.append(f"Fila {row_num}: 'plan_name'/'nombre_plan' es requerido")
                            stats['plans']['skipped'] += 1
                            skipped += 1
                            continue

                        plan = WorkoutProgram.objects.filter(name=plan_name, is_template=True).first()
                        plan_fields = {
                            'name': plan_name,
                            'description': str(get_value(row, complete_aliases['plan_description'], '') or '').strip(),
                            'difficulty': normalize_choice(str(get_value(row, complete_aliases['plan_difficulty'], 'beginner') or 'beginner'), difficulty_map) or 'beginner',
                            'goal': normalize_choice(str(get_value(row, complete_aliases['plan_goal'], 'general_fitness') or 'general_fitness'), goal_map) or 'general_fitness',
                            'location': normalize_choice(str(get_value(row, complete_aliases['plan_location'], 'any') or 'any'), location_map) or 'any',
                            'duration_weeks': parse_int(get_value(row, complete_aliases['plan_duration_weeks']), default=4),
                            'days_per_week': parse_int(get_value(row, complete_aliases['plan_days_per_week']), default=3),
                            'estimated_duration_minutes': parse_int(get_value(row, complete_aliases['plan_estimated_duration_minutes']), default=60),
                            'equipment_needed': parse_list(get_value(row, complete_aliases['plan_equipment_needed'])),
                            'tags': parse_list(get_value(row, complete_aliases['plan_tags'])),
                            'image_url': str(get_value(row, complete_aliases['plan_image_url'], '') or '').strip(),
                            'is_active': parse_bool(get_value(row, complete_aliases['plan_is_active']), default=True),
                            'is_template': parse_bool(get_value(row, complete_aliases['plan_is_template']), default=True),
                            'is_system': False,
                        }

                        if plan:
                            for key, value in plan_fields.items():
                                setattr(plan, key, value)
                            plan.save()
                            stats['plans']['updated'] += 1
                        else:
                            plan = WorkoutProgram.objects.create(**plan_fields)
                            stats['plans']['created'] += 1

                        day_number = parse_int(get_value(row, complete_aliases['day_number']), default=None)
                        if day_number is None:
                            continue

                        day = WorkoutDay.objects.filter(program=plan, day_number=day_number).first()
                        day_fields = {
                            'program': plan,
                            'day_number': day_number,
                            'name': (str(get_value(row, complete_aliases['day_name'], f'Día {day_number}') or f'Día {day_number}').strip()) or f'Día {day_number}',
                            'day_of_week': normalize_choice(str(get_value(row, complete_aliases['day_of_week'], '') or '').strip(), day_map),
                            'is_rest_day': parse_bool(get_value(row, complete_aliases['day_is_rest_day']), default=False),
                            'duration_minutes': parse_int(get_value(row, complete_aliases['day_duration_minutes']), default=None),
                            'focus': str(get_value(row, complete_aliases['day_focus'], '') or '').strip(),
                            'notes': str(get_value(row, complete_aliases['day_notes'], '') or '').strip(),
                            'order_index': parse_int(get_value(row, complete_aliases['day_order']), default=day_number),
                        }

                        if day:
                            for key, value in day_fields.items():
                                setattr(day, key, value)
                            day.save()
                            stats['days']['updated'] += 1
                        else:
                            day = WorkoutDay.objects.create(**day_fields)
                            stats['days']['created'] += 1

                        exercise_name = str(get_value(row, complete_aliases['exercise_name'], '') or '').strip()
                        if not exercise_name:
                            continue

                        exercise = Exercise.objects.filter(name=exercise_name).first()
                        if not exercise:
                            errors.append(f"Fila {row_num}: Ejercicio '{exercise_name}' no encontrado")
                            stats['exercises']['skipped'] += 1
                            skipped += 1
                            continue

                        exercise_order = parse_int(get_value(row, complete_aliases['exercise_order']), default=1)
                        workout_exercise = WorkoutDayExercise.objects.filter(
                            workout_day=day,
                            order_index=exercise_order,
                        ).first()

                        exercise_fields = {
                            'workout_day': day,
                            'exercise': exercise,
                            'sets': parse_int(get_value(row, complete_aliases['sets']), default=3),
                            'reps': (str(get_value(row, complete_aliases['reps'], '10') or '10').strip()) or '10',
                            'weight': str(get_value(row, complete_aliases['weight'], '') or '').strip(),
                            'duration_seconds': parse_int(get_value(row, complete_aliases['duration_seconds']), default=None),
                            'rest_seconds': parse_int(get_value(row, complete_aliases['rest_seconds']), default=60),
                            'notes': str(get_value(row, complete_aliases['exercise_notes'], '') or '').strip(),
                            'order_index': exercise_order,
                            'superset_group': parse_int(get_value(row, complete_aliases['superset_group']), default=None),
                        }

                        if workout_exercise:
                            for key, value in exercise_fields.items():
                                setattr(workout_exercise, key, value)
                            workout_exercise.save()
                            stats['exercises']['updated'] += 1
                        else:
                            WorkoutDayExercise.objects.create(**exercise_fields)
                            stats['exercises']['created'] += 1

                        substitutes_raw = str(get_value(row, complete_aliases['substitutes'], '') or '').strip()
                        if substitutes_raw:
                            substitute_names = [name.strip() for name in substitutes_raw.split('|') if name and name.strip()]
                            for index, substitute_name in enumerate(substitute_names, start=1):
                                substitute = Exercise.objects.filter(name=substitute_name).first()
                                if not substitute:
                                    errors.append(f"Fila {row_num}: Sustituto '{substitute_name}' no encontrado")
                                    stats['substitutes']['skipped'] += 1
                                    skipped += 1
                                    continue

                                relation = ExerciseSubstitution.objects.filter(
                                    exercise=exercise,
                                    substitute=substitute,
                                ).first()
                                if relation:
                                    relation.priority = index
                                    relation.save()
                                    stats['substitutes']['updated'] += 1
                                else:
                                    ExerciseSubstitution.objects.create(
                                        exercise=exercise,
                                        substitute=substitute,
                                        priority=index,
                                    )
                                    stats['substitutes']['created'] += 1

                    except Exception as e:
                        errors.append(f"Fila {row_num}: {str(e)}")
                        skipped += 1

            return Response({
                'message': 'Importación CSV completa finalizada',
                'stats': stats,
                'errors': errors if errors else None
            })

        for row_num, row in enumerate(reader, start=2):
            try:
                if not row_has_content_dict(row):
                    continue

                name = str(get_value(row, simple_aliases['name'], '') or '').strip()
                if not name:
                    errors.append(f"Fila {row_num}: 'name'/'nombre' es requerido")
                    skipped += 1
                    continue

                plan = WorkoutProgram.objects.filter(name=name, is_template=True).first()
                fields = {
                    'description': str(get_value(row, simple_aliases['description'], '') or '').strip(),
                    'difficulty': normalize_choice(str(get_value(row, simple_aliases['difficulty'], 'beginner') or 'beginner'), difficulty_map) or 'beginner',
                    'goal': normalize_choice(str(get_value(row, simple_aliases['goal'], 'general_fitness') or 'general_fitness'), goal_map) or 'general_fitness',
                    'location': normalize_choice(str(get_value(row, simple_aliases['location'], 'any') or 'any'), location_map) or 'any',
                    'duration_weeks': parse_int(get_value(row, simple_aliases['duration_weeks']), default=4),
                    'days_per_week': parse_int(get_value(row, simple_aliases['days_per_week']), default=3),
                    'estimated_duration_minutes': parse_int(get_value(row, simple_aliases['estimated_duration_minutes']), default=60),
                    'equipment_needed': parse_list(get_value(row, simple_aliases['equipment_needed'])),
                    'tags': parse_list(get_value(row, simple_aliases['tags'])),
                    'image_url': str(get_value(row, simple_aliases['image_url'], '') or '').strip(),
                    'is_active': parse_bool(get_value(row, simple_aliases['is_active'], 'true'), default=True),
                    'is_template': parse_bool(get_value(row, simple_aliases['is_template'], 'true'), default=True),
                    'is_system': False,
                }

                if plan:
                    for k, v in fields.items():
                        setattr(plan, k, v)
                    plan.save()
                    updated += 1
                else:
                    WorkoutProgram.objects.create(name=name, **fields)
                    created += 1
            except Exception as e:
                errors.append(f"Fila {row_num}: {str(e)}")
                skipped += 1

        message = f"Importación completada. {created} planes creados, {updated} actualizados"
        if skipped > 0:
            message += f", {skipped} omitidos"

        return Response({
            'message': message,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors if errors else None
        })

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Importa planes completos desde Excel sin IDs, actualizando por nombre de plan."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Archivo no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            
            wb = load_workbook(file)
            stats = {
                'plans': {'created': 0, 'updated': 0, 'skipped': 0},
                'days': {'created': 0, 'updated': 0, 'skipped': 0, 'deleted': 0},
                'exercises': {'created': 0, 'updated': 0, 'skipped': 0, 'deleted': 0},
                'substitutes': {'created': 0, 'updated': 0, 'skipped': 0},
            }
            errors = []
            imported_day_ids_by_plan = {}
            imported_exercise_ids_by_day = {}
            
            def parse_bool(value, default=False):
                if value is None or str(value).strip() == '':
                    return default
                return str(value).strip().lower() in ['sí', 'si', 'true', '1', 'yes', 'y']

            def parse_list(value):
                if value is None:
                    return []
                if isinstance(value, list):
                    return value
                return [item.strip() for item in str(value).split(',') if item and item.strip()]

            def parse_int(value, default=None):
                if value is None:
                    return default
                text = str(value).strip()
                if text == '':
                    return default
                try:
                    return int(text)
                except (TypeError, ValueError):
                    try:
                        return int(float(text.replace(',', '.')))
                    except (TypeError, ValueError):
                        return default

            difficulty_map = {
                'principiante': 'beginner',
                'intermedio': 'intermediate',
                'avanzado': 'advanced',
            }
            goal_map = {
                'pérdida de peso': 'weight_loss',
                'perdida de peso': 'weight_loss',
                'ganancia muscular': 'muscle_gain',
                'fuerza': 'strength',
                'resistencia': 'endurance',
                'fitness general': 'general_fitness',
                'recomposición corporal': 'body_recomposition',
                'recomposicion corporal': 'body_recomposition',
            }
            location_map = {
                'gimnasio': 'gym',
                'casa': 'home',
                'exterior': 'outdoor',
                'cualquier lugar': 'any',
            }
            day_map = {
                'lunes': 'monday',
                'martes': 'tuesday',
                'miércoles': 'wednesday',
                'miercoles': 'wednesday',
                'jueves': 'thursday',
                'viernes': 'friday',
                'sábado': 'saturday',
                'sabado': 'saturday',
                'domingo': 'sunday',
            }

            def normalize_choice(value, translation_map):
                normalized = str(value).strip().lower() if value is not None else ''
                return translation_map.get(normalized, str(value).strip() if value is not None else '')

            def normalize_text(value):
                import unicodedata
                text = str(value or '').strip().lower()
                text = unicodedata.normalize('NFKD', text)
                text = ''.join(ch for ch in text if not unicodedata.combining(ch))
                return ' '.join(text.split())

            def normalize_header(value):
                import unicodedata
                text = str(value or '').strip().lower()
                text = unicodedata.normalize('NFKD', text)
                text = ''.join(ch for ch in text if not unicodedata.combining(ch))
                return ''.join(ch for ch in text if ch.isalnum())

            def build_header_index(ws):
                header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                return {normalize_header(name): idx for idx, name in enumerate(header) if name is not None and str(name).strip() != ''}

            def get_cell(row, header_index, aliases, fallback_index=None, default=None):
                for alias in aliases:
                    idx = header_index.get(normalize_header(alias))
                    if idx is not None and idx < len(row):
                        value = row[idx]
                        if value not in [None, '']:
                            return value
                if fallback_index is not None and fallback_index < len(row):
                    value = row[fallback_index]
                    if value not in [None, '']:
                        return value
                return default

            def get_sheet(*aliases):
                wanted = {normalize_header(alias) for alias in aliases}
                for sheet_name in wb.sheetnames:
                    if normalize_header(sheet_name) in wanted:
                        return wb[sheet_name]
                return None

            def row_has_content_tuple(row):
                for value in row:
                    if value is None:
                        continue
                    if isinstance(value, str):
                        if value.strip() != '':
                            return True
                    else:
                        return True
                return False

            User = get_user_model()
            users_by_email = {
                str(user.email).strip().lower(): user
                for user in User.objects.filter(email__isnull=False).exclude(email='')
            }

            exercises_cache = {
                normalize_text(exercise.name): exercise
                for exercise in Exercise.objects.all().only('id', 'name')
            }

            def resolve_user(email_value):
                email = str(email_value or '').strip().lower()
                if not email:
                    return None
                return users_by_email.get(email)

            def is_user_category(category, user_email=None):
                return normalize_text(category) == 'usuario' or bool(str(user_email or '').strip())

            def deactivate_other_active_user_plans(plan):
                if not getattr(plan, 'user_id', None) or not getattr(plan, 'is_active', False):
                    return
                from django.utils import timezone

                WorkoutProgram.objects.filter(
                    user=plan.user,
                    is_active=True,
                ).exclude(pk=plan.pk).update(
                    is_active=False,
                    end_date=timezone.now().date(),
                )

            def resolve_plan(name, category=None, user_email=None):
                if not name:
                    return None

                qs = WorkoutProgram.objects.filter(name=name)
                user = resolve_user(user_email)
                category_key = normalize_text(category)

                if user:
                    by_user = qs.filter(user=user).first()
                    if by_user:
                        return by_user
                    if is_user_category(category, user_email):
                        return None

                if category_key == 'sistema':
                    system_plan = qs.filter(is_system=True).first()
                    if system_plan:
                        return system_plan
                elif category_key == 'usuario':
                    user_plan = qs.exclude(user__isnull=True).first()
                    if user_plan:
                        return user_plan
                elif category_key == 'plantilla':
                    template_plan = qs.filter(is_template=True).first()
                    if template_plan:
                        return template_plan

                if user:
                    same_name_user_null = qs.filter(user__isnull=True).first()
                    if same_name_user_null:
                        return same_name_user_null

                return qs.first()

            def get_or_create_referenced_plan(plan_name, category=None, user_email=None):
                plan = resolve_plan(plan_name, category=category, user_email=user_email)
                if plan:
                    return plan, False

                category_key = normalize_text(category)
                user = resolve_user(user_email)
                if is_user_category(category, user_email) and not user:
                    raise ValueError(f"Usuario '{user_email}' no encontrado para el plan '{plan_name}'")

                is_system = category_key == 'sistema'
                is_template = category_key == 'plantilla' and not user

                plan = WorkoutProgram.objects.create(
                    name=plan_name,
                    description='',
                    difficulty='beginner',
                    goal='general_fitness',
                    location='any',
                    duration_weeks=4,
                    days_per_week=3,
                    estimated_duration_minutes=60,
                    equipment_needed=[],
                    tags=[],
                    image_url='',
                    is_active=True,
                    is_template=is_template,
                    is_system=is_system and not user,
                    user=user,
                )
                deactivate_other_active_user_plans(plan)
                return plan, True

            def find_exercise_by_name(name):
                if not name:
                    return None
                exercise = Exercise.objects.filter(name=name).first()
                if exercise:
                    return exercise
                return exercises_cache.get(normalize_text(name))

            def remember_imported_day(day):
                imported_day_ids_by_plan.setdefault(day.program_id, set()).add(day.id)
                imported_exercise_ids_by_day.setdefault(day.id, set())

            def remember_imported_exercise(workout_exercise):
                remember_imported_day(workout_exercise.workout_day)
                imported_exercise_ids_by_day.setdefault(workout_exercise.workout_day_id, set()).add(workout_exercise.id)

            with transaction.atomic():
                # === IMPORTAR PLANES ===
                ws_plans = get_sheet('Planes')
                if ws_plans:
                    plans_headers = build_header_index(ws_plans)

                    for row_num, row in enumerate(ws_plans.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row_has_content_tuple(row):
                                continue

                            name = str(get_cell(row, plans_headers, ['Nombre', 'name', 'plan_name', 'nombre_plan'], fallback_index=3, default='') or '').strip()
                            if not name:
                                continue

                            category = str(get_cell(row, plans_headers, ['Categoría', 'Categoria', 'categoria_plan'], fallback_index=0, default='') or '').strip()
                            user_ref = str(get_cell(row, plans_headers, ['Usuario Referencia', 'usuario_referencia'], fallback_index=1, default='') or '').strip()
                            created_by_ref = str(get_cell(row, plans_headers, ['Creado Por', 'creado_por'], fallback_index=2, default='') or '').strip()

                            plan = resolve_plan(name, category=category, user_email=user_ref)
                            category_key = normalize_text(category)
                            user = resolve_user(user_ref)
                            if is_user_category(category, user_ref) and not user:
                                errors.append(f"Planes - Fila {row_num}: Usuario '{user_ref}' no encontrado")
                                stats['plans']['skipped'] += 1
                                continue
                            plan_is_active = parse_bool(get_cell(row, plans_headers, ['Activo', 'is_active', 'plan_is_active'], fallback_index=14), default=True)

                            fields = {
                                'name': name,
                                'description': str(get_cell(row, plans_headers, ['Descripción', 'Descripcion', 'description', 'plan_description'], fallback_index=4, default='') or '').strip(),
                                'difficulty': normalize_choice(str(get_cell(row, plans_headers, ['Dificultad', 'difficulty', 'plan_difficulty'], fallback_index=5, default='beginner') or 'beginner').strip(), difficulty_map) or 'beginner',
                                'goal': normalize_choice(str(get_cell(row, plans_headers, ['Objetivo', 'goal', 'plan_goal'], fallback_index=6, default='general_fitness') or 'general_fitness').strip(), goal_map) or 'general_fitness',
                                'location': normalize_choice(str(get_cell(row, plans_headers, ['Ubicación', 'Ubicacion', 'location', 'plan_location'], fallback_index=7, default='any') or 'any').strip(), location_map) or 'any',
                                'duration_weeks': parse_int(get_cell(row, plans_headers, ['Semanas', 'duration_weeks', 'plan_duration_weeks'], fallback_index=8), default=4),
                                'days_per_week': parse_int(get_cell(row, plans_headers, ['Días/Semana', 'Dias/Semana', 'days_per_week', 'plan_days_per_week'], fallback_index=9), default=3),
                                'estimated_duration_minutes': parse_int(get_cell(row, plans_headers, ['Duración (min)', 'Duracion (min)', 'estimated_duration_minutes', 'plan_estimated_duration_minutes'], fallback_index=10), default=60),
                                'equipment_needed': parse_list(get_cell(row, plans_headers, ['Equipo (coma separado)', 'equipment_needed', 'plan_equipment_needed'], fallback_index=11)),
                                'tags': parse_list(get_cell(row, plans_headers, ['Tags (coma separado)', 'tags', 'plan_tags'], fallback_index=12)),
                                'image_url': str(get_cell(row, plans_headers, ['Imagen URL', 'image_url', 'plan_image_url'], fallback_index=13, default='') or '').strip(),
                                'is_active': plan_is_active,
                                'is_template': False if user else parse_bool(get_cell(row, plans_headers, ['Plantilla', 'is_template', 'plan_is_template'], fallback_index=15), default=(category_key == 'plantilla')),
                                'is_system': False if user else category_key == 'sistema',
                                'user': user,
                                'created_by': resolve_user(created_by_ref),
                            }

                            if plan:
                                for k, v in fields.items():
                                    setattr(plan, k, v)
                                plan.save()
                                deactivate_other_active_user_plans(plan)
                                stats['plans']['updated'] += 1
                            else:
                                plan = WorkoutProgram.objects.create(**fields)
                                deactivate_other_active_user_plans(plan)
                                stats['plans']['created'] += 1

                        except Exception as e:
                            errors.append(f"Planes - Fila {row_num}: {str(e)}")
                            stats['plans']['skipped'] += 1

                # === IMPORTAR DÍAS (match por Nombre Plan + Número Día) ===
                ws_days = get_sheet('Días', 'Dias')
                if ws_days:
                    days_headers = build_header_index(ws_days)

                    for row_num, row in enumerate(ws_days.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row_has_content_tuple(row):
                                continue

                            plan_name = str(get_cell(row, days_headers, ['Nombre Plan', 'plan_name'], fallback_index=2, default='') or '').strip()
                            day_number = parse_int(get_cell(row, days_headers, ['Número Día', 'Numero Día', 'day_number'], fallback_index=3), default=None)
                            if not plan_name or day_number is None:
                                continue

                            name = str(get_cell(row, days_headers, ['Nombre Día', 'Nombre Dia', 'day_name'], fallback_index=4, default=f'Día {day_number}') or f'Día {day_number}').strip()
                            category = str(get_cell(row, days_headers, ['Categoría Plan', 'Categoria Plan', 'categoria_plan'], fallback_index=0, default='') or '').strip()
                            user_ref = str(get_cell(row, days_headers, ['Usuario Referencia', 'usuario_referencia'], fallback_index=1, default='') or '').strip()

                            plan, created_from_reference = get_or_create_referenced_plan(plan_name, category=category, user_email=user_ref)
                            if created_from_reference:
                                stats['plans']['created'] += 1

                            day = WorkoutDay.objects.filter(program=plan, day_number=day_number).first()

                            fields = {
                                'program': plan,
                                'name': name,
                                'day_number': day_number,
                                'day_of_week': normalize_choice(str(get_cell(row, days_headers, ['Día Semana', 'Dia Semana', 'day_of_week'], fallback_index=5, default='') or '').strip(), day_map),
                                'is_rest_day': parse_bool(get_cell(row, days_headers, ['Es Descanso', 'day_is_rest_day'], fallback_index=6), default=False),
                                'duration_minutes': parse_int(get_cell(row, days_headers, ['Duración (min)', 'Duracion (min)', 'day_duration_minutes'], fallback_index=7), default=None),
                                'focus': str(get_cell(row, days_headers, ['Enfoque', 'day_focus'], fallback_index=8, default='') or '').strip(),
                                'notes': str(get_cell(row, days_headers, ['Notas', 'day_notes'], fallback_index=9, default='') or '').strip(),
                                'order_index': parse_int(get_cell(row, days_headers, ['Orden', 'day_order'], fallback_index=10), default=day_number),
                            }

                            if day:
                                for k, v in fields.items():
                                    setattr(day, k, v)
                                day.save()
                                stats['days']['updated'] += 1
                            else:
                                day = WorkoutDay.objects.create(**fields)
                                stats['days']['created'] += 1
                            remember_imported_day(day)

                        except Exception as e:
                            errors.append(f"Días - Fila {row_num}: {str(e)}")
                            stats['days']['skipped'] += 1

                # === IMPORTAR EJERCICIOS (match por Plan + Día + Orden) ===
                ws_ex = get_sheet('Ejercicios')
                if ws_ex:
                    exercises_headers = build_header_index(ws_ex)

                    for row_num, row in enumerate(ws_ex.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row_has_content_tuple(row):
                                continue

                            plan_name = str(get_cell(row, exercises_headers, ['Nombre Plan', 'plan_name'], fallback_index=2, default='') or '').strip()
                            day_number = parse_int(get_cell(row, exercises_headers, ['Número Día', 'Numero Día', 'day_number'], fallback_index=3), default=None)
                            order_index = parse_int(get_cell(row, exercises_headers, ['Orden', 'exercise_order'], fallback_index=5), default=1)
                            exercise_name = str(get_cell(row, exercises_headers, ['Nombre Ejercicio', 'exercise_name'], fallback_index=6, default='') or '').strip()

                            if not plan_name or day_number is None:
                                errors.append(f"Ejercicios - Fila {row_num}: Plan, Número Día y Nombre Ejercicio son requeridos")
                                stats['exercises']['skipped'] += 1
                                continue

                            # En filas de descanso puede no haber ejercicio; no se considera error.
                            if not exercise_name:
                                continue

                            category = str(get_cell(row, exercises_headers, ['Categoría Plan', 'Categoria Plan', 'categoria_plan'], fallback_index=0, default='') or '').strip()
                            user_ref = str(get_cell(row, exercises_headers, ['Usuario Referencia', 'usuario_referencia'], fallback_index=1, default='') or '').strip()

                            plan, created_from_reference = get_or_create_referenced_plan(plan_name, category=category, user_email=user_ref)
                            if created_from_reference:
                                stats['plans']['created'] += 1

                            day = WorkoutDay.objects.filter(program=plan, day_number=day_number).first()
                            if not day:
                                day = WorkoutDay.objects.create(
                                    program=plan,
                                    name=f'Día {day_number}',
                                    day_number=day_number,
                                    day_of_week='',
                                    is_rest_day=False,
                                    duration_minutes=None,
                                    focus='',
                                    notes='',
                                    order_index=day_number,
                                )
                                stats['days']['created'] += 1

                            exercise = find_exercise_by_name(exercise_name)
                            if not exercise:
                                errors.append(f"Ejercicios - Fila {row_num}: Ejercicio '{exercise_name}' no encontrado")
                                stats['exercises']['skipped'] += 1
                                continue

                            workout_ex = WorkoutDayExercise.objects.filter(
                                workout_day=day,
                                order_index=order_index,
                            ).first()

                            fields = {
                                'workout_day': day,
                                'exercise': exercise,
                                'sets': parse_int(get_cell(row, exercises_headers, ['Series', 'sets'], fallback_index=7), default=3),
                                'reps': str(get_cell(row, exercises_headers, ['Reps', 'reps'], fallback_index=8, default='10') or '10').strip() or '10',
                                'weight': str(get_cell(row, exercises_headers, ['Peso', 'weight'], fallback_index=9, default='') or '').strip(),
                                'duration_seconds': parse_int(get_cell(row, exercises_headers, ['Duración (seg)', 'Duracion (seg)', 'duration_seconds'], fallback_index=10), default=None),
                                'rest_seconds': parse_int(get_cell(row, exercises_headers, ['Descanso (seg)', 'rest_seconds'], fallback_index=11), default=60),
                                'notes': str(get_cell(row, exercises_headers, ['Notas', 'exercise_notes'], fallback_index=12, default='') or '').strip(),
                                'order_index': order_index,
                                'superset_group': parse_int(get_cell(row, exercises_headers, ['Grupo Superset', 'superset_group'], fallback_index=13), default=None),
                            }

                            if workout_ex:
                                for k, v in fields.items():
                                    setattr(workout_ex, k, v)
                                workout_ex.save()
                                stats['exercises']['updated'] += 1
                            else:
                                workout_ex = WorkoutDayExercise.objects.create(**fields)
                                stats['exercises']['created'] += 1
                            remember_imported_exercise(workout_ex)

                        except Exception as e:
                            errors.append(f"Ejercicios - Fila {row_num}: {str(e)}")
                            stats['exercises']['skipped'] += 1

                # === IMPORTAR RESUMEN COMPLETO ===
                # Muchos clientes editan la primera hoja exportada. Procesarla al final
                # permite que sus cambios prevalezcan sobre las hojas detalladas.
                ws_summary = get_sheet('Resumen completo', 'Resumen')
                if ws_summary:
                    summary_headers = build_header_index(ws_summary)

                    for row_num, row in enumerate(ws_summary.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row_has_content_tuple(row):
                                continue

                            plan_name = str(get_cell(row, summary_headers, ['Nombre Plan', 'plan_name'], fallback_index=3, default='') or '').strip()
                            if not plan_name:
                                continue

                            category = str(get_cell(row, summary_headers, ['Categoría Plan', 'Categoria Plan', 'Categoría', 'Categoria', 'categoria_plan'], fallback_index=0, default='') or '').strip()
                            user_ref = str(get_cell(row, summary_headers, ['Usuario Referencia', 'usuario_referencia'], fallback_index=1, default='') or '').strip()
                            created_by_ref = str(get_cell(row, summary_headers, ['Creado Por', 'creado_por'], fallback_index=2, default='') or '').strip()
                            user = resolve_user(user_ref)
                            if is_user_category(category, user_ref) and not user:
                                errors.append(f"Resumen completo - Fila {row_num}: Usuario '{user_ref}' no encontrado")
                                stats['plans']['skipped'] += 1
                                continue

                            plan, created_from_reference = get_or_create_referenced_plan(plan_name, category=category, user_email=user_ref)
                            if created_from_reference:
                                stats['plans']['created'] += 1

                            plan_fields = {
                                'name': plan_name,
                                'description': str(get_cell(row, summary_headers, ['Descripción Plan', 'Descripcion Plan', 'description', 'plan_description'], fallback_index=4, default='') or '').strip(),
                                'difficulty': normalize_choice(str(get_cell(row, summary_headers, ['Dificultad', 'difficulty', 'plan_difficulty'], fallback_index=5, default=plan.difficulty or 'beginner') or plan.difficulty or 'beginner').strip(), difficulty_map) or 'beginner',
                                'goal': normalize_choice(str(get_cell(row, summary_headers, ['Objetivo', 'goal', 'plan_goal'], fallback_index=6, default=plan.goal or 'general_fitness') or plan.goal or 'general_fitness').strip(), goal_map) or 'general_fitness',
                                'location': normalize_choice(str(get_cell(row, summary_headers, ['Ubicación', 'Ubicacion', 'location', 'plan_location'], fallback_index=7, default=plan.location or 'any') or plan.location or 'any').strip(), location_map) or 'any',
                                'duration_weeks': parse_int(get_cell(row, summary_headers, ['Semanas', 'duration_weeks', 'plan_duration_weeks'], fallback_index=8), default=plan.duration_weeks or 4),
                                'days_per_week': parse_int(get_cell(row, summary_headers, ['Días/Semana', 'Dias/Semana', 'days_per_week', 'plan_days_per_week'], fallback_index=9), default=plan.days_per_week or 3),
                                'estimated_duration_minutes': parse_int(get_cell(row, summary_headers, ['Duración Plan (min)', 'Duracion Plan (min)', 'Duración (min)', 'Duracion (min)', 'estimated_duration_minutes', 'plan_estimated_duration_minutes'], fallback_index=10), default=plan.estimated_duration_minutes or 60),
                                'is_active': True,
                                'is_template': False if user else (normalize_text(category) == 'plantilla'),
                                'is_system': False if user else (normalize_text(category) == 'sistema'),
                                'user': user,
                                'created_by': resolve_user(created_by_ref),
                            }
                            for k, v in plan_fields.items():
                                setattr(plan, k, v)
                            plan.save()
                            deactivate_other_active_user_plans(plan)
                            stats['plans']['updated'] += 1

                            day_number = parse_int(get_cell(row, summary_headers, ['Número Día', 'Numero Día', 'day_number'], fallback_index=11), default=None)
                            if day_number is None:
                                continue

                            day = WorkoutDay.objects.filter(program=plan, day_number=day_number).first()
                            day_fields = {
                                'program': plan,
                                'name': str(get_cell(row, summary_headers, ['Nombre Día', 'Nombre Dia', 'day_name'], fallback_index=12, default=f'Día {day_number}') or f'Día {day_number}').strip(),
                                'day_number': day_number,
                                'day_of_week': normalize_choice(str(get_cell(row, summary_headers, ['Día Semana', 'Dia Semana', 'day_of_week'], fallback_index=13, default='') or '').strip(), day_map),
                                'is_rest_day': parse_bool(get_cell(row, summary_headers, ['Descanso', 'Es Descanso', 'day_is_rest_day'], fallback_index=14), default=False),
                                'duration_minutes': None,
                                'focus': '',
                                'notes': '',
                                'order_index': day_number,
                            }
                            if day:
                                for k, v in day_fields.items():
                                    setattr(day, k, v)
                                day.save()
                                stats['days']['updated'] += 1
                            else:
                                day = WorkoutDay.objects.create(**day_fields)
                                stats['days']['created'] += 1
                            remember_imported_day(day)

                            exercise_name = str(get_cell(row, summary_headers, ['Nombre Ejercicio', 'exercise_name'], fallback_index=16, default='') or '').strip()
                            if not exercise_name:
                                continue

                            exercise = find_exercise_by_name(exercise_name)
                            if not exercise:
                                errors.append(f"Resumen completo - Fila {row_num}: Ejercicio '{exercise_name}' no encontrado")
                                stats['exercises']['skipped'] += 1
                                continue

                            order_index = parse_int(get_cell(row, summary_headers, ['Orden Ejercicio', 'Orden', 'exercise_order'], fallback_index=15), default=1)
                            workout_ex = WorkoutDayExercise.objects.filter(
                                workout_day=day,
                                order_index=order_index,
                            ).first()
                            exercise_fields = {
                                'workout_day': day,
                                'exercise': exercise,
                                'sets': parse_int(get_cell(row, summary_headers, ['Series', 'sets'], fallback_index=17), default=3),
                                'reps': str(get_cell(row, summary_headers, ['Reps', 'reps'], fallback_index=18, default='10') or '10').strip() or '10',
                                'weight': str(get_cell(row, summary_headers, ['Peso', 'weight'], fallback_index=19, default='') or '').strip(),
                                'duration_seconds': parse_int(get_cell(row, summary_headers, ['Duración (seg)', 'Duracion (seg)', 'duration_seconds'], fallback_index=20), default=None),
                                'rest_seconds': parse_int(get_cell(row, summary_headers, ['Descanso (seg)', 'rest_seconds'], fallback_index=21), default=60),
                                'notes': '',
                                'order_index': order_index,
                                'superset_group': parse_int(get_cell(row, summary_headers, ['Grupo Superset', 'superset_group'], fallback_index=22), default=None),
                            }
                            if workout_ex:
                                for k, v in exercise_fields.items():
                                    setattr(workout_ex, k, v)
                                workout_ex.save()
                                stats['exercises']['updated'] += 1
                            else:
                                workout_ex = WorkoutDayExercise.objects.create(**exercise_fields)
                                stats['exercises']['created'] += 1
                            remember_imported_exercise(workout_ex)

                        except Exception as e:
                            errors.append(f"Resumen completo - Fila {row_num}: {str(e)}")
                            stats['exercises']['skipped'] += 1

                for day_id, imported_exercise_ids in imported_exercise_ids_by_day.items():
                    stale_exercises = WorkoutDayExercise.objects.filter(workout_day_id=day_id)
                    if imported_exercise_ids:
                        stale_exercises = stale_exercises.exclude(id__in=imported_exercise_ids)
                    deleted_count, _ = stale_exercises.delete()
                    stats['exercises']['deleted'] += deleted_count

                for plan_id, imported_day_ids in imported_day_ids_by_plan.items():
                    stale_days = WorkoutDay.objects.filter(program_id=plan_id)
                    if imported_day_ids:
                        stale_days = stale_days.exclude(id__in=imported_day_ids)
                    deleted_count, deleted_details = stale_days.delete()
                    stats['days']['deleted'] += deleted_details.get('workouts.WorkoutDay', 0)
                    stats['exercises']['deleted'] += deleted_details.get('workouts.WorkoutDayExercise', 0)

                # === IMPORTAR SUSTITUTOS (match por nombre ejercicio) ===
                ws_subs = get_sheet('Sustitutos')
                if ws_subs:
                    substitutes_headers = build_header_index(ws_subs)

                    for row_num, row in enumerate(ws_subs.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row_has_content_tuple(row):
                                continue

                            exercise_name = str(get_cell(row, substitutes_headers, ['Nombre Ejercicio', 'exercise_name'], fallback_index=0, default='') or '').strip()
                            substitute_name = str(get_cell(row, substitutes_headers, ['Nombre Sustituto', 'substitute_name'], fallback_index=1, default='') or '').strip()

                            if not exercise_name or not substitute_name:
                                errors.append(f"Sustitutos - Fila {row_num}: Nombre Ejercicio y Nombre Sustituto son requeridos")
                                stats['substitutes']['skipped'] += 1
                                continue

                            exercise = Exercise.objects.filter(name=exercise_name).first()
                            if not exercise:
                                errors.append(f"Sustitutos - Fila {row_num}: Ejercicio '{exercise_name}' no encontrado")
                                stats['substitutes']['skipped'] += 1
                                continue

                            substitute = Exercise.objects.filter(name=substitute_name).first()
                            if not substitute:
                                errors.append(f"Sustitutos - Fila {row_num}: Sustituto '{substitute_name}' no encontrado")
                                stats['substitutes']['skipped'] += 1
                                continue

                            sub_relation = ExerciseSubstitution.objects.filter(
                                exercise=exercise,
                                substitute=substitute
                            ).first()

                            fields = {
                                'exercise': exercise,
                                'substitute': substitute,
                                'priority': parse_int(get_cell(row, substitutes_headers, ['Prioridad', 'priority'], fallback_index=2), default=1),
                                'notes': str(get_cell(row, substitutes_headers, ['Notas', 'notes'], fallback_index=3, default='') or '').strip(),
                            }

                            if sub_relation:
                                sub_relation.priority = fields['priority']
                                sub_relation.notes = fields['notes']
                                sub_relation.save()
                                stats['substitutes']['updated'] += 1
                            else:
                                ExerciseSubstitution.objects.create(**fields)
                                stats['substitutes']['created'] += 1

                        except Exception as e:
                            errors.append(f"Sustitutos - Fila {row_num}: {str(e)}")
                            stats['substitutes']['skipped'] += 1

            return Response({
                'message': 'Importación completada exitosamente',
                'stats': stats,
                'errors': errors if errors else None
            })
            
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc()
            }, status=status.HTTP_400_BAD_REQUEST)
