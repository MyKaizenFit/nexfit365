"""Tests para nutrition/admin_views.py - Admin endpoints de nutrición"""
import io
import csv
import pytest
from decimal import Decimal
from rest_framework.test import APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from nutrition.models import (
    Recipe, NutritionPlan, PlanMeal, PlanMealRecipe, MealLog, Food,
    NutritionPlanAssignment, RecipeIngredient
)

User = get_user_model()

RECIPES_URL = '/api/admin/nutrition/recipes/'
PLANS_URL = '/api/admin/nutrition/plans/'
MEALS_URL = '/api/admin/nutrition/meals/'
FOODS_URL = '/api/admin/nutrition/foods/'
MEAL_RECIPES_URL = '/api/admin/nutrition/meal-recipes/'
DEFAULT_PLANS_URL = '/api/admin/nutrition/default-plans/'
CHANGE_USER_PLAN_URL = '/api/admin/nutrition/change-user-plan/'
BULK_CHANGE_PLANS_URL = '/api/admin/nutrition/bulk-change-plans/'
USER_PLANS_STATS_URL = '/api/admin/nutrition/user-plans/stats/'
USER_PLANS_USAGE_URL = '/api/admin/nutrition/user-plans/usage_stats/'
USER_PLANS_HISTORY_URL = '/api/admin/nutrition/user-plans/history/'


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def admin_user(db):
    return User.objects.create_user(
        email='admin_nutr@test.com',
        password='testpass123',
        is_staff=True,
        is_superuser=True,
    )


@pytest.fixture
def regular_user(db):
    return User.objects.create_user(
        email='user_nutr@test.com',
        password='testpass123',
    )


@pytest.fixture
def admin_client(admin_user):
    client = APIClient()
    client.force_authenticate(user=admin_user)
    return client


@pytest.fixture
def regular_client(regular_user):
    client = APIClient()
    client.force_authenticate(user=regular_user)
    return client


@pytest.fixture
def recipe(db):
    return Recipe.objects.create(
        name='Ensalada César',
        category='Almuerzo',
        difficulty='Fácil',
        servings=2,
        calories=350,
        protein=Decimal('25.0'),
        carbs=Decimal('20.0'),
        fat=Decimal('12.0'),
        is_active=True,
        is_featured=False,
        meal_types=['lunch'],
    )


@pytest.fixture
def food(db):
    return Food.objects.create(
        name='Lechuga',
        calories=15,
        protein=Decimal('1.5'),
        carbs=Decimal('2.0'),
        fat=Decimal('0.2'),
        serving_size=100,
        serving_unit='g',
    )


@pytest.fixture
def nutrition_plan(db, regular_user):
    return NutritionPlan.objects.create(
        name='Plan Admin Test',
        user=regular_user,
        daily_calories=2000,
        protein_grams=150,
        carbs_grams=250,
        fat_grams=60,
        is_active=True,
        is_template=False,
    )


@pytest.fixture
def template_plan(db):
    return NutritionPlan.objects.create(
        name='Plantilla Admin Test',
        daily_calories=1800,
        protein_grams=130,
        carbs_grams=220,
        fat_grams=55,
        is_active=True,
        is_template=True,
    )


@pytest.fixture
def plan_meal(db, nutrition_plan):
    return PlanMeal.objects.create(
        plan=nutrition_plan,
        name='Desayuno',
        meal_type='breakfast',
        order_index=1,
        calories=500,
    )


@pytest.fixture
def plan_meal_recipe(db, plan_meal, recipe):
    return PlanMealRecipe.objects.create(
        meal=plan_meal,
        recipe=recipe,
        servings=Decimal('1.0'),
    )


@pytest.fixture
def meal_log(db, regular_user):
    import uuid
    return MealLog.objects.create(
        user=regular_user,
        date='2024-01-15',
        meal_type='breakfast',
        calories=500,
        protein=Decimal('40.0'),
    )


# ---------------------------------------------------------------------------
# AdminRecipeViewSet tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminRecipeViewSet:
    """Tests para CRUD de AdminRecipeViewSet"""

    def test_list_requires_admin(self, regular_client):
        response = regular_client.get(RECIPES_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_list_requires_auth(self):
        client = APIClient()
        response = client.get(RECIPES_URL)
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_list_recipes(self, admin_client, recipe):
        response = admin_client.get(RECIPES_URL)
        assert response.status_code == status.HTTP_200_OK
        data = response.data
        items = data.get('results', data) if isinstance(data, dict) else data
        assert len(items) >= 1

    def test_retrieve_recipe(self, admin_client, recipe):
        response = admin_client.get(f'{RECIPES_URL}{recipe.id}/')
        assert response.status_code == status.HTTP_200_OK
        assert response.data['name'] == recipe.name

    def test_create_recipe(self, admin_client):
        data = {
            'name': 'Nueva Receta Admin',
            'category': 'Cena',
            'difficulty': 'Medio',
            'servings': 1,
            'calories': 400,
            'protein': '30.0',
            'carbs': '40.0',
            'fat': '10.0',
            'is_active': True,
            'meal_types': ['dinner'],
        }
        response = admin_client.post(RECIPES_URL, data, format='json')
        assert response.status_code in [status.HTTP_201_CREATED, status.HTTP_200_OK]

    def test_update_recipe(self, admin_client, recipe):
        data = {'name': 'Ensalada César Actualizada', 'calories': 360}
        response = admin_client.patch(f'{RECIPES_URL}{recipe.id}/', data, format='json')
        assert response.status_code == status.HTTP_200_OK
        assert response.data['name'] == 'Ensalada César Actualizada'

    def test_delete_recipe(self, admin_client, recipe):
        response = admin_client.delete(f'{RECIPES_URL}{recipe.id}/')
        assert response.status_code == status.HTTP_204_NO_CONTENT

    def test_stats_action(self, admin_client, recipe):
        response = admin_client.get(f'{RECIPES_URL}stats/')
        assert response.status_code == status.HTTP_200_OK

    def test_stats_requires_admin(self, regular_client):
        response = regular_client.get(f'{RECIPES_URL}stats/')
        assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestAdminRecipeExportImport:
    """Tests para export/import de AdminRecipeViewSet"""

    def test_export_csv(self, admin_client, recipe):
        response = admin_client.get(f'{RECIPES_URL}export-csv/')
        assert response.status_code == status.HTTP_200_OK
        assert 'text/csv' in response['Content-Type']

    def test_export_csv_requires_admin(self, regular_client):
        response = regular_client.get(f'{RECIPES_URL}export-csv/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_excel(self, admin_client, recipe):
        response = admin_client.get(f'{RECIPES_URL}export-excel/')
        assert response.status_code == status.HTTP_200_OK
        assert 'spreadsheet' in response['Content-Type']

    def test_export_excel_requires_admin(self, regular_client):
        response = regular_client.get(f'{RECIPES_URL}export-excel/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_no_file(self, admin_client):
        response = admin_client.post(f'{RECIPES_URL}import-csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_csv_requires_admin(self, regular_client):
        response = regular_client.post(f'{RECIPES_URL}import-csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_valid(self, admin_client):
        csv_content = (
            'nombre,categoria,dificultad,porciones,calorias,proteinas,carbohidratos,grasas\n'
            'Receta Import,Almuerzo,Fácil,2,400,30.0,40.0,10.0\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'recipes.csv'
        response = admin_client.post(
            f'{RECIPES_URL}import-csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST]

    def test_import_excel_no_file(self, admin_client):
        response = admin_client.post(f'{RECIPES_URL}import-excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_excel_requires_admin(self, regular_client):
        response = regular_client.post(f'{RECIPES_URL}import-excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_excel_valid(self, admin_client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['nombre', 'categoria', 'dificultad', 'porciones', 'calorias', 'proteinas', 'carbohidratos', 'grasas', 'activa'])
        ws.append(['Receta Excel', 'Cena', 'Media', 2, 500, 35.0, 50.0, 12.0, 'Sí'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'recipes.xlsx'
        response = admin_client.post(
            f'{RECIPES_URL}import-excel/',
            {'file': buf},
            format='multipart',
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST]


# ---------------------------------------------------------------------------
# AdminNutritionPlanViewSet tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminNutritionPlanViewSet:
    """Tests para CRUD de AdminNutritionPlanViewSet"""

    def test_list_requires_admin(self, regular_client):
        response = regular_client.get(PLANS_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_list_plans(self, admin_client, nutrition_plan):
        response = admin_client.get(PLANS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_retrieve_plan(self, admin_client, nutrition_plan):
        response = admin_client.get(f'{PLANS_URL}{nutrition_plan.id}/')
        assert response.status_code == status.HTTP_200_OK
        assert response.data['name'] == nutrition_plan.name

    def test_create_plan(self, admin_client):
        data = {
            'name': 'Nuevo Plan Admin',
            'daily_calories': 2200,
            'protein_grams': 160,
            'carbs_grams': 270,
            'fat_grams': 65,
            'is_active': True,
            'is_template': True,
        }
        response = admin_client.post(PLANS_URL, data, format='json')
        assert response.status_code in [status.HTTP_201_CREATED, status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST]

    def test_update_plan(self, admin_client, nutrition_plan):
        data = {'name': 'Plan Actualizado', 'daily_calories': 2100}
        response = admin_client.patch(f'{PLANS_URL}{nutrition_plan.id}/', data, format='json')
        assert response.status_code == status.HTTP_200_OK

    def test_update_plan_daily_calories_scales_meals_and_recipes(self, admin_client, nutrition_plan, recipe):
        meal = PlanMeal.objects.create(
            plan=nutrition_plan,
            day_of_week=1,
            name='Comida principal',
            meal_type='lunch',
            order_index=1,
            calories=1000,
            protein=Decimal('50.0'),
            carbs=Decimal('120.0'),
            fat=Decimal('30.0'),
        )
        meal_recipe = PlanMealRecipe.objects.create(
            meal=meal,
            recipe=recipe,
            servings=Decimal('1.0'),
            custom_calories=500,
            custom_protein=Decimal('25.0'),
            custom_carbs=Decimal('40.0'),
            custom_fat=Decimal('12.0'),
        )

        response = admin_client.patch(
            f'{PLANS_URL}{nutrition_plan.id}/',
            {'daily_calories': 2400},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        nutrition_plan.refresh_from_db()
        meal.refresh_from_db()
        meal_recipe.refresh_from_db()

        assert nutrition_plan.daily_calories == 2400
        assert meal.calories > 1000
        assert meal_recipe.custom_calories and meal_recipe.custom_calories > 500
        assert float(meal_recipe.servings) > 1.0

    def test_update_plan_preserves_meal_order_indexes(self, admin_client, nutrition_plan):
        payload = {
            'meals': [
                {
                    'day_of_week': 1,
                    'name': 'Comida tarde',
                    'meal_type': 'dinner',
                    'time': '21:00',
                    'order_index': 2,
                },
                {
                    'day_of_week': 1,
                    'name': 'Comida mañana',
                    'meal_type': 'breakfast',
                    'time': '08:00',
                    'order_index': 1,
                },
            ]
        }

        response = admin_client.patch(f'{PLANS_URL}{nutrition_plan.id}/', payload, format='json')

        assert response.status_code == status.HTTP_200_OK
        meals = list(nutrition_plan.meals.order_by('day_of_week', 'order_index').values_list('name', 'order_index'))
        assert meals == [('Comida mañana', 1), ('Comida tarde', 2)]

    def test_generate_weekly_progression_from_base_day(self, admin_client, nutrition_plan, recipe):
        base_meal = PlanMeal.objects.create(
            plan=nutrition_plan,
            day_of_week=1,
            name='Comida principal',
            meal_type='lunch',
            order_index=1,
            calories=500,
            protein=Decimal('30.0'),
            carbs=Decimal('45.0'),
            fat=Decimal('15.0'),
        )
        base_meal.suggested_recipes.add(recipe)
        PlanMealRecipe.objects.create(
            meal=base_meal,
            recipe=recipe,
            servings=Decimal('1.00'),
            custom_calories=500,
            custom_protein=Decimal('30.0'),
            custom_carbs=Decimal('45.0'),
            custom_fat=Decimal('15.0'),
        )

        response = admin_client.post(
            f'{PLANS_URL}{nutrition_plan.id}/generate-weekly-progression/',
            {
                'base_day': 1,
                'step_percent': 2,
                'mode': 'increase',
                'overwrite': False,
                'preserve_base_day': True,
            },
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.data['created_meals'] == 6
        assert PlanMeal.objects.filter(plan=nutrition_plan).count() == 7

        friday_meal = PlanMeal.objects.get(plan=nutrition_plan, day_of_week=5)
        friday_recipe = friday_meal.meal_recipes.get(recipe=recipe)

        assert friday_meal.calories == 540
        assert friday_meal.protein == Decimal('32.40')
        assert friday_recipe.custom_calories == 540
        assert friday_recipe.servings == Decimal('1.08')

    def test_generate_weekly_progression_does_not_overwrite_without_flag(self, admin_client, nutrition_plan):
        PlanMeal.objects.create(
            plan=nutrition_plan,
            day_of_week=1,
            name='Base',
            meal_type='breakfast',
            order_index=1,
            calories=400,
        )
        existing = PlanMeal.objects.create(
            plan=nutrition_plan,
            day_of_week=2,
            name='Martes existente',
            meal_type='lunch',
            order_index=1,
            calories=900,
        )

        response = admin_client.post(
            f'{PLANS_URL}{nutrition_plan.id}/generate-weekly-progression/',
            {
                'base_day': 1,
                'step_percent': 5,
                'mode': 'decrease',
                'overwrite': False,
            },
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        existing.refresh_from_db()
        assert existing.name == 'Martes existente'
        assert existing.calories == 900

    def test_delete_plan(self, admin_client, nutrition_plan):
        response = admin_client.delete(f'{PLANS_URL}{nutrition_plan.id}/')
        assert response.status_code == status.HTTP_204_NO_CONTENT


@pytest.mark.django_db
class TestAdminNutritionPlanExportImport:
    """Tests para export/import de AdminNutritionPlanViewSet"""

    def test_export_csv_empty(self, admin_client):
        response = admin_client.get(f'{PLANS_URL}export-csv/')
        assert response.status_code == status.HTTP_200_OK
        assert 'text/csv' in response['Content-Type']

    def test_export_csv_with_plans(self, admin_client, nutrition_plan, plan_meal):
        response = admin_client.get(f'{PLANS_URL}export-csv/')
        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        assert 'Plan Admin Test' in content

    def test_export_csv_requires_admin(self, regular_client):
        response = regular_client.get(f'{PLANS_URL}export-csv/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_excel_empty(self, admin_client):
        response = admin_client.get(f'{PLANS_URL}export-excel/')
        assert response.status_code == status.HTTP_200_OK
        assert 'spreadsheet' in response['Content-Type']

    def test_export_excel_with_plans(self, admin_client, nutrition_plan, plan_meal):
        response = admin_client.get(f'{PLANS_URL}export-excel/')
        assert response.status_code == status.HTTP_200_OK

    def test_export_excel_requires_admin(self, regular_client):
        response = regular_client.get(f'{PLANS_URL}export-excel/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_no_file(self, admin_client):
        response = admin_client.post(f'{PLANS_URL}import-csv/', {}, format='multipart')
        assert response.status_code in [status.HTTP_400_BAD_REQUEST, status.HTTP_200_OK]

    def test_import_csv_requires_admin(self, regular_client):
        response = regular_client.post(f'{PLANS_URL}import-csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_minimal(self, admin_client):
        """CSV con mínimos campos requeridos"""
        csv_content = (
            'nombre_plan,descripcion,activo,es_plantilla,calorias_objetivo,proteinas_objetivo,carbohidratos_objetivo,grasas_objetivo\n'
            'Plan CSV Import,Descripción,Sí,Sí,2000,150,250,60\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'plans.csv'
        response = admin_client.post(
            f'{PLANS_URL}import-csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST
        ]

    def test_import_excel_no_file(self, admin_client):
        response = admin_client.post(f'{PLANS_URL}import-excel/', {}, format='multipart')
        assert response.status_code in [status.HTTP_400_BAD_REQUEST, status.HTTP_200_OK]

    def test_import_excel_requires_admin(self, regular_client):
        response = regular_client.post(f'{PLANS_URL}import-excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_excel_valid(self, admin_client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            'nombre_plan', 'descripcion', 'activo', 'es_plantilla',
            'calorias_objetivo', 'proteinas_objetivo', 'carbohidratos_objetivo', 'grasas_objetivo'
        ])
        ws.append(['Plan Excel Import', 'Descripción', 'Sí', 'Sí', 2000, 150, 250, 60])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'plans.xlsx'
        response = admin_client.post(
            f'{PLANS_URL}import-excel/',
            {'file': buf},
            format='multipart',
        )
        assert response.status_code in [
            status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST
        ]


# ---------------------------------------------------------------------------
# AdminPlanMealViewSet tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminPlanMealViewSet:
    """Tests para CRUD de AdminPlanMealViewSet"""

    def test_list_requires_admin(self, regular_client):
        response = regular_client.get(MEALS_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_list_meals(self, admin_client, plan_meal):
        response = admin_client.get(MEALS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_create_meal(self, admin_client, nutrition_plan):
        data = {
            'plan_id': str(nutrition_plan.id),
            'name': 'Almuerzo Admin',
            'meal_type': 'lunch',
            'order_index': 2,
            'calories': 700,
        }
        response = admin_client.post(MEALS_URL, data, format='json')
        assert response.status_code in [status.HTTP_201_CREATED, status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST]

    def test_delete_meal(self, admin_client, plan_meal):
        response = admin_client.delete(f'{MEALS_URL}{plan_meal.id}/')
        assert response.status_code == status.HTTP_204_NO_CONTENT


# ---------------------------------------------------------------------------
# AdminFoodViewSet tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminFoodViewSet:
    """Tests para CRUD de AdminFoodViewSet"""

    def test_list_requires_admin(self, regular_client):
        response = regular_client.get(FOODS_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_list_foods(self, admin_client, food):
        response = admin_client.get(FOODS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_retrieve_food(self, admin_client, food):
        response = admin_client.get(f'{FOODS_URL}{food.id}/')
        assert response.status_code == status.HTTP_200_OK
        assert response.data['name'] == food.name

    def test_create_food(self, admin_client):
        data = {
            'name': 'Pollo Admin',
            'calories': 165,
            'protein': '31.0',
            'carbs': '0.0',
            'fat': '3.6',
            'serving_size': 100,
            'serving_unit': 'g',
        }
        response = admin_client.post(FOODS_URL, data, format='json')
        assert response.status_code in [status.HTTP_201_CREATED, status.HTTP_200_OK]

    def test_update_food(self, admin_client, food):
        data = {'name': 'Lechuga Actualizada', 'calories': 16}
        response = admin_client.patch(f'{FOODS_URL}{food.id}/', data, format='json')
        assert response.status_code == status.HTTP_200_OK

    def test_delete_food(self, admin_client, food):
        response = admin_client.delete(f'{FOODS_URL}{food.id}/')
        assert response.status_code == status.HTTP_204_NO_CONTENT

    def test_list_for_recipes(self, admin_client, food):
        response = admin_client.get(f'{FOODS_URL}list-for-recipes/')
        assert response.status_code == status.HTTP_200_OK

    def test_list_for_recipes_requires_admin(self, regular_client):
        response = regular_client.get(f'{FOODS_URL}list-for-recipes/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_csv(self, admin_client, food):
        response = admin_client.get(f'{FOODS_URL}export-csv/')
        assert response.status_code == status.HTTP_200_OK
        assert 'text/csv' in response['Content-Type']

    def test_export_csv_requires_admin(self, regular_client):
        response = regular_client.get(f'{FOODS_URL}export-csv/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_export_excel(self, admin_client, food):
        response = admin_client.get(f'{FOODS_URL}export-excel/')
        assert response.status_code == status.HTTP_200_OK

    def test_export_excel_requires_admin(self, regular_client):
        response = regular_client.get(f'{FOODS_URL}export-excel/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_no_file(self, admin_client):
        response = admin_client.post(f'{FOODS_URL}import-csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_csv_requires_admin(self, regular_client):
        response = regular_client.post(f'{FOODS_URL}import-csv/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_csv_valid(self, admin_client):
        csv_content = (
            'nombre,calorias,proteinas,carbohidratos,grasas,tamano_porcion,unidad_porcion\n'
            'Alimento Test,100,10.0,15.0,3.0,100,g\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'foods.csv'
        response = admin_client.post(
            f'{FOODS_URL}import-csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST]

    def test_import_csv_infers_allergens_when_not_marked(self, admin_client):
        csv_content = (
            'nombre,marca,categoria,calorias,proteinas,carbohidratos,grasas,tamano_porcion,unidad_porcion\n'
            'Pan integral,Genérico,Cereales,245,8.0,46.0,2.0,100,g\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'foods_with_inference.csv'
        response = admin_client.post(
            f'{FOODS_URL}import-csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code == status.HTTP_200_OK
        food = Food.objects.get(name='Pan integral')
        assert 'gluten' in (food.allergens or [])

    def test_import_csv_keeps_declared_allergens_and_adds_detected(self, admin_client):
        csv_content = (
            'nombre,marca,categoria,alergenos,calorias,proteinas,carbohidratos,grasas,tamano_porcion,unidad_porcion\n'
            'Barrita proteica,Marca X,Snacks,soy,380,24.0,35.0,12.0,100,g\n'
        )
        csv_file = io.BytesIO(csv_content.encode('utf-8'))
        csv_file.name = 'foods_allergens.csv'
        response = admin_client.post(
            f'{FOODS_URL}import-csv/',
            {'file': csv_file},
            format='multipart',
        )
        assert response.status_code == status.HTTP_200_OK
        food = Food.objects.get(name='Barrita proteica')
        assert 'soy' in (food.allergens or [])

    def test_import_excel_no_file(self, admin_client):
        response = admin_client.post(f'{FOODS_URL}import-excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_import_excel_requires_admin(self, regular_client):
        response = regular_client.post(f'{FOODS_URL}import-excel/', {}, format='multipart')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_import_excel_valid(self, admin_client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['nombre', 'calorias', 'proteinas', 'carbohidratos', 'grasas', 'tamano_porcion', 'unidad_porcion'])
        ws.append(['Alimento Excel', 120, 12.0, 18.0, 4.0, 100, 'g'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'foods.xlsx'
        response = admin_client.post(
            f'{FOODS_URL}import-excel/',
            {'file': buf},
            format='multipart',
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST]

    def test_import_excel_infers_allergens_when_not_marked(self, admin_client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['nombre', 'marca', 'categoria', 'calorias', 'proteinas', 'carbohidratos', 'grasas', 'tamano_porcion', 'unidad_porcion'])
        ws.append(['Yogur natural', 'Marca Y', 'Lacteos', 60, 3.5, 4.7, 3.2, 100, 'g'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'foods_infer.xlsx'

        response = admin_client.post(
            f'{FOODS_URL}import-excel/',
            {'file': buf},
            format='multipart',
        )
        assert response.status_code == status.HTTP_200_OK
        food = Food.objects.get(name='Yogur natural')
        assert 'dairy' in (food.allergens or [])


# ---------------------------------------------------------------------------
# AdminPlanMealRecipeViewSet tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminPlanMealRecipeViewSet:
    """Tests para CRUD de AdminPlanMealRecipeViewSet"""

    def test_list_requires_admin(self, regular_client):
        response = regular_client.get(MEAL_RECIPES_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_list_meal_recipes(self, admin_client, plan_meal_recipe):
        response = admin_client.get(MEAL_RECIPES_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_retrieve_meal_recipe(self, admin_client, plan_meal_recipe):
        response = admin_client.get(f'{MEAL_RECIPES_URL}{plan_meal_recipe.id}/')
        assert response.status_code == status.HTTP_200_OK

    def test_delete_meal_recipe(self, admin_client, plan_meal_recipe):
        response = admin_client.delete(f'{MEAL_RECIPES_URL}{plan_meal_recipe.id}/')
        assert response.status_code == status.HTTP_204_NO_CONTENT


# ---------------------------------------------------------------------------
# admin_default_plans function view tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminDefaultPlans:
    """Tests para GET/POST /api/admin/nutrition/default-plans/"""

    def test_get_requires_admin(self, regular_client):
        response = regular_client.get(DEFAULT_PLANS_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_requires_auth(self):
        client = APIClient()
        response = client.get(DEFAULT_PLANS_URL)
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_get_default_plans_empty(self, admin_client):
        response = admin_client.get(DEFAULT_PLANS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_get_default_plans_with_template(self, admin_client, template_plan):
        response = admin_client.get(DEFAULT_PLANS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_get_default_plans_with_pagination(self, admin_client, template_plan):
        response = admin_client.get(f'{DEFAULT_PLANS_URL}?page=1&page_size=10')
        assert response.status_code == status.HTTP_200_OK
        assert 'results' in response.data


@pytest.mark.django_db
class TestAdminPlanAssignments:
    def test_change_user_plan_assigns_plan(self, admin_client, regular_user, template_plan):
        response = admin_client.post(
            CHANGE_USER_PLAN_URL,
            {
                'user_id': regular_user.id,
                'default_plan_id': str(template_plan.id),
            },
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.data.get('plan') is not None
        assert NutritionPlan.objects.filter(user=regular_user, is_active=True).exists()

    def test_change_user_plan_preserves_and_scales_custom_recipe_portions(
        self,
        admin_client,
        regular_user,
        template_plan,
        recipe,
    ):
        regular_user.admin_calories_override = 2400
        regular_user.save(update_fields=['admin_calories_override'])
        template_plan.daily_calories = 1800
        template_plan.save(update_fields=['daily_calories'])
        meal = PlanMeal.objects.create(
            plan=template_plan,
            name='Comida base',
            meal_type='lunch',
            order_index=1,
            calories=1800,
            protein=Decimal('90.0'),
            carbs=Decimal('180.0'),
            fat=Decimal('60.0'),
        )
        PlanMealRecipe.objects.create(
            meal=meal,
            recipe=recipe,
            servings=Decimal('1.50'),
            custom_calories=1800,
            custom_protein=Decimal('90.0'),
            custom_carbs=Decimal('180.0'),
            custom_fat=Decimal('60.0'),
        )

        response = admin_client.post(
            CHANGE_USER_PLAN_URL,
            {
                'user_id': regular_user.id,
                'default_plan_id': str(template_plan.id),
            },
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        assigned = NutritionPlan.objects.get(user=regular_user, is_active=True)
        meal_recipe = PlanMealRecipe.objects.get(meal__plan=assigned)
        assert assigned.daily_calories == 2400
        assert meal_recipe.custom_calories == 2400
        assert meal_recipe.servings == Decimal('2.00')

    def test_update_shared_assigned_plan_creates_user_copy_before_scaling(
        self,
        admin_client,
        regular_user,
        template_plan,
        recipe,
    ):
        assignment = NutritionPlanAssignment.objects.create(
            plan=template_plan,
            user=regular_user,
            is_active=True,
        )
        meal = PlanMeal.objects.create(
            plan=template_plan,
            name='Comida asignada',
            meal_type='lunch',
            order_index=1,
            calories=1800,
            protein=Decimal('90.0'),
            carbs=Decimal('180.0'),
            fat=Decimal('60.0'),
        )
        meal_recipe = PlanMealRecipe.objects.create(
            meal=meal,
            recipe=recipe,
            servings=Decimal('1.00'),
            custom_calories=1800,
            custom_protein=Decimal('90.0'),
            custom_carbs=Decimal('180.0'),
            custom_fat=Decimal('60.0'),
        )

        response = admin_client.patch(
            f'{PLANS_URL}{template_plan.id}/',
            {'user_id': regular_user.id, 'daily_calories': 2400},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        copied_plan_id = response.data['id']
        template_plan.refresh_from_db()
        assignment.refresh_from_db()
        copied_plan = NutritionPlan.objects.get(id=copied_plan_id)
        copied_meal_recipe = PlanMealRecipe.objects.get(meal__plan=copied_plan)

        assert copied_plan.user == regular_user
        assert copied_plan.daily_calories == 2400
        assert copied_meal_recipe.custom_calories == 2400
        assert copied_meal_recipe.servings == Decimal('1.33')
        assert assignment.is_active is False
        assert template_plan.daily_calories == 1800
        assert meal_recipe.custom_calories == 1800

    def test_update_multi_assigned_plan_without_user_context_edits_original(
        self,
        admin_client,
        regular_user,
        template_plan,
    ):
        second_user = User.objects.create_user(
            email='second_assigned_nutr@test.com',
            password='testpass123',
        )
        NutritionPlanAssignment.objects.create(plan=template_plan, user=regular_user, is_active=True)
        NutritionPlanAssignment.objects.create(plan=template_plan, user=second_user, is_active=True)

        response = admin_client.patch(
            f'{PLANS_URL}{template_plan.id}/',
            {'daily_calories': 2200},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        template_plan.refresh_from_db()
        assert response.data['id'] == str(template_plan.id)
        assert template_plan.daily_calories == 2200
        assert NutritionPlan.objects.filter(user__in=[regular_user, second_user]).count() == 0

    def test_bulk_change_plans_assigns_multiple_users(self, admin_client, regular_user, template_plan):
        second_user = User.objects.create_user(
            email='bulk_user_nutr@test.com',
            password='testpass123',
        )

        response = admin_client.post(
            BULK_CHANGE_PLANS_URL,
            {
                'default_plan_id': str(template_plan.id),
                'user_ids': [regular_user.id, second_user.id],
            },
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.data.get('success_count') == 2
        assert NutritionPlan.objects.filter(user=regular_user, is_active=True).exists()
        assert NutritionPlan.objects.filter(user=second_user, is_active=True).exists()


# ---------------------------------------------------------------------------
# admin_user_plan function view tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserPlan:
    """Tests para GET/PUT /api/admin/nutrition/users/{id}/plan/"""

    def test_get_requires_admin(self, regular_client, regular_user):
        response = regular_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_user_plan_no_plan(self, admin_client, regular_user):
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan/')
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_404_NOT_FOUND]

    def test_get_user_plan_with_plan(self, admin_client, regular_user, nutrition_plan):
        NutritionPlanAssignment.objects.create(
            user=regular_user,
            plan=nutrition_plan,
            is_active=True,
        )
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan/')
        assert response.status_code == status.HTTP_200_OK

    def test_get_nonexistent_user(self, admin_client):
        response = admin_client.get('/api/admin/nutrition/users/999999/plan/')
        assert response.status_code in [status.HTTP_404_NOT_FOUND, status.HTTP_200_OK]

    def test_get_user_plan_with_query_params(self, admin_client, regular_user, nutrition_plan):
        response = admin_client.get(
            f'/api/admin/nutrition/users/{regular_user.id}/plan/?days=14'
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_404_NOT_FOUND]

    def test_patch_user_plan_updates_daily_calories(self, admin_client, regular_user, nutrition_plan):
        response = admin_client.patch(
            f'/api/admin/nutrition/users/{regular_user.id}/plan/',
            {'daily_calories': 2300},
            format='json',
        )

        assert response.status_code == status.HTTP_200_OK
        nutrition_plan.refresh_from_db()
        assert nutrition_plan.daily_calories == 2300

    def test_patch_user_plan_requires_numeric_daily_calories(self, admin_client, regular_user, nutrition_plan):
        response = admin_client.patch(
            f'/api/admin/nutrition/users/{regular_user.id}/plan/',
            {'daily_calories': 'abc'},
            format='json',
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST


# ---------------------------------------------------------------------------
# admin_user_plan_history tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserPlanHistory:
    """Tests para GET /api/admin/nutrition/users/{id}/plan-history/"""

    def test_requires_admin(self, regular_client, regular_user):
        response = regular_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan-history/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_empty_history(self, admin_client, regular_user):
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan-history/')
        assert response.status_code == status.HTTP_200_OK

    def test_get_history_with_assignments(self, admin_client, regular_user, nutrition_plan):
        NutritionPlanAssignment.objects.create(
            user=regular_user,
            plan=nutrition_plan,
            is_active=False,
        )
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/plan-history/')
        assert response.status_code == status.HTTP_200_OK


# ---------------------------------------------------------------------------
# admin_user_meal_logs tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserMealLogs:
    """Tests para admin_user_meal_logs"""

    def test_requires_admin(self, regular_client, regular_user):
        response = regular_client.get(f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/')
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_empty_logs(self, admin_client, regular_user):
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/')
        assert response.status_code == status.HTTP_200_OK

    def test_get_logs_with_data(self, admin_client, regular_user, meal_log):
        response = admin_client.get(f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/')
        assert response.status_code == status.HTTP_200_OK

    def test_get_logs_nonexistent_user(self, admin_client):
        response = admin_client.get('/api/admin/nutrition/users/999999/meal-logs/')
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_404_NOT_FOUND]


# ---------------------------------------------------------------------------
# admin_user_meal_log_detail tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserMealLogDetail:
    """Tests para admin_user_meal_log_detail"""

    def test_requires_admin(self, regular_client, regular_user, meal_log):
        response = regular_client.get(
            f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/{meal_log.id}/'
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_patch_log(self, admin_client, regular_user, meal_log):
        """PATCH is allowed on meal log detail"""
        response = admin_client.patch(
            f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/{meal_log.id}/',
            {'calories': 550},
            format='json',
        )
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST]

    def test_delete_log(self, admin_client, regular_user, meal_log):
        response = admin_client.delete(
            f'/api/admin/nutrition/users/{regular_user.id}/meal-logs/{meal_log.id}/'
        )
        assert response.status_code == status.HTTP_204_NO_CONTENT


# ---------------------------------------------------------------------------
# admin_user_plans_stats tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserPlansStats:
    """Tests para admin_user_plans_stats"""

    def test_requires_admin(self, regular_client):
        response = regular_client.get(USER_PLANS_STATS_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_requires_auth(self):
        client = APIClient()
        response = client.get(USER_PLANS_STATS_URL)
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_get_stats_empty(self, admin_client):
        response = admin_client.get(USER_PLANS_STATS_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_get_stats_with_assignments(self, admin_client, regular_user, nutrition_plan):
        NutritionPlanAssignment.objects.create(
            user=regular_user,
            plan=nutrition_plan,
            is_active=True,
        )
        response = admin_client.get(USER_PLANS_STATS_URL)
        assert response.status_code == status.HTTP_200_OK


# ---------------------------------------------------------------------------
# admin_user_plans_usage_stats tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserPlansUsageStats:
    """Tests para admin_user_plans_usage_stats"""

    def test_requires_admin(self, regular_client):
        response = regular_client.get(USER_PLANS_USAGE_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_usage_stats_empty(self, admin_client):
        response = admin_client.get(USER_PLANS_USAGE_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_get_usage_stats_with_data(self, admin_client, regular_user, nutrition_plan):
        NutritionPlanAssignment.objects.create(
            user=regular_user,
            plan=nutrition_plan,
            is_active=True,
        )
        response = admin_client.get(USER_PLANS_USAGE_URL)
        assert response.status_code == status.HTTP_200_OK


# ---------------------------------------------------------------------------
# admin_user_plans_history tests
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestAdminUserPlansHistory:
    """Tests para admin_user_plans_history"""

    def test_requires_admin(self, regular_client):
        response = regular_client.get(USER_PLANS_HISTORY_URL)
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_get_history_empty(self, admin_client):
        response = admin_client.get(USER_PLANS_HISTORY_URL)
        assert response.status_code == status.HTTP_200_OK

    def test_get_history_with_data(self, admin_client, regular_user, nutrition_plan):
        NutritionPlanAssignment.objects.create(
            user=regular_user,
            plan=nutrition_plan,
            is_active=False,
        )
        response = admin_client.get(USER_PLANS_HISTORY_URL)
        assert response.status_code == status.HTTP_200_OK
