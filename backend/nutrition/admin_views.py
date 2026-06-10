# nutrition/admin_views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes as perm_classes, parser_classes as parser_decorator
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import transaction
from django.db import DatabaseError
from django.db.models import Count, Avg, Sum, Min, Max, Q, Exists, OuterRef
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
import requests
import re
import unicodedata
from urllib.parse import urlparse

from .models import (
    Recipe, NutritionPlan, PlanMeal, Food, MealLog, NutritionPlanHistory,
    PlanMealRecipe, NutritionPlanAssignment, CommunityRecipePost,
    CommunityRecipeLike, EquivalenceCategory
)
from .admin_serializers import (
    AdminRecipeSerializer, AdminNutritionPlanSerializer,
    AdminPlanMealSerializer, AdminFoodSerializer, PlanMealRecipeSerializer,
    AdminNutritionPlanMinimalSerializer, EquivalenceCategorySerializer
)
from .serializers import MealLogSerializer, NutritionPlanHistorySerializer
from .serializers import AdminCommunityRecipePostSerializer
from .services import PersonalizedNutritionService
from notifications.models import Notification

User = get_user_model()


class AdminCommunityRecipePostViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CommunityRecipePost.objects.all()
    serializer_class = AdminCommunityRecipePostSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'description', 'ingredients', 'instructions', 'author__email']
    filterset_fields = ['post_type']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_queryset(self):
        liked_subquery = CommunityRecipeLike.objects.filter(post=OuterRef('pk'), user=self.request.user)
        queryset = (
            CommunityRecipePost.objects.all()
            .select_related('author')
            .prefetch_related('comments__author')
            .annotate(
                likes_count=Count('likes', distinct=True),
                comments_count=Count('comments', distinct=True),
                liked_by_me=Exists(liked_subquery),
            )
            .order_by('-created_at')
        )
        return queryset

    @action(detail=True, methods=['post'], url_path='delete-with-reason')
    def delete_with_reason(self, request, pk=None):
        post = self.get_object()
        reason = str(request.data.get('reason') or '').strip()
        if not reason:
            return Response({'detail': 'La razón es obligatoria.'}, status=status.HTTP_400_BAD_REQUEST)

        author = post.author
        title = post.title
        Notification.objects.create(
            user=author,
            type='nutrition',
            title='Publicación de Team SK eliminada',
            message=f'Tu publicación "{title}" ha sido eliminada por el equipo. Motivo: {reason}',
            data={
                'priority': 'high',
                'source': 'community_recipe_moderation',
                'post_id': str(post.id),
                'post_title': title,
                'reason': reason,
            },
            action_url='/dashboard?section=recipe-community',
            expires_at=timezone.now() + timedelta(days=14),
        )
        post.delete()
        return Response({'message': 'Publicación eliminada y usuaria notificada.'})


class AdminRecipeViewSet(viewsets.ModelViewSet):
    queryset = Recipe.objects.all()
    serializer_class = AdminRecipeSerializer
    permission_classes = [IsAdminUser]
    parser_classes = [JSONParser, MultiPartParser, FormParser]  # Agregado JSONParser para crear recetas con JSON

    def list(self, request, *args, **kwargs):
        """Override list to add cache prevention headers"""
        try:
            response = super().list(request, *args, **kwargs)
        except DatabaseError:
            response = Response({
                'count': 0,
                'next': None,
                'previous': None,
                'results': [],
                'warning': 'Resultados parciales por incidencia temporal en base de datos'
            }, status=status.HTTP_200_OK)

        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    @action(detail=True, methods=['post'], url_path='upload-image')
    def upload_image(self, request, pk=None):
        recipe = self.get_object()
        image = request.FILES.get('image')

        if not image:
            return Response({'detail': 'image es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        allowed_types = {'image/jpeg', 'image/png', 'image/webp'}
        if image.content_type not in allowed_types:
            return Response({'detail': 'Formato no soportado'}, status=status.HTTP_400_BAD_REQUEST)

        max_size_bytes = 5 * 1024 * 1024
        if image.size and image.size > max_size_bytes:
            return Response({'detail': 'La imagen excede 5MB'}, status=status.HTTP_400_BAD_REQUEST)

        recipe.image = image
        recipe.save(update_fields=['image'])

        serializer = self.get_serializer(recipe)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='set-image-url')
    def set_image_url(self, request, pk=None):
        """Actualizar image_url (para URLs externas desde Google Drive, etc)"""
        recipe = self.get_object()
        image_url = request.data.get('image_url', '').strip()

        if not image_url:
            return Response({'detail': 'image_url es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar que sea una URL válida
        try:
            result = urlparse(image_url)
            if not all([result.scheme, result.netloc]):
                return Response({'detail': 'URL inválida'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'detail': f'Error validando URL: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar que sea accesible (HEAD request, timeout 5 segundos)
        try:
            headers = {'User-Agent': 'MyKaizenFit/1.0'}
            response = requests.head(image_url, timeout=5, allow_redirects=True, headers=headers)
            
            # Aceptar 200-299 como OK
            if response.status_code < 200 or response.status_code >= 300:
                return Response(
                    {'detail': f'URL no es accesible (status {response.status_code})'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que sea una imagen
            content_type = response.headers.get('content-type', '').lower()
            valid_types = ['image/jpeg', 'image/png', 'image/webp', 'image/gif']
            if not any(vtype in content_type for vtype in valid_types):
                return Response(
                    {'detail': f'URL no es una imagen válida (type: {content_type})'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except requests.exceptions.Timeout:
            return Response({'detail': 'URL tardó demasiado en responder'}, status=status.HTTP_400_BAD_REQUEST)
        except requests.exceptions.RequestException as e:
            return Response({'detail': f'Error accediendo a URL: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Actualizar URL
        recipe.image_url = image_url
        recipe.save(update_fields=['image_url'])

        serializer = self.get_serializer(recipe)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Estadísticas de recetas"""
        total_recipes = Recipe.objects.count()
        
        # Estadísticas por categoría
        by_category = Recipe.objects.values('category').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Promedios nutricionales
        avg_nutrition = Recipe.objects.aggregate(
            avg_calories=Avg('calories'),
            avg_protein=Avg('protein'),
            avg_carbs=Avg('carbs'),
            avg_fat=Avg('fat'),
        )
        
        # Recetas más usadas (en planes)
        popular_recipes = Recipe.objects.annotate(
            usage_count=Count('suggested_in_meals')
        ).order_by('-usage_count')[:10]
        
        # Conteo por dificultad
        by_difficulty = Recipe.objects.values('difficulty').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return Response({
            'total_recipes': total_recipes,
            'by_category': list(by_category),
            'by_difficulty': list(by_difficulty),
            'average_nutrition': {
                'calories': round(avg_nutrition['avg_calories'] or 0, 1),
                'protein': round(avg_nutrition['avg_protein'] or 0, 1),
                'carbs': round(avg_nutrition['avg_carbs'] or 0, 1),
                'fat': round(avg_nutrition['avg_fat'] or 0, 1),
            },
            'popular_recipes': [
                {
                    'id': str(r.id),
                    'name': r.name,
                    'usage_count': r.usage_count
                } for r in popular_recipes
            ]
        })

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """Exporta todas las recetas en formato CSV con ingredientes estructurados.
        Formato ingredientes: 'Nombre|cantidad|unidad|notas; Nombre2|cantidad|unidad|notas'
        Las macros se muestran como referencia pero se recalculan automáticamente al importar.
        """
        import csv
        from django.http import HttpResponse
        recipes = self.get_queryset().prefetch_related('recipe_ingredients__food')
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="recetas_exportacion.csv"'
        fieldnames = [
            'nombre', 'descripción', 'categoría', 'dificultad', 'porciones', 'tiempo_preparacion_minutos',
            'ingredientes', 'instrucciones', 'imagen_url',
            'calorias_ref', 'proteinas_ref', 'carbos_ref', 'grasas_ref'
        ]
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        
        for recipe in recipes:
            # Formato de ingredientes estructurado: Nombre|cantidad|unidad|notas; ....
            recipe_ingredients = recipe.recipe_ingredients.all()
            ingredients_str = ''
            if recipe_ingredients:
                ingredient_parts = []
                for ing in recipe_ingredients:
                    # Formato: NombreAlimento|cantidad|unidad|notas
                    part = f"{ing.food.name}|{ing.quantity}|{ing.unit}|{ing.notes or ''}"
                    ingredient_parts.append(part)
                ingredients_str = '; '.join(ingredient_parts)
            
            writer.writerow({
                'nombre': recipe.name,
                'descripción': recipe.description or '',
                'categoría': recipe.category or '',
                'dificultad': recipe.difficulty or '',
                'porciones': recipe.servings or 1,
                'tiempo_preparacion_minutos': recipe.prep_time_minutes or 0,
                'ingredientes': ingredients_str,
                'instrucciones': recipe.instructions or '',
                'imagen_url': recipe.image_url or '',
                'calorias_ref': recipe.calories or 0,
                'proteinas_ref': recipe.protein or 0,
                'carbos_ref': recipe.carbs or 0,
                'grasas_ref': recipe.fat or 0,
            })
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporta recetas en Excel con tres hojas:
        1. Recetas: Todas las recetas con ingredientes estructurados
        2. Ingredientes_Disponibles: Listado de todos los ingredientes con códigos para validación
        3. Referencias: Valores válidos para categoría y dificultad
        """
        import io
        import math
        import re
        import xlsxwriter
        from django.http import HttpResponse
        from collections import defaultdict

        illegal_xlsx_chars = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F\uFFFE\uFFFF]')

        def safe_text(value):
            if value is None:
                return ''
            text = str(value)
            text = ''.join(ch for ch in text if not (0xD800 <= ord(ch) <= 0xDFFF))
            return illegal_xlsx_chars.sub('', text)

        def safe_number(value, default=0):
            if value is None or value == '':
                return default
            try:
                num = float(value)
            except (TypeError, ValueError):
                return default
            if not math.isfinite(num):
                return default
            return num
        
        recipes = self.get_queryset().prefetch_related('recipe_ingredients__food')
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'strings_to_urls': False})
        
        # ========== HOJA 1: RECETAS ==========
        worksheet = workbook.add_worksheet('Recetas')
        
        headers = [
            'nombre', 'descripción', 'categoría', 'dificultad', 'porciones', 'tiempo_preparacion_minutos',
            'ingredientes', 'instrucciones', 'imagen_url',
            'calorias_ref', 'proteinas_ref', 'carbos_ref', 'grasas_ref'
        ]
        
        # Escribir headers con formato
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Recopilar ingredientes únicos mientras escribimos las recetas
        all_ingredients = defaultdict(set)
        
        for row_idx, recipe in enumerate(recipes, start=1):
            recipe_ingredients = recipe.recipe_ingredients.all()
            ingredients_str = ''
            if recipe_ingredients:
                ingredient_parts = []
                for ing in recipe_ingredients:
                    # Recopilar ingredientes
                    all_ingredients[ing.food.name].add((ing.unit, ing.food.id))
                    # Formato: NombreAlimento|cantidad|unidad|notas
                    part = f"{ing.food.name}|{ing.quantity}|{ing.unit}|{ing.notes or ''}"
                    ingredient_parts.append(part)
                ingredients_str = '; '.join(ingredient_parts)
            
            worksheet.write(row_idx, 0, safe_text(recipe.name))
            worksheet.write(row_idx, 1, safe_text(recipe.description or ''))
            worksheet.write(row_idx, 2, safe_text(recipe.category or ''))
            worksheet.write(row_idx, 3, safe_text(recipe.difficulty or ''))
            worksheet.write(row_idx, 4, recipe.servings or 1)
            worksheet.write(row_idx, 5, recipe.prep_time_minutes or 0)
            worksheet.write(row_idx, 6, safe_text(ingredients_str))
            worksheet.write(row_idx, 7, safe_text(recipe.instructions or ''))
            worksheet.write(row_idx, 8, safe_text(recipe.image_url or ''))
            worksheet.write(row_idx, 9, safe_number(recipe.calories, 0))
            worksheet.write(row_idx, 10, safe_number(recipe.protein, 0))
            worksheet.write(row_idx, 11, safe_number(recipe.carbs, 0))
            worksheet.write(row_idx, 12, safe_number(recipe.fat, 0))
        
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 35)
        worksheet.set_column('G:G', 60)
        worksheet.set_column('H:H', 50)

        # ========== HOJA 2: INGREDIENTES DISPONIBLES ==========
        ing_worksheet = workbook.add_worksheet('Ingredientes_Disponibles')
        
        # Headers para ingredientes
        ing_headers = ['Nombre Alimento', 'Unidad Recomendada', 'Calorías/100g', 'Proteína/100g', 'Carbos/100g', 'Grasa/100g', 'ID_BD']
        for col, header in enumerate(ing_headers):
            ing_worksheet.write(0, col, header, header_format)
        
        # Escribir todos los ingredientes disponibles ordenados alfabéticamente
        all_foods = Food.objects.all().order_by('name')
        for row_idx, food in enumerate(all_foods, start=1):
            ing_worksheet.write(row_idx, 0, safe_text(food.name))
            ing_worksheet.write(row_idx, 1, 'g')  # Unidad recomendada
            ing_worksheet.write(row_idx, 2, safe_number(food.calories, 0))
            ing_worksheet.write(row_idx, 3, safe_number(food.protein, 0))
            ing_worksheet.write(row_idx, 4, safe_number(food.carbs, 0))
            ing_worksheet.write(row_idx, 5, safe_number(food.fat, 0))
            ing_worksheet.write(row_idx, 6, safe_text(food.id))
        
        ing_worksheet.set_column('A:A', 30)
        ing_worksheet.set_column('B:B', 15)

        # ========== HOJA 3: REFERENCIAS ==========
        ref_worksheet = workbook.add_worksheet('Referencias')
        ref_worksheet.write(0, 0, 'Campo', header_format)
        ref_worksheet.write(0, 1, 'Valores válidos', header_format)
        ref_worksheet.write(0, 2, 'Notas', header_format)
        refs = [
            ('categoría', 'Desayuno, Almuerzo, Cena, Snack, Postre, Bebida', ''),
            ('dificultad', 'Fácil, Medio, Difícil', ''),
            ('porciones', '1, 2, 3, 4, ...', 'Número entero positivo'),
            ('tiempo_preparacion_minutos', '0, 5, 10, 15, 30, 60, ...', 'Número entero en minutos'),
            ('ingredientes', 'Nombre|cantidad|unidad|notas; Nombre2|cantidad2|unidad2|notas2', 'Separar ingredientes con " ; ", campos con "|"'),
            ('imagen_url', 'https://... (opcional)', 'URL de la imagen de la receta'),
            ('calorias_ref', 'Número', 'Solo referencia, se recalcula automáticamente'),
            ('proteinas_ref', 'Número decimal', 'Solo referencia, se recalcula automáticamente'),
            ('carbos_ref', 'Número decimal', 'Solo referencia, se recalcula automáticamente'),
            ('grasas_ref', 'Número decimal', 'Solo referencia, se recalcula automáticamente'),
        ]
        for row_idx, (campo, valores, notas) in enumerate(refs, start=1):
            ref_worksheet.write(row_idx, 0, safe_text(campo))
            ref_worksheet.write(row_idx, 1, safe_text(valores))
            ref_worksheet.write(row_idx, 2, safe_text(notas))
        ref_worksheet.set_column('A:A', 28)
        ref_worksheet.set_column('B:B', 55)
        ref_worksheet.set_column('C:C', 50)
        
        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="recetas_exportacion.xlsx"'
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    @action(detail=False, methods=['post'], url_path='import-csv')
    @parser_decorator([MultiPartParser, FormParser])
    def import_csv(self, request):
        """Importa recetas desde CSV con ingredientes estructurados.
        
        Acepta columnas en español o inglés.
        Formato ingredientes: 'NombreAlimento|cantidad|unidad|notas; NombreAlimento2|...'
        Las macros se calculan AUTOMÁTICAMENTE basándose en los ingredientes.
        Las columnas *_ref / *_ref se ignoran (son solo de referencia).
        
        - Busca recetas existentes por nombre
        - Actualiza si hay cambios
        - Crea nuevas recetas si no existen  
        - Recalcula macros automáticamente
        - Nunca borra
        """
        import csv
        import io
        import re
        import unicodedata
        from django.http import JsonResponse
        from .models import RecipeIngredient
        
        # Mapas de alias: columna exportada → campo interno
        field_aliases = {
            'nombre': 'name', 'name': 'name',
            'descripción': 'description', 'descripcion': 'description', 'description': 'description',
            'categoría': 'category', 'categoria': 'category', 'category': 'category',
            'dificultad': 'difficulty', 'difficulty': 'difficulty',
            'porciones': 'servings', 'servings': 'servings',
            'tiempo_preparacion_minutos': 'prep_time_minutes', 'prep_time_minutes': 'prep_time_minutes',
            'ingredientes': 'ingredients', 'ingredients': 'ingredients',
            'instrucciones': 'instructions', 'instructions': 'instructions',
            'imagen_url': 'image_url', 'image_url': 'image_url', 'url_imagen': 'image_url',
            'calorias_ref': 'calories_ref', 'calories_ref': 'calories_ref',
            'proteinas_ref': 'protein_ref', 'protein_ref': 'protein_ref',
            'carbos_ref': 'carbs_ref', 'carbs_ref': 'carbs_ref',
            'grasas_ref': 'fat_ref', 'fat_ref': 'fat_ref',
        }

        def get_value(row, canonical, default=''):
            """Busca un campo por su nombre canónico en una fila con aliases."""
            # Primero intentar directamente
            if canonical in row:
                return row[canonical]
            # Buscar en aliases
            for key, mapped in field_aliases.items():
                if mapped == canonical and key in row:
                    return row[key]
            return default

        def normalize_text(value):
            text = str(value or '').strip().lower()
            text = unicodedata.normalize('NFKD', text)
            text = ''.join(ch for ch in text if not unicodedata.combining(ch))
            text = re.sub(r'[^a-z0-9\s]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            return text

        ingredient_aliases = {
            'aove': 'aceite de oliva',
            'tomates cherry': 'tomate cherry',
            'huevo': 'huevos',
            'impulsor': 'levadura quimica',
            'arroz': 'arroz blanco',
        }

        def parse_quantity(raw_value):
            normalized = str(raw_value or '').strip().replace(',', '.')
            return float(normalized)

        def split_ingredient_parts(ingredients_raw):
            text = str(ingredients_raw or '')
            if not text.strip():
                return []
            return [p.strip() for p in re.split(r'[;\n\r]+', text) if p and p.strip()]

        def parse_ingredient_fields(part):
            # Formato principal: Nombre|cantidad|unidad|notas
            if '|' in part:
                fields = [f.strip() for f in part.split('|')]
                if len(fields) >= 3:
                    return fields[0], fields[1], fields[2], fields[3] if len(fields) > 3 else ''
                raise ValueError('se esperan al menos 3 campos separados por "|"')

            # Fallback: Nombre,cantidad,unidad,notas
            if ',' in part:
                fields = [f.strip() for f in part.split(',')]
                if len(fields) >= 3:
                    return fields[0], fields[1], fields[2], fields[3] if len(fields) > 3 else ''

            raise ValueError('formato invalido, usa Nombre|cantidad|unidad|notas')

        def resolve_food(food_name):
            original = str(food_name or '').strip()
            if not original:
                return None

            normalized = normalize_text(original)
            aliased = ingredient_aliases.get(normalized, original)

            # 1) Coincidencia exacta (original y alias)
            for candidate in [original, aliased]:
                food = Food.objects.filter(name__iexact=candidate).first()
                if food:
                    return food

            # 2) Singular/plural basico
            base = normalized[:-1] if normalized.endswith('s') else normalized
            for candidate in [base, f"{base}s"]:
                if candidate:
                    food = Food.objects.filter(name__icontains=candidate).first()
                    if food:
                        return food

            # 3) Coincidencia por tokens relevantes
            tokens = [t for t in normalize_text(aliased).split() if len(t) > 2]
            if tokens:
                qs = Food.objects.all()
                for token in tokens[:3]:
                    qs = qs.filter(name__icontains=token)
                food = qs.first()
                if food:
                    return food

            return None

        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'error': 'No file provided'}, status=400)
            
            text_file = io.TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.DictReader(text_file)
            
            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []
            failed_count = 0
            rejections = []
            
            for row_num, row in enumerate(reader, start=2):
                try:
                    name = get_value(row, 'name', '').strip()
                    if not name:
                        errors.append(f"Fila {row_num}: nombre es obligatorio")
                        continue
                    
                    # Buscar receta existente por nombre
                    existing_recipe = Recipe.objects.filter(name=name).first()
                    
                    # Preparar datos básicos (SIN MACROS - se calculan después)
                    recipe_data = {
                        'name': name,
                        'description': get_value(row, 'description', '') or '',
                        'category': get_value(row, 'category', '') or '',
                        'difficulty': get_value(row, 'difficulty', '') or '',
                        'servings': int(get_value(row, 'servings', 1) or 1),
                        'prep_time_minutes': int(get_value(row, 'prep_time_minutes', 0) or 0),
                        'instructions': get_value(row, 'instructions', '') or '',
                        'image_url': (get_value(row, 'image_url', '') or '').strip(),
                    }
                    
                    # Parsear ingredientes (formato: Nombre|cantidad|unidad|notas; ...)
                    ingredients_str = get_value(row, 'ingredients', '') or ''
                    parsed_ingredients = []
                    invalid_ingredients = []
                    
                    if ingredients_str:
                        ingredient_parts = split_ingredient_parts(ingredients_str)
                        for part in ingredient_parts:
                            try:
                                food_name, quantity_str, unit, notes = parse_ingredient_fields(part)
                                quantity = parse_quantity(quantity_str)

                                # Buscar alimento en la BD
                                food = resolve_food(food_name)
                                if not food:
                                    invalid_ingredients.append(f"'{food_name}': alimento no encontrado")
                                    continue

                                parsed_ingredients.append({
                                    'food': food,
                                    'quantity': quantity,
                                    'unit': unit,
                                    'notes': notes
                                })
                            except (ValueError, IndexError) as e:
                                invalid_ingredients.append(f"'{part}': {str(e)}")
                                continue

                    # Si hay ingredientes inválidos, rechazar receta completa para evitar cargas parciales
                    if invalid_ingredients:
                        failed_count += 1
                        rejections.append(
                            f"Fila {row_num} ('{name}'): Rechazada - Ingredientes inválidos: {', '.join(invalid_ingredients)}"
                        )
                        continue
                    
                    # Crear o actualizar receta
                    if existing_recipe:
                        # Actualizar campos básicos
                        has_changes = False
                        for field, value in recipe_data.items():
                            if getattr(existing_recipe, field) != value:
                                setattr(existing_recipe, field, value)
                                has_changes = True
                        
                        # Limpiar ingredientes anteriores y crear nuevos
                        existing_recipe.recipe_ingredients.all().delete()
                        for idx, ing_data in enumerate(parsed_ingredients):
                            RecipeIngredient.objects.create(
                                recipe=existing_recipe,
                                food=ing_data['food'],
                                quantity=ing_data['quantity'],
                                unit=ing_data['unit'],
                                notes=ing_data['notes'],
                                order=idx
                            )
                        
                        # Recalcular macros automáticamente
                        if parsed_ingredients:
                            existing_recipe.calculate_macros_from_ingredients()
                        
                        if has_changes or parsed_ingredients:
                            updated_count += 1
                        else:
                            skipped_count += 1
                    else:
                        # Crear nueva receta
                        new_recipe = Recipe.objects.create(**recipe_data)
                        
                        # Crear ingredientes
                        for idx, ing_data in enumerate(parsed_ingredients):
                            RecipeIngredient.objects.create(
                                recipe=new_recipe,
                                food=ing_data['food'],
                                quantity=ing_data['quantity'],
                                unit=ing_data['unit'],
                                notes=ing_data['notes'],
                                order=idx
                            )
                        
                        # Recalcular macros automáticamente
                        if parsed_ingredients:
                            new_recipe.calculate_macros_from_ingredients()
                        
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f"Fila {row_num}: {str(e)}")
                    continue
            
            message = (
                f"Importación completada: {created_count} creadas, {updated_count} actualizadas, "
                f"{skipped_count} omitidas, {failed_count} rechazadas"
            )
            if errors:
                message += f". {len(errors)} error(es) encontrados."

            result = {
                'success': True,
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'rejected': failed_count,
                'message': message,
            }

            if rejections:
                result['rejections'] = rejections[:20]
                result['rejection_count'] = len(rejections)

            if errors:
                result['errors'] = errors[:10]
                result['error_count'] = len(errors)

            return JsonResponse(result)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    @action(detail=False, methods=['post'], url_path='import-excel')
    @parser_decorator([MultiPartParser, FormParser])
    def import_excel(self, request):
        """Importa recetas desde Excel con validación inteligente de ingredientes.
        
        Acepta columnas en español o inglés.
        Formato ingredientes: 'NombreAlimento|cantidad|unidad|notas; NombreAlimento2|...'
        
        VALIDACIÓN: 
        - Valida TODOS los ingredientes ANTES de procesar
        - Si hay ingredientes inválidos/inexistentes, NO procesa esa receta
        - Continúa procesando las demás recetas válidas
        - Notifica cuáles recetas fueron rechazadas y por qué
        
        Las macros se calculan AUTOMÁTICAMENTE basándose en los ingredientes.
        Las columnas *_ref / *_ref se ignoran (son solo de referencia).
        """
        from openpyxl import load_workbook
        from django.http import JsonResponse
        from .models import RecipeIngredient
        import io
        import re
        import unicodedata
        
        # Mapas de alias: columna exportada → campo interno
        field_aliases = {
            'nombre': 'name', 'name': 'name',
            'descripción': 'description', 'descripcion': 'description', 'description': 'description',
            'categoría': 'category', 'categoria': 'category', 'category': 'category',
            'dificultad': 'difficulty', 'difficulty': 'difficulty',
            'porciones': 'servings', 'servings': 'servings',
            'tiempo_preparacion_minutos': 'prep_time_minutes', 'prep_time_minutes': 'prep_time_minutes',
            'ingredientes': 'ingredients', 'ingredients': 'ingredients',
            'instrucciones': 'instructions', 'instructions': 'instructions',
            'imagen_url': 'image_url', 'image_url': 'image_url', 'url_imagen': 'image_url',
            'calorias_ref': 'calories_ref', 'calories_ref': 'calories_ref',
            'proteinas_ref': 'protein_ref', 'protein_ref': 'protein_ref',
            'carbos_ref': 'carbs_ref', 'carbs_ref': 'carbs_ref',
            'grasas_ref': 'fat_ref', 'fat_ref': 'fat_ref',
        }

        def get_value(row_data, canonical, default=''):
            if canonical in row_data:
                return row_data[canonical]
            for key, mapped in field_aliases.items():
                if mapped == canonical and key in row_data:
                    return row_data[key]
            return default

        def normalize_text(value):
            text = str(value or '').strip().lower()
            text = unicodedata.normalize('NFKD', text)
            text = ''.join(ch for ch in text if not unicodedata.combining(ch))
            text = re.sub(r'[^a-z0-9\s]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            return text

        ingredient_aliases = {
            'aove': 'aceite de oliva',
            'tomates cherry': 'tomate cherry',
            'huevo': 'huevos',
            'impulsor': 'levadura quimica',
            'arroz': 'arroz blanco',
        }

        def parse_quantity(raw_value):
            normalized = str(raw_value or '').strip().replace(',', '.')
            return float(normalized)

        def split_ingredient_parts(ingredients_raw):
            text = str(ingredients_raw or '')
            if not text.strip():
                return []
            return [p.strip() for p in re.split(r'[;\n\r]+', text) if p and p.strip()]

        def parse_ingredient_fields(part):
            # Formato principal: Nombre|cantidad|unidad|notas
            if '|' in part:
                fields = [f.strip() for f in part.split('|')]
                if len(fields) >= 3:
                    return fields[0], fields[1], fields[2], fields[3] if len(fields) > 3 else ''
                raise ValueError('se esperan al menos 3 campos separados por "|"')

            # Fallback: Nombre,cantidad,unidad,notas
            if ',' in part:
                fields = [f.strip() for f in part.split(',')]
                if len(fields) >= 3:
                    return fields[0], fields[1], fields[2], fields[3] if len(fields) > 3 else ''

            raise ValueError('formato invalido, usa Nombre|cantidad|unidad|notas')

        def resolve_food(food_name):
            original = str(food_name or '').strip()
            if not original:
                return None

            normalized = normalize_text(original)
            aliased = ingredient_aliases.get(normalized, original)

            # 1) Coincidencia exacta (original y alias)
            for candidate in [original, aliased]:
                food = Food.objects.filter(name__iexact=candidate).first()
                if food:
                    return food

            # 2) Singular/plural basico
            base = normalized[:-1] if normalized.endswith('s') else normalized
            for candidate in [base, f"{base}s"]:
                if candidate:
                    food = Food.objects.filter(name__icontains=candidate).first()
                    if food:
                        return food

            # 3) Coincidencia por tokens relevantes
            tokens = [t for t in normalize_text(aliased).split() if len(t) > 2]
            if tokens:
                qs = Food.objects.all()
                for token in tokens[:3]:
                    qs = qs.filter(name__icontains=token)
                food = qs.first()
                if food:
                    return food

            return None

        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'error': 'No file provided'}, status=400)
            
            # Leer archivo Excel y seleccionar automaticamente la hoja correcta de recetas.
            workbook = load_workbook(io.BytesIO(file.read()))
            worksheet = None

            def normalize_header(value):
                return str(value or '').strip().lower()

            # 1) Intentar por nombre de hoja esperado
            for ws in workbook.worksheets:
                title = (ws.title or '').strip().lower()
                if title in ('recetas', 'recipes'):
                    worksheet = ws
                    break

            # 2) Si no existe, buscar una hoja con cabeceras de recetas (name/nombre + ingredients/ingredientes)
            if worksheet is None:
                for ws in workbook.worksheets:
                    title = (ws.title or '').strip().lower()
                    if title in ('referencias', 'ingredientes_disponibles', 'ingredients_available'):
                        continue

                    headers_probe = {
                        normalize_header(cell.value)
                        for cell in ws[1]
                        if cell.value is not None
                    }
                    if (
                        ('nombre' in headers_probe or 'name' in headers_probe)
                        and ('ingredientes' in headers_probe or 'ingredients' in headers_probe)
                    ):
                        worksheet = ws
                        break

            if worksheet is None:
                return JsonResponse(
                    {'error': 'No se encontro una hoja de recetas valida (Recetas/Recipes).'},
                    status=400
                )
            
            # Obtener headers desde la primera fila
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
            
            created_count = 0
            updated_count = 0
            skipped_count = 0
            failed_count = 0
            errors = []
            rejections = []  # Recetas rechazadas por ingredientes inválidos
            
            # Procesar filas de datos
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear diccionario con headers y valores
                    row_data = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_data[header] = row[idx]
                    
                    name = str(get_value(row_data, 'name') or '').strip()
                    if not name:
                        errors.append(f"Fila {row_num}: nombre es obligatorio")
                        continue
                    
                    # ========== VALIDACIÓN DE INGREDIENTES PRIMERO ==========
                    ingredients_str = get_value(row_data, 'ingredients', '') or ''
                    parsed_ingredients = []
                    invalid_ingredients = []
                    
                    if ingredients_str:
                        ingredient_parts = split_ingredient_parts(ingredients_str)
                        for part in ingredient_parts:
                            try:
                                food_name, quantity_str, unit, notes = parse_ingredient_fields(part)

                                # Validar cantidad
                                try:
                                    quantity = parse_quantity(quantity_str)
                                except ValueError:
                                    invalid_ingredients.append(f"'{food_name}': cantidad inválida '{quantity_str}'")
                                    continue

                                # Buscar alimento en la BD
                                food = resolve_food(food_name)
                                if not food:
                                    invalid_ingredients.append(f"'{food_name}': alimento no encontrado en la base de datos")
                                    continue

                                parsed_ingredients.append({
                                    'food': food,
                                    'quantity': quantity,
                                    'unit': unit,
                                    'notes': notes
                                })
                            except (ValueError, IndexError) as e:
                                invalid_ingredients.append(f"'{part}': {str(e)}")
                                continue
                    
                    # Si hay ingredientes INVÁLIDOS, RECHAZAR la receta completamente
                    if invalid_ingredients:
                        rejection_msg = f"Fila {row_num} ('{name}'): Rechazada - Ingredientes inválidos: {', '.join(invalid_ingredients)}"
                        rejections.append(rejection_msg)
                        failed_count += 1
                        continue
                    
                    # ========== PROCESAMIENTO SOLO SI TODO ES VÁLIDO ==========
                    recipe_data = {
                        'name': name,
                        'description': str(get_value(row_data, 'description') or '') or '',
                        'category': str(get_value(row_data, 'category') or '') or '',
                        'difficulty': str(get_value(row_data, 'difficulty') or '') or '',
                        'servings': int(get_value(row_data, 'servings') or 1),
                        'prep_time_minutes': int(get_value(row_data, 'prep_time_minutes') or 0),
                        'instructions': str(get_value(row_data, 'instructions') or '') or '',
                        'image_url': str(get_value(row_data, 'image_url') or '').strip(),
                    }
                    
                    # Buscar receta existente por nombre
                    existing_recipe = Recipe.objects.filter(name=name).first()
                    
                    if existing_recipe:
                        # Actualizar campos básicos
                        has_changes = False
                        for field, value in recipe_data.items():
                            if getattr(existing_recipe, field) != value:
                                setattr(existing_recipe, field, value)
                                has_changes = True
                        
                        # Limpiar ingredientes anteriores y crear nuevos
                        existing_recipe.recipe_ingredients.all().delete()
                        for idx, ing_data in enumerate(parsed_ingredients):
                            RecipeIngredient.objects.create(
                                recipe=existing_recipe,
                                food=ing_data['food'],
                                quantity=ing_data['quantity'],
                                unit=ing_data['unit'],
                                notes=ing_data['notes'],
                                order=idx
                            )
                        
                        # Recalcular macros automáticamente
                        if parsed_ingredients:
                            existing_recipe.calculate_macros_from_ingredients()
                        
                        if has_changes or parsed_ingredients:
                            updated_count += 1
                        else:
                            skipped_count += 1
                    else:
                        # Crear nueva receta
                        new_recipe = Recipe.objects.create(**recipe_data)
                        
                        # Crear ingredientes
                        for idx, ing_data in enumerate(parsed_ingredients):
                            RecipeIngredient.objects.create(
                                recipe=new_recipe,
                                food=ing_data['food'],
                                quantity=ing_data['quantity'],
                                unit=ing_data['unit'],
                                notes=ing_data['notes'],
                                order=idx
                            )
                        
                        # Recalcular macros automáticamente
                        if parsed_ingredients:
                            new_recipe.calculate_macros_from_ingredients()
                        
                        created_count += 1
                        
                except Exception as e:
                    errors.append(f"Fila {row_num}: Error inesperado - {str(e)}")
                    continue
            
            message = f"Importación completada: {created_count} creadas, {updated_count} actualizadas, {skipped_count} omitidas, {failed_count} RECHAZADAS (ingredientes inválidos)"
            
            result = {
                'success': True,
                'created': created_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'rejected': failed_count,
                'message': message,
            }
            
            if rejections:
                result['rejections'] = rejections[:20]
                result['rejection_count'] = len(rejections)
            
            if errors:
                result['errors'] = errors[:10]
                result['error_count'] = len(errors)
            
            return JsonResponse(result)
        except Exception as e:
            return JsonResponse({'error': f'Error procesando archivo: {str(e)}'}, status=400)


class AdminNutritionPlanViewSet(viewsets.ModelViewSet):
    queryset = NutritionPlan.objects.all().prefetch_related(
        'meals',
        'meals__suggested_recipes',
        'meals__meal_recipes',
        'meals__meal_recipes__recipe',
    )
    serializer_class = AdminNutritionPlanSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_template', 'is_system', 'is_active', 'user']
    search_fields = ['name', 'description', 'user__email', 'assignments__user__email']
    ordering_fields = ['created_at', 'updated_at', 'name', 'daily_calories', 'user__email', 'is_template', 'is_system', 'is_active']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if getattr(self, 'action', None) == 'list':
            return AdminNutritionPlanMinimalSerializer
        return AdminNutritionPlanSerializer

    def get_queryset(self):
        qs = NutritionPlan.objects.all()
        user_param = self.request.query_params.get('user')
        user_isnull = self.request.query_params.get('user__isnull')

        if user_param:
            qs = qs.filter(Q(user_id=user_param) | Q(assignments__user_id=user_param))

        if user_isnull == 'true':
            qs = qs.filter(Q(user__isnull=True) & Q(assignments__isnull=True))
        elif user_isnull == 'false':
            qs = qs.filter(Q(user__isnull=False) | Q(assignments__isnull=False))

        qs = qs.distinct()
        # Listado: evita cargar todo el árbol de comidas por defecto
        if getattr(self, 'action', None) == 'list':
            return qs.prefetch_related('meals', 'assignments__user').select_related('user')
        return qs.prefetch_related(
            'meals',
            'meals__suggested_recipes',
            'meals__meal_recipes',
            'meals__meal_recipes__recipe',
            'assignments__user',
        ).select_related('user')

    def _extract_assigned_user_ids(self, data):
        has_assigned = 'assigned_user_ids' in data or 'user_ids' in data or 'user_id' in data
        raw_ids = data.get('assigned_user_ids') or data.get('user_ids')
        if raw_ids is None and data.get('user_id') is not None:
            raw_ids = [data.get('user_id')]
        if raw_ids is None and not has_assigned:
            return None
        if raw_ids is None:
            return []
        if not isinstance(raw_ids, list):
            raw_ids = [raw_ids]
        normalized = []
        for uid in raw_ids:
            if uid is None or uid == 'none':
                continue
            try:
                normalized.append(int(uid))
            except (TypeError, ValueError):
                continue
        return list(dict.fromkeys(normalized))

    def _sync_assignments(self, plan: NutritionPlan, user_ids):
        user_ids = list(dict.fromkeys(user_ids))
        existing_ids = set(plan.assignments.values_list('user_id', flat=True))
        incoming_ids = set(user_ids)

        remove_ids = existing_ids - incoming_ids
        if remove_ids:
            plan.assignments.filter(user_id__in=remove_ids).delete()

        add_ids = incoming_ids - existing_ids
        for uid in add_ids:
            NutritionPlanAssignment.objects.create(
                plan=plan,
                user_id=uid,
                is_active=plan.is_active,
            )

        if incoming_ids:
            plan.assignments.filter(user_id__in=incoming_ids).update(is_active=plan.is_active)
            if plan.is_active:
                NutritionPlanAssignment.objects.filter(
                    user_id__in=incoming_ids,
                    is_active=True,
                ).exclude(plan=plan).update(is_active=False)
                NutritionPlan.objects.filter(
                    user_id__in=incoming_ids,
                    is_active=True,
                ).exclude(pk=plan.pk).update(is_active=False, end_date=timezone.now().date())

        if len(user_ids) == 1:
            if plan.user_id != user_ids[0]:
                plan.user_id = user_ids[0]
                plan.save(update_fields=['user'])
        else:
            if plan.user_id is not None:
                plan.user = None
                plan.save(update_fields=['user'])

    def _copy_plan_for_user(self, source_plan: NutritionPlan, user) -> NutritionPlan:
        """Create an individual editable copy so admin changes do not affect shared plans."""
        NutritionPlanAssignment.objects.filter(user=user, is_active=True).update(is_active=False)
        NutritionPlan.objects.filter(user=user, is_active=True).update(
            is_active=False,
            end_date=timezone.now().date(),
        )

        copied_plan = NutritionPlan.objects.create(
            user=user,
            name=source_plan.name,
            description=source_plan.description,
            daily_calories=source_plan.daily_calories,
            protein_grams=source_plan.protein_grams,
            carbs_grams=source_plan.carbs_grams,
            fat_grams=source_plan.fat_grams,
            fiber_grams=source_plan.fiber_grams,
            protein_percentage=source_plan.protein_percentage,
            carbs_percentage=source_plan.carbs_percentage,
            fat_percentage=source_plan.fat_percentage,
            goal=source_plan.goal,
            diet_type=source_plan.diet_type,
            meals_per_day=source_plan.meals_per_day,
            duration_weeks=source_plan.duration_weeks,
            portion_multiplier=source_plan.portion_multiplier,
            is_template=False,
            is_system=False,
            is_active=True,
            start_date=timezone.now().date(),
            end_date=source_plan.end_date,
            tags=source_plan.tags,
            image_url=source_plan.image_url,
            created_by=source_plan.created_by,
        )

        for source_meal in source_plan.meals.prefetch_related(
            'suggested_recipes',
            'meal_recipes__recipe',
        ).order_by('day_of_week', 'order_index', 'created_at'):
            copied_meal = PlanMeal.objects.create(
                plan=copied_plan,
                day_of_week=source_meal.day_of_week,
                name=source_meal.name,
                meal_type=source_meal.meal_type,
                time=source_meal.time,
                calories=source_meal.calories,
                protein=source_meal.protein,
                carbs=source_meal.carbs,
                fat=source_meal.fat,
                description=source_meal.description,
                order_index=source_meal.order_index,
            )
            copied_meal.suggested_recipes.set(source_meal.suggested_recipes.all())
            for meal_recipe in source_meal.meal_recipes.all():
                PlanMealRecipe.objects.create(
                    meal=copied_meal,
                    recipe=meal_recipe.recipe,
                    servings=meal_recipe.servings,
                    custom_calories=meal_recipe.custom_calories,
                    custom_protein=meal_recipe.custom_protein,
                    custom_carbs=meal_recipe.custom_carbs,
                    custom_fat=meal_recipe.custom_fat,
                    display_order=meal_recipe.display_order,
                )

        NutritionPlanAssignment.objects.create(
            plan=copied_plan,
            user=user,
            is_active=True,
        )
        return copied_plan

    def _owner_for_individual_edit(self, plan: NutritionPlan, assigned_user_ids):
        if assigned_user_ids and len(assigned_user_ids) == 1:
            try:
                return User.objects.get(pk=assigned_user_ids[0])
            except User.DoesNotExist:
                return None
        if plan.user_id:
            return plan.user
        active_assignments = list(plan.assignments.filter(is_active=True).select_related('user')[:2])
        if len(active_assignments) == 1:
            return active_assignments[0].user
        return None

    def _build_recipe_map(self, meals_payload):
        recipe_ids = set()
        for meal_data in meals_payload or []:
            if not isinstance(meal_data, dict):
                continue
            for mr in meal_data.get('meal_recipes') or []:
                if not isinstance(mr, dict):
                    continue
                recipe_id = mr.get('recipe_id') or (mr.get('recipe') or {}).get('id')
                if recipe_id:
                    recipe_ids.add(str(recipe_id))
        recipes = Recipe.objects.filter(id__in=list(recipe_ids))
        return {str(r.id): r for r in recipes}

    def _compute_meal_macros(self, meal_recipes, recipe_map):
        if not meal_recipes:
            return {'calories': 0, 'protein': 0, 'carbs': 0, 'fat': 0}

        options = []
        for mr in meal_recipes:
            if not isinstance(mr, dict):
                continue
            recipe_id = mr.get('recipe_id') or (mr.get('recipe') or {}).get('id')
            recipe = recipe_map.get(str(recipe_id)) if recipe_id else None
            if not recipe:
                continue

            servings = float(mr.get('servings') or 1)
            calories = mr.get('custom_calories')
            protein = mr.get('custom_protein')
            carbs = mr.get('custom_carbs')
            fat = mr.get('custom_fat')

            options.append({
                'calories': float(calories) if calories is not None else float(recipe.calories or 0) * servings,
                'protein': float(protein) if protein is not None else float(recipe.protein or 0) * servings,
                'carbs': float(carbs) if carbs is not None else float(recipe.carbs or 0) * servings,
                'fat': float(fat) if fat is not None else float(recipe.fat or 0) * servings,
            })

        if not options:
            return {'calories': 0, 'protein': 0, 'carbs': 0, 'fat': 0}

        total = {
            'calories': sum(o['calories'] for o in options),
            'protein': sum(o['protein'] for o in options),
            'carbs': sum(o['carbs'] for o in options),
            'fat': sum(o['fat'] for o in options),
        }
        count = len(options)
        return {
            'calories': total['calories'] / count,
            'protein': total['protein'] / count,
            'carbs': total['carbs'] / count,
            'fat': total['fat'] / count,
        }

    def _recompute_plan_macros(self, plan: NutritionPlan):
        day_totals = {}
        meals = list(plan.meals.all())

        for day in range(1, 8):
            totals = {'calories': 0, 'protein': 0, 'carbs': 0, 'fat': 0}
            for meal in meals:
                if meal.day_of_week and meal.day_of_week != day:
                    continue
                totals['calories'] += float(meal.calories or 0)
                totals['protein'] += float(meal.protein or 0)
                totals['carbs'] += float(meal.carbs or 0)
                totals['fat'] += float(meal.fat or 0)
            day_totals[day] = totals

        active_days = [t for t in day_totals.values() if sum(t.values()) > 0]
        if not active_days:
            return

        avg = {
            'calories': sum(t['calories'] for t in active_days) / len(active_days),
            'protein': sum(t['protein'] for t in active_days) / len(active_days),
            'carbs': sum(t['carbs'] for t in active_days) / len(active_days),
            'fat': sum(t['fat'] for t in active_days) / len(active_days),
        }

        plan.daily_calories = int(round(avg['calories']))
        plan.protein_grams = int(round(avg['protein']))
        plan.carbs_grams = int(round(avg['carbs']))
        plan.fat_grams = int(round(avg['fat']))
        plan.save(update_fields=['daily_calories', 'protein_grams', 'carbs_grams', 'fat_grams'])

    def _to_int(self, value):
        try:
            if value is None:
                return None
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def _to_decimal(self, value, default='0'):
        try:
            if value is None or value == '':
                value = default
            return Decimal(str(value))
        except Exception:
            return Decimal(str(default))

    def _to_bool(self, value, default=False):
        if value is None or value == '':
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {'1', 'true', 'yes', 'y', 'si', 'sí'}

    def _scale_decimal(self, value, multiplier):
        return (Decimal(str(value or 0)) * multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def _scale_int(self, value, multiplier):
        return int((Decimal(str(value or 0)) * multiplier).quantize(Decimal('1'), rounding=ROUND_HALF_UP))

    def _replace_plan_meals(self, plan: NutritionPlan, meals_payload):
        """
        Reemplaza TODAS las comidas del plan por las proporcionadas (enfoque robusto).
        Espera una lista de objetos con campos de PlanMeal y opcionalmente:
          - suggested_recipes_ids: [recipe_id,...]
          - meal_recipes: [{recipe_id, servings, custom_*, display_order}, ...]
        """
        if meals_payload is None:
            return
        if not isinstance(meals_payload, list):
            return

        recipe_map = self._build_recipe_map(meals_payload)

        # Eliminar comidas anteriores (cascade elimina PlanMealRecipe)
        plan.meals.all().delete()

        for idx, meal_data in enumerate(meals_payload):
            if not isinstance(meal_data, dict):
                continue

            suggested_ids = meal_data.get('suggested_recipes_ids')
            meal_recipes = meal_data.get('meal_recipes')

            meal = PlanMeal.objects.create(
                plan=plan,
                day_of_week=meal_data.get('day_of_week') or None,
                name=meal_data.get('name') or f'Comida {idx + 1}',
                meal_type=meal_data.get('meal_type') or 'lunch',
                time=meal_data.get('time') or None,
                calories=0,
                protein=0,
                carbs=0,
                fat=0,
                description=meal_data.get('description') or '',
                order_index=meal_data.get('order_index') or (idx + 1),
            )

            # ManyToMany simple
            if isinstance(suggested_ids, list) and suggested_ids:
                meal.suggested_recipes.set(Recipe.objects.filter(id__in=suggested_ids))

            effective_meal_recipes = []

            # Cantidades personalizadas por receta
            if isinstance(meal_recipes, list) and meal_recipes:
                for mr in meal_recipes:
                    if not isinstance(mr, dict):
                        continue
                    recipe_id = mr.get('recipe_id') or (mr.get('recipe') or {}).get('id')
                    if not recipe_id:
                        continue
                    try:
                        recipe = Recipe.objects.get(id=recipe_id)
                    except Recipe.DoesNotExist:
                        continue

                    payload = {
                        'meal': meal,
                        'recipe': recipe,
                        'servings': mr.get('servings') or 1.0,
                        'display_order': mr.get('display_order') or 0,
                        'custom_calories': mr.get('custom_calories', None),
                        'custom_protein': mr.get('custom_protein', None),
                        'custom_carbs': mr.get('custom_carbs', None),
                        'custom_fat': mr.get('custom_fat', None),
                    }
                    PlanMealRecipe.objects.create(**payload)
                    effective_meal_recipes.append(mr)

            # Compatibilidad: si no hay meal_recipes, crear desde suggested_ids
            if not effective_meal_recipes and isinstance(suggested_ids, list) and suggested_ids:
                for idx_s, recipe_id in enumerate(suggested_ids):
                    try:
                        recipe = Recipe.objects.get(id=recipe_id)
                    except Recipe.DoesNotExist:
                        continue
                    PlanMealRecipe.objects.create(
                        meal=meal,
                        recipe=recipe,
                        servings=1.0,
                        display_order=idx_s,
                    )
                    effective_meal_recipes.append({'recipe_id': recipe_id, 'servings': 1.0, 'display_order': idx_s})

            computed = self._compute_meal_macros(effective_meal_recipes, recipe_map)
            PlanMeal.objects.filter(pk=meal.pk).update(
                calories=int(round(computed['calories'])),
                protein=round(computed['protein'], 2),
                carbs=round(computed['carbs'], 2),
                fat=round(computed['fat'], 2),
            )

    @transaction.atomic
    @action(detail=True, methods=['post'], url_path='generate-weekly-progression')
    def generate_weekly_progression(self, request, pk=None):
        """
        Genera una semana completa a partir de un día base del plan.
        Escala comidas y opciones de receta por día para crear progresión semanal.
        """
        plan = self.get_object()
        base_day = self._to_int(request.data.get('base_day')) or 1
        if base_day < 1 or base_day > 7:
            return Response({'detail': 'base_day debe estar entre 1 y 7.'}, status=status.HTTP_400_BAD_REQUEST)

        target_days = request.data.get('target_days') or [1, 2, 3, 4, 5, 6, 7]
        if not isinstance(target_days, list):
            target_days = [target_days]
        normalized_target_days = []
        for day in target_days:
            day_int = self._to_int(day)
            if day_int and 1 <= day_int <= 7:
                normalized_target_days.append(day_int)
        target_days = sorted(set(normalized_target_days))
        if not target_days:
            return Response({'detail': 'target_days debe incluir al menos un día entre 1 y 7.'}, status=status.HTTP_400_BAD_REQUEST)

        step_percent = self._to_decimal(request.data.get('step_percent'), default='0')
        mode = str(request.data.get('mode') or 'increase').strip().lower()
        overwrite = self._to_bool(request.data.get('overwrite'), default=False)
        preserve_base_day = self._to_bool(request.data.get('preserve_base_day'), default=True)

        if mode not in {'increase', 'decrease'}:
            return Response({'detail': 'mode debe ser increase o decrease.'}, status=status.HTTP_400_BAD_REQUEST)
        if step_percent < Decimal('0') or step_percent > Decimal('25'):
            return Response({'detail': 'step_percent debe estar entre 0 y 25.'}, status=status.HTTP_400_BAD_REQUEST)

        base_meals = list(
            plan.meals.filter(day_of_week=base_day)
            .prefetch_related('suggested_recipes', 'meal_recipes__recipe')
            .order_by('order_index', 'created_at')
        )
        if not base_meals:
            base_meals = list(
                plan.meals.filter(day_of_week__isnull=True)
                .prefetch_related('suggested_recipes', 'meal_recipes__recipe')
                .order_by('order_index', 'created_at')
            )

        if not base_meals:
            return Response(
                {'detail': 'El plan no tiene comidas en el día base ni comidas generales para copiar.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_meals = 0
        updated_days = []

        for day in target_days:
            if day == base_day and preserve_base_day:
                continue
            if plan.meals.filter(day_of_week=day).exists():
                if not overwrite:
                    continue
                plan.meals.filter(day_of_week=day).delete()

            distance = day - base_day
            direction = Decimal('-1') if mode == 'decrease' else Decimal('1')
            multiplier = Decimal('1') + (direction * step_percent * Decimal(distance) / Decimal('100'))
            if multiplier < Decimal('0.10'):
                multiplier = Decimal('0.10')

            for base_meal in base_meals:
                meal = PlanMeal.objects.create(
                    plan=plan,
                    day_of_week=day,
                    name=base_meal.name,
                    meal_type=base_meal.meal_type,
                    time=base_meal.time,
                    calories=self._scale_int(base_meal.calories, multiplier),
                    protein=self._scale_decimal(base_meal.protein, multiplier),
                    carbs=self._scale_decimal(base_meal.carbs, multiplier),
                    fat=self._scale_decimal(base_meal.fat, multiplier),
                    description=base_meal.description,
                    order_index=base_meal.order_index,
                )
                meal.suggested_recipes.set(base_meal.suggested_recipes.all())

                for meal_recipe in base_meal.meal_recipes.all():
                    custom_calories = (
                        self._scale_int(meal_recipe.custom_calories, multiplier)
                        if meal_recipe.custom_calories is not None
                        else None
                    )
                    PlanMealRecipe.objects.create(
                        meal=meal,
                        recipe=meal_recipe.recipe,
                        servings=self._scale_decimal(meal_recipe.servings, multiplier),
                        custom_calories=custom_calories,
                        custom_protein=(
                            self._scale_decimal(meal_recipe.custom_protein, multiplier)
                            if meal_recipe.custom_protein is not None
                            else None
                        ),
                        custom_carbs=(
                            self._scale_decimal(meal_recipe.custom_carbs, multiplier)
                            if meal_recipe.custom_carbs is not None
                            else None
                        ),
                        custom_fat=(
                            self._scale_decimal(meal_recipe.custom_fat, multiplier)
                            if meal_recipe.custom_fat is not None
                            else None
                        ),
                        display_order=meal_recipe.display_order,
                    )
                created_meals += 1
            updated_days.append(day)

        self._recompute_plan_macros(plan)
        plan.refresh_from_db()

        return Response({
            'message': 'Progresión semanal generada',
            'created_meals': created_meals,
            'updated_days': updated_days,
            'daily_calories': plan.daily_calories,
        }, status=status.HTTP_200_OK)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        meals_payload = request.data.get('meals')
        assigned_user_ids = self._extract_assigned_user_ids(request.data)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        plan = serializer.save()

        if assigned_user_ids is not None:
            self._sync_assignments(plan, assigned_user_ids)

        # Normalizar flags según si es plan de usuario o plantilla
        has_assignees = plan.assignments.exists()
        if (plan.user_id or has_assignees) and not plan.is_system:
            if plan.is_template:
                plan.is_template = False
                plan.save(update_fields=['is_template'])
        if not plan.user_id and not has_assignees and not plan.is_system and not plan.is_template:
            # Por defecto, un admin creando sin usuario asignado es plantilla
            plan.is_template = True
            plan.save(update_fields=['is_template'])

        self._replace_plan_meals(plan, meals_payload)
        self._recompute_plan_macros(plan)

        plan.refresh_from_db()
        plan = NutritionPlan.objects.prefetch_related(
            'meals',
            'meals__suggested_recipes',
            'meals__meal_recipes',
            'meals__meal_recipes__recipe',
            'assignments__user',
        ).get(pk=plan.pk)
        return Response(self.get_serializer(plan).data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        meals_payload = request.data.get('meals')
        assigned_user_ids = self._extract_assigned_user_ids(request.data)
        prev_is_active = instance.is_active

        requested_daily_calories = self._to_int(request.data.get('daily_calories'))
        requested_protein = self._to_int(request.data.get('protein_grams'))
        requested_carbs = self._to_int(request.data.get('carbs_grams'))
        requested_fat = self._to_int(request.data.get('fat_grams'))

        edit_owner = self._owner_for_individual_edit(instance, assigned_user_ids)
        if edit_owner and instance.user_id != edit_owner.id:
            instance = self._copy_plan_for_user(instance, edit_owner)
            prev_is_active = instance.is_active

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        plan = serializer.save()

        if plan.is_active != prev_is_active:
            plan.assignments.all().update(is_active=plan.is_active)

        if assigned_user_ids is not None:
            self._sync_assignments(plan, assigned_user_ids)

        # Normalizar flags si cambió asignación de usuario
        has_assignees = plan.assignments.exists()
        if (plan.user_id or has_assignees) and not plan.is_system:
            if plan.is_template:
                plan.is_template = False
                plan.save(update_fields=['is_template'])
        if not plan.user_id and not has_assignees and not plan.is_system and not plan.is_template:
            plan.is_template = True
            plan.save(update_fields=['is_template'])

        self._replace_plan_meals(plan, meals_payload)
        self._recompute_plan_macros(plan)

        # Si un admin fija calorías manualmente en un plan de usuario,
        # reescalamos comidas/recetas para no perder ese valor al recalcular.
        calorie_owner = plan.user
        if calorie_owner is None:
            active_assignment = plan.assignments.filter(is_active=True).select_related('user').order_by('-assigned_at').first()
            calorie_owner = active_assignment.user if active_assignment else None

        if requested_daily_calories and calorie_owner and requested_daily_calories > 0:
            current_calories = int(plan.daily_calories or 0)
            if current_calories != requested_daily_calories:
                service = PersonalizedNutritionService(calorie_owner)
                plan = service.adjust_plan_calories(
                    plan,
                    requested_daily_calories - current_calories,
                    reason='admin_manual_update',
                    notes=f'Ajuste manual desde admin a {requested_daily_calories} kcal',
                )

        update_fields = []
        if requested_protein is not None:
            plan.protein_grams = requested_protein
            update_fields.append('protein_grams')
        if requested_carbs is not None:
            plan.carbs_grams = requested_carbs
            update_fields.append('carbs_grams')
        if requested_fat is not None:
            plan.fat_grams = requested_fat
            update_fields.append('fat_grams')
        if update_fields:
            plan.save(update_fields=update_fields)

        plan.refresh_from_db()
        plan = NutritionPlan.objects.prefetch_related(
            'meals',
            'meals__suggested_recipes',
            'meals__meal_recipes',
            'meals__meal_recipes__recipe',
            'assignments__user',
        ).get(pk=plan.pk)
        return Response(self.get_serializer(plan).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='allergen-check')
    def allergen_check(self, request, pk=None):
        """
        Dado un plan y una lista de user_ids, devuelve conflictos de alérgenos
        entre las recetas del plan y los alérgenos declarados de cada usuario.
        Respuesta: {warnings: [{user_id, user_email, user_name, conflicts: [{recipe_name, allergens}]}]}
        """
        from django.contrib.auth import get_user_model
        User = get_user_model()

        user_ids = request.data.get('user_ids', [])
        if not isinstance(user_ids, list):
            user_ids = [user_ids]
        user_ids = [int(uid) for uid in user_ids if uid]

        plan = self.get_object()
        recipe_allergens = []
        for meal in plan.meals.prefetch_related('meal_recipes__recipe').all():
            for pmr in meal.meal_recipes.all():
                r = pmr.recipe
                if r and r.allergens:
                    recipe_allergens.append({'recipe_name': r.name, 'allergens': list(r.allergens)})

        ALLERGEN_LABELS = {
            'gluten': 'Gluten', 'dairy': 'Lácteos', 'eggs': 'Huevo',
            'nuts': 'Frutos secos', 'soy': 'Soja', 'fish': 'Pescado',
            'shellfish': 'Marisco', 'sesame': 'Sésamo',
        }

        warnings = []
        for user in User.objects.filter(id__in=user_ids):
            user_allergens = set(getattr(user, 'allergies', None) or [])
            if not user_allergens:
                continue
            user_conflicts = []
            for ra in recipe_allergens:
                recipe_set = set(ra['allergens'])
                conflict = user_allergens & recipe_set
                if conflict:
                    user_conflicts.append({
                        'recipe_name': ra['recipe_name'],
                        'allergens': [ALLERGEN_LABELS.get(a, a) for a in conflict],
                    })
            if user_conflicts:
                warnings.append({
                    'user_id': user.id,
                    'user_email': user.email,
                    'user_name': f'{user.first_name} {user.last_name}'.strip() or user.email,
                    'conflicts': user_conflicts,
                })

        return Response({'warnings': warnings})

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """Exporta planes en CSV con paridad de edición del panel admin.
        Tipos de fila soportados:
        - plan: datos del plan
        - comida: datos de comidas por plan/día/orden
        - opcion_receta: opciones de receta por comida (servings/custom macros/display_order)
        - asignacion_usuario: asignaciones por email
        """
        import csv
        from django.http import HttpResponse
        plans = NutritionPlan.objects.prefetch_related(
            'meals__meal_recipes__recipe',
            'meals__suggested_recipes',
            'assignments__user',
        ).order_by('name')
        goal_labels = {value: label for value, label in NutritionPlan.GOAL_CHOICES}
        diet_labels = {value: label for value, label in NutritionPlan.DIET_TYPE_CHOICES}
        meal_type_labels = {value: label for value, label in PlanMeal.MEAL_TYPE_CHOICES}
        meal_type_labels.update({
            'morning_snack': 'Merienda',
            'afternoon_snack': 'Merienda',
            'pre_workout': 'Merienda',
            'post_workout': 'Merienda',
        })
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="planes_menu_exportacion.csv"'
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        fieldnames = [
            'tipo_fila',
            'plan_nombre', 'descripcion', 'objetivo', 'tipo_dieta',
            'calorias_diarias', 'proteinas_g', 'carbohidratos_g', 'grasas_g', 'fibra_g',
            'porcentaje_proteinas', 'porcentaje_carbohidratos', 'porcentaje_grasas',
            'comidas_por_dia', 'duracion_semanas', 'multiplicador_porcion',
            'es_plantilla', 'es_sistema', 'activo',
            # Campos exclusivos de comidas (vacíos en filas de plan)
            'comida_nombre', 'tipo_comida', 'dia_semana', 'orden_comida', 'hora',
            'calorias_comida', 'proteinas_comida', 'carbohidratos_comida', 'grasas_comida',
            'descripcion_comida', 'recetas_sugeridas',
            # Campos exclusivos de opcion_receta
            'receta_nombre', 'imagen_url', 'orden_visualizacion', 'porciones',
            'calorias_personalizadas', 'proteinas_personalizadas', 'carbohidratos_personalizados', 'grasas_personalizadas',
            # Campos exclusivos de asignacion_usuario
            'usuario_email', 'asignacion_activa',
        ]
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        empty_meal = {k: '' for k in ['comida_nombre', 'tipo_comida', 'dia_semana', 'orden_comida', 'hora',
                                       'calorias_comida', 'proteinas_comida', 'carbohidratos_comida',
                                       'grasas_comida', 'descripcion_comida', 'recetas_sugeridas']}
        empty_option = {k: '' for k in ['receta_nombre', 'imagen_url', 'orden_visualizacion', 'porciones',
                        'calorias_personalizadas', 'proteinas_personalizadas', 'carbohidratos_personalizados', 'grasas_personalizadas']}
        empty_assignment = {k: '' for k in ['usuario_email', 'asignacion_activa']}
        empty_plan = {k: '' for k in ['descripcion', 'objetivo', 'tipo_dieta',
                                       'calorias_diarias', 'proteinas_g', 'carbohidratos_g', 'grasas_g', 'fibra_g',
                                       'porcentaje_proteinas', 'porcentaje_carbohidratos', 'porcentaje_grasas',
                                       'comidas_por_dia', 'duracion_semanas', 'multiplicador_porcion',
                                       'es_plantilla', 'es_sistema', 'activo']}
        for plan in plans:
            writer.writerow({
                'tipo_fila': 'plan',
                'plan_nombre': plan.name,
                'descripcion': plan.description or '',
                'objetivo': goal_labels.get(plan.goal, plan.goal or 'Mantener peso'),
                'tipo_dieta': diet_labels.get(plan.diet_type, plan.diet_type or 'Normal'),
                'calorias_diarias': plan.daily_calories or 0,
                'proteinas_g': plan.protein_grams or 0,
                'carbohidratos_g': plan.carbs_grams or 0,
                'grasas_g': plan.fat_grams or 0,
                'fibra_g': plan.fiber_grams or 0,
                'porcentaje_proteinas': plan.protein_percentage if plan.protein_percentage is not None else '',
                'porcentaje_carbohidratos': plan.carbs_percentage if plan.carbs_percentage is not None else '',
                'porcentaje_grasas': plan.fat_percentage if plan.fat_percentage is not None else '',
                'comidas_por_dia': plan.meals_per_day or 5,
                'duracion_semanas': plan.duration_weeks or 4,
                'multiplicador_porcion': float(plan.portion_multiplier or 1.0),
                'es_plantilla': 'Si' if plan.is_template else 'No',
                'es_sistema': 'Si' if plan.is_system else 'No',
                'activo': 'Si' if plan.is_active else 'No',
                **empty_meal,
                **empty_option,
                **empty_assignment,
            })
            for meal in plan.meals.order_by('day_of_week', 'order_index'):
                recipe_names = set()
                for mr in meal.meal_recipes.all():
                    recipe_names.add(mr.recipe.name)
                for sr in meal.suggested_recipes.all():
                    recipe_names.add(sr.name)
                recetas_str = '; '.join(sorted(recipe_names))
                writer.writerow({
                    'tipo_fila': 'comida',
                    'plan_nombre': plan.name,
                    **empty_plan,
                    'comida_nombre': meal.name,
                    'tipo_comida': meal_type_labels.get(meal.meal_type, meal.meal_type),
                    'dia_semana': meal.day_of_week if meal.day_of_week is not None else '',
                    'orden_comida': meal.order_index if meal.order_index is not None else '',
                    'hora': str(meal.time) if meal.time else '',
                    'calorias_comida': meal.calories or 0,
                    'proteinas_comida': float(meal.protein or 0),
                    'carbohidratos_comida': float(meal.carbs or 0),
                    'grasas_comida': float(meal.fat or 0),
                    'descripcion_comida': meal.description or '',
                    'recetas_sugeridas': recetas_str,
                    **empty_option,
                    **empty_assignment,
                })

                for option in meal.meal_recipes.all().order_by('display_order', 'created_at'):
                    writer.writerow({
                        'tipo_fila': 'opcion_receta',
                        'plan_nombre': plan.name,
                        **empty_plan,
                        'comida_nombre': meal.name,
                        'tipo_comida': meal_type_labels.get(meal.meal_type, meal.meal_type),
                        'dia_semana': meal.day_of_week if meal.day_of_week is not None else '',
                        'orden_comida': meal.order_index if meal.order_index is not None else '',
                        'hora': str(meal.time) if meal.time else '',
                        'calorias_comida': meal.calories or 0,
                        'proteinas_comida': float(meal.protein or 0),
                        'carbohidratos_comida': float(meal.carbs or 0),
                        'grasas_comida': float(meal.fat or 0),
                        'descripcion_comida': meal.description or '',
                        'recetas_sugeridas': recetas_str,
                        'receta_nombre': option.recipe.name,
                        'imagen_url': option.recipe.image_url or '',
                        'orden_visualizacion': option.display_order,
                        'porciones': float(option.servings or 1),
                        'calorias_personalizadas': option.custom_calories if option.custom_calories is not None else '',
                        'proteinas_personalizadas': float(option.custom_protein) if option.custom_protein is not None else '',
                        'carbohidratos_personalizados': float(option.custom_carbs) if option.custom_carbs is not None else '',
                        'grasas_personalizadas': float(option.custom_fat) if option.custom_fat is not None else '',
                        **empty_assignment,
                    })

            for assignment in plan.assignments.select_related('user').all().order_by('user__email'):
                writer.writerow({
                    'tipo_fila': 'asignacion_usuario',
                    'plan_nombre': plan.name,
                    **empty_plan,
                    **empty_meal,
                    **empty_option,
                    'usuario_email': assignment.user.email,
                    'asignacion_activa': 'Si' if assignment.is_active else 'No',
                })
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporta planes de menú en Excel con paridad completa del panel admin.

        Hojas:
        1. Planes
        2. Comidas
        3. Opciones_Receta_Comida
        4. Asignaciones_Usuarios
        5. Recetas_Disponibles
        6. Usuarios_Disponibles
        7. Referencias
        """
        import io
        import xlsxwriter
        from django.http import HttpResponse
        plans = list(NutritionPlan.objects.prefetch_related(
            'meals__meal_recipes__recipe',
            'meals__suggested_recipes',
            'assignments__user',
        ).order_by('name'))
        goal_labels = {value: label for value, label in NutritionPlan.GOAL_CHOICES}
        diet_labels = {value: label for value, label in NutritionPlan.DIET_TYPE_CHOICES}
        meal_type_labels = {value: label for value, label in PlanMeal.MEAL_TYPE_CHOICES}
        meal_type_labels.update({
            'morning_snack': 'Merienda',
            'afternoon_snack': 'Merienda',
            'pre_workout': 'Merienda',
            'post_workout': 'Merienda',
        })
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'strings_to_urls': False})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        ref_format = workbook.add_format({'bold': True, 'bg_color': '#C6EFCE', 'italic': True})

        # ========== HOJA 1: PLANES ==========
        ws_plans = workbook.add_worksheet('Planes')
        plan_headers = [
            'plan_nombre', 'descripcion', 'objetivo', 'tipo_dieta',
            'calorias_diarias', 'proteinas_g', 'carbohidratos_g', 'grasas_g', 'fibra_g',
            'porcentaje_proteinas', 'porcentaje_carbohidratos', 'porcentaje_grasas',
            'comidas_por_dia', 'duracion_semanas', 'multiplicador_porcion',
            'es_plantilla', 'es_sistema', 'activo',
        ]
        for col, h in enumerate(plan_headers):
            ws_plans.write(0, col, h, header_format)
        for row_idx, plan in enumerate(plans, start=1):
            ws_plans.write(row_idx, 0, plan.name)
            ws_plans.write(row_idx, 1, plan.description or '')
            ws_plans.write(row_idx, 2, goal_labels.get(plan.goal, plan.goal or 'Mantener peso'))
            ws_plans.write(row_idx, 3, diet_labels.get(plan.diet_type, plan.diet_type or 'Normal'))
            ws_plans.write(row_idx, 4, plan.daily_calories or 0)
            ws_plans.write(row_idx, 5, plan.protein_grams or 0)
            ws_plans.write(row_idx, 6, plan.carbs_grams or 0)
            ws_plans.write(row_idx, 7, plan.fat_grams or 0)
            ws_plans.write(row_idx, 8, plan.fiber_grams or 0)
            ws_plans.write(row_idx, 9, plan.protein_percentage if plan.protein_percentage is not None else '')
            ws_plans.write(row_idx, 10, plan.carbs_percentage if plan.carbs_percentage is not None else '')
            ws_plans.write(row_idx, 11, plan.fat_percentage if plan.fat_percentage is not None else '')
            ws_plans.write(row_idx, 12, plan.meals_per_day or 5)
            ws_plans.write(row_idx, 13, plan.duration_weeks or 4)
            ws_plans.write(row_idx, 14, float(plan.portion_multiplier or 1.0))
            ws_plans.write(row_idx, 15, 'Si' if plan.is_template else 'No')
            ws_plans.write(row_idx, 16, 'Si' if plan.is_system else 'No')
            ws_plans.write(row_idx, 17, 'Si' if plan.is_active else 'No')
        ws_plans.set_column('A:A', 32)
        ws_plans.set_column('B:B', 35)
        ws_plans.set_column('C:C', 16)

        # ========== HOJA 2: COMIDAS ==========
        ws_meals = workbook.add_worksheet('Comidas')
        meal_headers = [
            'plan_nombre', 'dia_semana', 'orden_comida', 'comida_nombre', 'tipo_comida', 'hora',
            'calorias_comida', 'proteinas_comida', 'carbohidratos_comida', 'grasas_comida',
            'descripcion_comida', 'recetas_sugeridas',
        ]
        for col, h in enumerate(meal_headers):
            ws_meals.write(0, col, h, header_format)
        meal_row = 1
        for plan in plans:
            for meal in plan.meals.order_by('day_of_week', 'order_index'):
                recipe_names = set()
                for mr in meal.meal_recipes.all():
                    recipe_names.add(mr.recipe.name)
                for sr in meal.suggested_recipes.all():
                    recipe_names.add(sr.name)
                recetas_str = '; '.join(sorted(recipe_names))
                ws_meals.write(meal_row, 0, plan.name)
                ws_meals.write(meal_row, 1, meal.day_of_week if meal.day_of_week is not None else '')
                ws_meals.write(meal_row, 2, meal.order_index if meal.order_index is not None else '')
                ws_meals.write(meal_row, 3, meal.name)
                ws_meals.write(meal_row, 4, meal_type_labels.get(meal.meal_type, meal.meal_type))
                ws_meals.write(meal_row, 5, str(meal.time) if meal.time else '')
                ws_meals.write(meal_row, 6, meal.calories or 0)
                ws_meals.write(meal_row, 7, float(meal.protein or 0))
                ws_meals.write(meal_row, 8, float(meal.carbs or 0))
                ws_meals.write(meal_row, 9, float(meal.fat or 0))
                ws_meals.write(meal_row, 10, meal.description or '')
                ws_meals.write(meal_row, 11, recetas_str)
                meal_row += 1
        ws_meals.set_column('A:A', 30)
        ws_meals.set_column('D:D', 28)
        ws_meals.set_column('L:L', 50)

        # ========== HOJA 3: OPCIONES RECETA COMIDA ==========
        ws_options = workbook.add_worksheet('Opciones_Receta_Comida')
        option_headers = [
            'plan_nombre', 'dia_semana', 'orden_comida', 'comida_nombre', 'tipo_comida',
            'receta_nombre', 'imagen_url', 'orden_visualizacion', 'porciones',
            'calorias_personalizadas', 'proteinas_personalizadas', 'carbohidratos_personalizados', 'grasas_personalizadas',
            'categoria_receta', 'tipos_comida_receta',
        ]
        for col, h in enumerate(option_headers):
            ws_options.write(0, col, h, header_format)
        option_row = 1
        for plan in plans:
            for meal in plan.meals.order_by('day_of_week', 'order_index'):
                for option in meal.meal_recipes.all().order_by('display_order', 'created_at'):
                    ws_options.write(option_row, 0, plan.name)
                    ws_options.write(option_row, 1, meal.day_of_week if meal.day_of_week is not None else '')
                    ws_options.write(option_row, 2, meal.order_index if meal.order_index is not None else '')
                    ws_options.write(option_row, 3, meal.name)
                    ws_options.write(option_row, 4, meal_type_labels.get(meal.meal_type, meal.meal_type))
                    ws_options.write(option_row, 5, option.recipe.name)
                    ws_options.write(option_row, 6, option.recipe.image_url or '')
                    ws_options.write(option_row, 7, option.display_order if option.display_order is not None else 0)
                    ws_options.write(option_row, 8, float(option.servings or 1))
                    ws_options.write(option_row, 9, option.custom_calories if option.custom_calories is not None else '')
                    ws_options.write(option_row, 10, float(option.custom_protein) if option.custom_protein is not None else '')
                    ws_options.write(option_row, 11, float(option.custom_carbs) if option.custom_carbs is not None else '')
                    ws_options.write(option_row, 12, float(option.custom_fat) if option.custom_fat is not None else '')
                    ws_options.write(option_row, 13, option.recipe.category or '')
                    meal_types_readable = []
                    for mt in (option.recipe.meal_types or []):
                        meal_types_readable.append(meal_type_labels.get(mt, mt))
                    ws_options.write(option_row, 14, ','.join(meal_types_readable))
                    option_row += 1
        ws_options.set_column('A:A', 30)
        ws_options.set_column('D:D', 28)
        ws_options.set_column('F:F', 34)
        ws_options.set_column('N:O', 22)

        # ========== HOJA 4: ASIGNACIONES USUARIOS ==========
        ws_assign = workbook.add_worksheet('Asignaciones_Usuarios')
        assign_headers = ['plan_nombre', 'usuario_email', 'asignacion_activa']
        for col, h in enumerate(assign_headers):
            ws_assign.write(0, col, h, header_format)
        assign_row = 1
        for plan in plans:
            for assignment in plan.assignments.select_related('user').all().order_by('user__email'):
                ws_assign.write(assign_row, 0, plan.name)
                ws_assign.write(assign_row, 1, assignment.user.email)
                ws_assign.write(assign_row, 2, 'Si' if assignment.is_active else 'No')
                assign_row += 1
        ws_assign.set_column('A:A', 30)
        ws_assign.set_column('B:B', 36)

        # ========== HOJA 5: RECETAS DISPONIBLES ==========
        from .models import Recipe
        ws_ref = workbook.add_worksheet('Recetas_Disponibles')
        ref_headers = ['nombre_receta', 'categoria', 'dificultad', 'calorias', 'proteinas', 'carbohidratos', 'grasas']
        for col, h in enumerate(ref_headers):
            ws_ref.write(0, col, h, ref_format)
        recipes = Recipe.objects.order_by('name').values_list('name', 'category', 'difficulty', 'calories', 'protein', 'carbs', 'fat')
        for r_idx, (name, cat, diff, cal, prot, carbs, fat) in enumerate(recipes, start=1):
            ws_ref.write(r_idx, 0, name or '')
            ws_ref.write(r_idx, 1, cat or '')
            ws_ref.write(r_idx, 2, diff or '')
            ws_ref.write(r_idx, 3, cal or 0)
            ws_ref.write(r_idx, 4, float(prot or 0))
            ws_ref.write(r_idx, 5, float(carbs or 0))
            ws_ref.write(r_idx, 6, float(fat or 0))
        ws_ref.set_column('A:A', 30)
        ws_ref.set_column('B:B', 20)

        # ========== HOJA 6: USUARIOS DISPONIBLES ==========
        ws_users = workbook.add_worksheet('Usuarios_Disponibles')
        user_headers = ['usuario_email', 'usuario_id', 'is_active', 'is_staff']
        for col, h in enumerate(user_headers):
            ws_users.write(0, col, h, ref_format)
        users = User.objects.order_by('email').values_list('email', 'id', 'is_active', 'is_staff')
        for u_idx, (email, uid, is_active, is_staff) in enumerate(users, start=1):
            ws_users.write(u_idx, 0, email or '')
            ws_users.write(u_idx, 1, str(uid))
            ws_users.write(u_idx, 2, 'Si' if is_active else 'No')
            ws_users.write(u_idx, 3, 'Si' if is_staff else 'No')
        ws_users.set_column('A:A', 36)

        # ========== HOJA 7: REFERENCIAS ==========
        ws_refs = workbook.add_worksheet('Referencias')
        refs_headers = ['tipo', 'valor', 'descripcion']
        for col, h in enumerate(refs_headers):
            ws_refs.write(0, col, h, ref_format)
        ref_rows = []
        ref_rows.extend([('objetivo', l, f'Objetivo nutricional: {l}') for _, l in NutritionPlan.GOAL_CHOICES])
        ref_rows.extend([('tipo_dieta', l, f'Tipo de dieta: {l}') for _, l in NutritionPlan.DIET_TYPE_CHOICES])
        ref_rows.extend([('tipo_comida', l, f'Tipo de comida: {l}') for _, l in PlanMeal.MEAL_TYPE_CHOICES])
        ref_rows.extend([
            ('booleano', 'Si/No', 'También acepta true/false, 1/0, yes/no'),
            ('dia_semana', '1..7', '1=Lunes .. 7=Domingo'),
            ('hora', 'HH:MM', 'Ejemplo: 08:30'),
            ('opciones_receta', 'porciones>0', 'campos personalizados opcionales (vacío = usar receta base)'),
            ('imagen_url', 'https://...', 'URL de imagen de receta (informativo y respaldo en importación)')
        ])
        for idx, (rtype, value, desc) in enumerate(ref_rows, start=1):
            ws_refs.write(idx, 0, rtype)
            ws_refs.write(idx, 1, value)
            ws_refs.write(idx, 2, desc)
        ws_refs.set_column('A:A', 20)
        ws_refs.set_column('B:B', 30)
        ws_refs.set_column('C:C', 60)

        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="planes_menu_exportacion.xlsx"'
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    @action(detail=False, methods=['post'], url_path='import-csv',
            parser_classes=[MultiPartParser, FormParser])
    def import_csv(self, request):
        """Importa planes de menú desde CSV con validación estricta por fila.

        Regla estricta: si una fila tiene un valor inválido, no se aplica ningún
        cambio de esa fila (ni create ni update).
        """
        import csv
        import unicodedata
        from decimal import Decimal
        from datetime import time
        from django.core.files.uploadedfile import UploadedFile

        goal_values = {v for v, _ in NutritionPlan.GOAL_CHOICES}
        diet_values = {v for v, _ in NutritionPlan.DIET_TYPE_CHOICES}
        meal_type_values = {v for v, _ in PlanMeal.MEAL_TYPE_CHOICES}

        def normalize_text(value):
            text = str(value or '').strip().lower()
            text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
            return text.replace('-', '_').replace(' ', '_')

        goal_input_map = {
            'lose_weight': 'lose_weight',
            'perder_peso': 'lose_weight',
            'gain_muscle': 'gain_muscle',
            'ganar_musculo': 'gain_muscle',
            'maintain': 'maintain',
            'mantener_peso': 'maintain',
            'body_recomposition': 'body_recomposition',
            'recomposicion_corporal': 'body_recomposition',
            'performance': 'performance',
            'rendimiento_deportivo': 'performance',
        }

        diet_input_map = {
            'normal': 'normal',
            'vegetarian': 'vegetarian',
            'vegetariano': 'vegetarian',
            'vegan': 'vegan',
            'vegano': 'vegan',
            'keto': 'keto',
            'paleo': 'paleo',
            'mediterranean': 'mediterranean',
            'mediterranea': 'mediterranean',
            'low_carb': 'low_carb',
            'bajo_en_carbohidratos': 'low_carb',
            'high_protein': 'high_protein',
            'alto_en_proteinas': 'high_protein',
        }

        meal_type_input_map = {
            'breakfast': 'breakfast',
            'desayuno': 'breakfast',
            'lunch': 'lunch',
            'comida': 'lunch',
            'almuerzo': 'lunch',
            'snack': 'snack',
            'merienda': 'snack',
            'morning_snack': 'snack',
            'afternoon_snack': 'snack',
            'pre_workout': 'snack',
            'post_workout': 'snack',
            'snack_manana': 'snack',
            'snack_tarde': 'snack',
            'media_manana': 'snack',
            'media_tarde': 'snack',
            'dinner': 'dinner',
            'cena': 'dinner',
        }

        field_aliases = {
            'nombre': 'name', 'name': 'name',
            'plan_nombre': 'name',
            'descripcion': 'description', 'description': 'description',
            'objetivo': 'goal', 'goal': 'goal',
            'tipo_dieta': 'diet_type', 'diet_type': 'diet_type',
            'calorias_diarias': 'daily_calories', 'daily_calories': 'daily_calories',
            'proteinas_g': 'protein_grams', 'protein_grams': 'protein_grams',
            'carbohidratos_g': 'carbs_grams', 'carbs_grams': 'carbs_grams',
            'grasas_g': 'fat_grams', 'fat_grams': 'fat_grams',
            'fibra_g': 'fiber_grams', 'fiber_grams': 'fiber_grams',
            'porcentaje_proteinas': 'protein_percentage', 'protein_percentage': 'protein_percentage',
            'porcentaje_carbohidratos': 'carbs_percentage', 'carbs_percentage': 'carbs_percentage',
            'porcentaje_grasas': 'fat_percentage', 'fat_percentage': 'fat_percentage',
            'comidas_por_dia': 'meals_per_day', 'meals_per_day': 'meals_per_day',
            'duracion_semanas': 'duration_weeks', 'duration_weeks': 'duration_weeks',
            'multiplicador_porcion': 'portion_multiplier', 'portion_multiplier': 'portion_multiplier',
            'es_plantilla': 'is_template', 'is_template': 'is_template',
            'es_sistema': 'is_system', 'is_system': 'is_system',
            'activo': 'is_active', 'is_active': 'is_active',
        }

        def gv(row, canonical, default=''):
            if canonical in row:
                return row[canonical]
            for k, v in field_aliases.items():
                if v == canonical and k in row:
                    return row[k]
            return default

        def parse_bool(val):
            raw = str(val).strip().lower()
            if raw in ('true', 'si', 'sí', 's', '1', 'yes', 'y'):
                return True, None
            if raw in ('false', 'no', 'n', '0'):
                return False, None
            return None, f"Valor booleano inválido: {val}"

        def parse_int(val, field_name, minimum=None, maximum=None, allow_blank=False):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                if allow_blank:
                    return None, None
                return None, f"{field_name} es obligatorio"
            try:
                parsed = int(float(raw))
            except (TypeError, ValueError):
                return None, f"{field_name} inválido: {val}"
            if minimum is not None and parsed < minimum:
                return None, f"{field_name} debe ser >= {minimum}"
            if maximum is not None and parsed > maximum:
                return None, f"{field_name} debe ser <= {maximum}"
            return parsed, None

        def parse_float(val, field_name, minimum=None, allow_blank=False):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                if allow_blank:
                    return None, None
                return None, f"{field_name} es obligatorio"
            try:
                parsed = float(raw)
            except (TypeError, ValueError):
                return None, f"{field_name} inválido: {val}"
            if minimum is not None and parsed < minimum:
                return None, f"{field_name} debe ser >= {minimum}"
            return parsed, None

        def parse_time_hhmm(val):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                return None, None
            try:
                parts = raw.split(':')
                if len(parts) < 2:
                    raise ValueError()
                return time(int(parts[0]), int(parts[1])), None
            except (TypeError, ValueError):
                return None, f"Hora inválida: {val} (formato esperado HH:MM)"

        def parse_percent(val, field_name):
            if val is None or str(val).strip() == '':
                return None, None
            return parse_int(val, field_name, minimum=0, maximum=100)

        def parse_decimal_or_none(val, field_name):
            if val is None or str(val).strip() == '':
                return None, None
            parsed, err = parse_float(val, field_name, minimum=0)
            if err:
                return None, err
            return Decimal(str(parsed)), None

        def default_if_blank(value, default):
            if value is None:
                return default
            if isinstance(value, str) and value.strip() == '':
                return default
            return value

        def parse_plan_fields(row):
            errors = []
            raw_goal = str(gv(row, 'goal', 'maintain') or 'maintain').strip()
            goal = goal_input_map.get(normalize_text(raw_goal), raw_goal)
            if goal not in goal_values:
                errors.append(f"objetivo inválido: {raw_goal}")

            raw_diet_type = str(gv(row, 'diet_type', 'normal') or 'normal').strip()
            diet_type = diet_input_map.get(normalize_text(raw_diet_type), raw_diet_type)
            if diet_type not in diet_values:
                errors.append(f"tipo_dieta inválido: {raw_diet_type}")

            daily_calories, err = parse_int(default_if_blank(gv(row, 'daily_calories', 2000), 2000), 'calorias_diarias', minimum=0)
            if err:
                errors.append(err)
            protein_grams, err = parse_int(default_if_blank(gv(row, 'protein_grams', 150), 150), 'proteinas_g', minimum=0)
            if err:
                errors.append(err)
            carbs_grams, err = parse_int(default_if_blank(gv(row, 'carbs_grams', 200), 200), 'carbohidratos_g', minimum=0)
            if err:
                errors.append(err)
            fat_grams, err = parse_int(default_if_blank(gv(row, 'fat_grams', 65), 65), 'grasas_g', minimum=0)
            if err:
                errors.append(err)
            fiber_grams, err = parse_int(default_if_blank(gv(row, 'fiber_grams', 25), 25), 'fibra_g', minimum=0)
            if err:
                errors.append(err)

            protein_pct, err = parse_percent(gv(row, 'protein_percentage', ''), 'porcentaje_proteinas')
            if err:
                errors.append(err)
            carbs_pct, err = parse_percent(gv(row, 'carbs_percentage', ''), 'porcentaje_carbohidratos')
            if err:
                errors.append(err)
            fat_pct, err = parse_percent(gv(row, 'fat_percentage', ''), 'porcentaje_grasas')
            if err:
                errors.append(err)

            meals_per_day, err = parse_int(default_if_blank(gv(row, 'meals_per_day', 5), 5), 'comidas_por_dia', minimum=1, maximum=8)
            if err:
                errors.append(err)
            duration_weeks, err = parse_int(default_if_blank(gv(row, 'duration_weeks', 4), 4), 'duracion_semanas', minimum=1)
            if err:
                errors.append(err)
            portion_multiplier, err = parse_float(default_if_blank(gv(row, 'portion_multiplier', 1.0), 1.0), 'multiplicador_porcion', minimum=0.1)
            if err:
                errors.append(err)

            is_template, err = parse_bool(default_if_blank(gv(row, 'is_template', 'false'), 'false'))
            if err:
                errors.append(f"es_plantilla: {err}")
            is_system, err = parse_bool(default_if_blank(gv(row, 'is_system', 'false'), 'false'))
            if err:
                errors.append(f"es_sistema: {err}")
            is_active, err = parse_bool(default_if_blank(gv(row, 'is_active', 'true'), 'true'))
            if err:
                errors.append(f"activo: {err}")

            if errors:
                return None, errors

            return {
                'description': str(gv(row, 'description', '') or ''),
                'goal': goal,
                'diet_type': diet_type,
                'daily_calories': daily_calories,
                'protein_grams': protein_grams,
                'carbs_grams': carbs_grams,
                'fat_grams': fat_grams,
                'fiber_grams': fiber_grams,
                'protein_percentage': protein_pct,
                'carbs_percentage': carbs_pct,
                'fat_percentage': fat_pct,
                'meals_per_day': meals_per_day,
                'duration_weeks': duration_weeks,
                'portion_multiplier': portion_multiplier,
                'is_template': is_template,
                'is_system': is_system,
                'is_active': is_active,
            }, None

        def parse_meal_fields(row):
            errors = []
            raw_meal_type = str(row.get('tipo_comida', '') or 'lunch').strip() or 'lunch'
            meal_type = meal_type_input_map.get(normalize_text(raw_meal_type), raw_meal_type)
            if meal_type not in meal_type_values:
                errors.append(f"tipo_comida inválido: {raw_meal_type}")

            day_of_week, err = parse_int(row.get('dia_semana', ''), 'dia_semana', minimum=1, maximum=7, allow_blank=True)
            if err:
                errors.append(err)
            order_index, err = parse_int(row.get('orden_comida', ''), 'orden_comida', minimum=1, allow_blank=True)
            if err:
                errors.append(err)
            if order_index is None:
                errors.append('orden_comida es obligatorio')

            meal_time, err = parse_time_hhmm(row.get('hora', ''))
            if err:
                errors.append(err)
            calories, err = parse_int(row.get('calorias_comida', ''), 'calorias_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            protein, err = parse_float(row.get('proteinas_comida', ''), 'proteinas_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            carbs, err = parse_float(row.get('carbohidratos_comida', ''), 'carbohidratos_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            fat, err = parse_float(row.get('grasas_comida', ''), 'grasas_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)

            if errors:
                return None, errors

            return {
                'meal_type': meal_type,
                'day_of_week': day_of_week,
                'order_index': order_index,
                'time': meal_time,
                'calories': calories,
                'protein': Decimal(str(protein)) if protein is not None else None,
                'carbs': Decimal(str(carbs)) if carbs is not None else None,
                'fat': Decimal(str(fat)) if fat is not None else None,
                'name': str(row.get('comida_nombre', '') or '').strip() or f'Comida {order_index}',
                'description': str(row.get('descripcion_comida', '') or ''),
                'recetas_sugeridas': str(row.get('recetas_sugeridas', '') or '').strip(),
            }, None

        def infer_meal_macros_from_suggested_recipes(recetas_sugeridas):
            recipe_names = [name.strip() for name in str(recetas_sugeridas or '').split(';') if name.strip()]
            if not recipe_names:
                return None
            recipes = list(Recipe.objects.filter(name__in=recipe_names))
            if not recipes:
                return None

            def avg_numeric(attr):
                values = [float(getattr(recipe, attr) or 0) for recipe in recipes]
                return (sum(values) / len(values)) if values else 0.0

            return {
                'calories': int(round(avg_numeric('calories'))),
                'protein': Decimal(str(round(avg_numeric('protein'), 2))),
                'carbs': Decimal(str(round(avg_numeric('carbs'), 2))),
                'fat': Decimal(str(round(avg_numeric('fat'), 2))),
            }

        def parse_option_fields(row):
            def optv(data, *keys, default=''):
                for key in keys:
                    if key in data and data.get(key) not in (None, ''):
                        return data.get(key)
                return default

            errors = []
            display_order, err = parse_int(optv(row, 'orden_visualizacion', 'display_order', default=0), 'orden_visualizacion', minimum=0)
            if err:
                errors.append(err)
            servings, err = parse_float(optv(row, 'porciones', 'servings', default=1), 'porciones', minimum=0.01)
            if err:
                errors.append(err)
            custom_calories, err = parse_int(
                optv(row, 'calorias_personalizadas', 'custom_calories', default=''),
                'calorias_personalizadas',
                minimum=0,
                allow_blank=True,
            )
            if err:
                errors.append(err)
            custom_protein, err = parse_decimal_or_none(
                optv(row, 'proteinas_personalizadas', 'custom_protein', default=''),
                'proteinas_personalizadas',
            )
            if err:
                errors.append(err)
            custom_carbs, err = parse_decimal_or_none(
                optv(row, 'carbohidratos_personalizados', 'custom_carbs', default=''),
                'carbohidratos_personalizados',
            )
            if err:
                errors.append(err)
            custom_fat, err = parse_decimal_or_none(
                optv(row, 'grasas_personalizadas', 'custom_fat', default=''),
                'grasas_personalizadas',
            )
            if err:
                errors.append(err)

            if errors:
                return None, errors
            return {
                'display_order': display_order,
                'servings': Decimal(str(servings)),
                'custom_calories': custom_calories,
                'custom_protein': custom_protein,
                'custom_carbs': custom_carbs,
                'custom_fat': custom_fat,
            }, None

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo CSV no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)
        decoded = file.read().decode('utf-8')
        rows = list(csv.DictReader(decoded.splitlines()))
        updated, created, skipped, rejected = 0, 0, 0, 0
        meals_created, meals_updated, meals_skipped, meals_rejected = 0, 0, 0, 0
        options_created, options_updated, options_skipped, options_rejected = 0, 0, 0, 0
        assignments_created, assignments_updated, assignments_skipped, assignments_rejected = 0, 0, 0, 0
        errors = []

        # Pass 1: plan rows (or legacy rows without tipo_fila)
        for idx, row in enumerate(rows, start=2):
            tipo_fila = str(row.get('tipo_fila', '') or '').strip().lower()
            if tipo_fila not in ('', 'plan'):
                continue
            name = str(gv(row, 'name', '') or '').strip()
            if not name:
                skipped += 1
                continue
            fields, row_errors = parse_plan_fields(row)
            if row_errors:
                rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'plan', 'errors': row_errors})
                continue
            plan = NutritionPlan.objects.filter(name=name).first()
            if plan:
                for k, v in fields.items():
                    setattr(plan, k, v)
                plan.save()
                updated += 1
            else:
                NutritionPlan.objects.create(name=name, **fields)
                created += 1

        # Pass 2: meal rows
        for idx, row in enumerate(rows, start=2):
            tipo_fila = str(row.get('tipo_fila', '') or '').strip().lower()
            if tipo_fila != 'comida':
                continue
            plan_name = str(row.get('plan_nombre', '') or '').strip()
            if not plan_name:
                meals_skipped += 1
                continue
            plan = NutritionPlan.objects.filter(name=plan_name).first()
            if not plan:
                plan = NutritionPlan.objects.create(
                    name=plan_name,
                    description='',
                    goal='maintain',
                    diet_type='normal',
                    daily_calories=2000,
                    protein_grams=150,
                    carbs_grams=200,
                    fat_grams=65,
                    fiber_grams=25,
                    protein_percentage=30,
                    carbs_percentage=40,
                    fat_percentage=30,
                    meals_per_day=5,
                    duration_weeks=4,
                    portion_multiplier=1.0,
                    is_template=True,
                    is_system=False,
                    is_active=True,
                )
                created += 1
            parsed, row_errors = parse_meal_fields(row)
            if row_errors:
                meals_rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'comida', 'errors': row_errors})
                continue

            inferred = infer_meal_macros_from_suggested_recipes(parsed['recetas_sugeridas'])
            calories = parsed['calories'] if parsed['calories'] is not None else (inferred['calories'] if inferred else 0)
            protein = parsed['protein'] if parsed['protein'] is not None else (inferred['protein'] if inferred else Decimal('0'))
            carbs = parsed['carbs'] if parsed['carbs'] is not None else (inferred['carbs'] if inferred else Decimal('0'))
            fat = parsed['fat'] if parsed['fat'] is not None else (inferred['fat'] if inferred else Decimal('0'))

            meal = PlanMeal.objects.filter(
                plan=plan,
                day_of_week=parsed['day_of_week'],
                order_index=parsed['order_index'],
            ).first()
            meal_fields = {
                'name': parsed['name'],
                'meal_type': parsed['meal_type'],
                'time': parsed['time'],
                'calories': calories,
                'protein': protein,
                'carbs': carbs,
                'fat': fat,
                'description': parsed['description'],
            }
            if meal:
                for k, v in meal_fields.items():
                    setattr(meal, k, v)
                meal.save()
                meals_updated += 1
            else:
                PlanMeal.objects.create(
                    plan=plan,
                    day_of_week=parsed['day_of_week'],
                    order_index=parsed['order_index'],
                    **meal_fields,
                )
                meals_created += 1

        # Pass 3: recipe-option rows
        option_recipes_by_meal = {}
        for idx, row in enumerate(rows, start=2):
            tipo_fila = str(row.get('tipo_fila', '') or '').strip().lower()
            if tipo_fila != 'opcion_receta':
                continue

            def _csv_optv(data, *keys, default=''):
                for key in keys:
                    if key in data and data.get(key) not in (None, ''):
                        return data.get(key)
                return default

            plan_name = str(row.get('plan_nombre', '') or '').strip()
            recipe_name = str(row.get('receta_nombre', '') or '').strip()
            image_url = str(_csv_optv(row, 'imagen_url', 'image_url', 'url_imagen', default='') or '').strip()
            if not plan_name or (not recipe_name and not image_url):
                options_skipped += 1
                continue
            plan = NutritionPlan.objects.filter(name=plan_name).first()
            if not plan:
                plan = NutritionPlan.objects.create(
                    name=plan_name,
                    description='',
                    goal='maintain',
                    diet_type='normal',
                    daily_calories=2000,
                    protein_grams=150,
                    carbs_grams=200,
                    fat_grams=65,
                    fiber_grams=25,
                    protein_percentage=30,
                    carbs_percentage=40,
                    fat_percentage=30,
                    meals_per_day=5,
                    duration_weeks=4,
                    portion_multiplier=1.0,
                    is_template=True,
                    is_system=False,
                    is_active=True,
                )
                created += 1

            day_of_week, err = parse_int(row.get('dia_semana', ''), 'dia_semana', minimum=1, maximum=7, allow_blank=True)
            order_index, err2 = parse_int(row.get('orden_comida', ''), 'orden_comida', minimum=1)
            local_errors = []
            if err:
                local_errors.append(err)
            if err2:
                local_errors.append(err2)
            parsed_option, option_errors = parse_option_fields(row)
            if option_errors:
                local_errors.extend(option_errors)
            meal = None
            if not local_errors:
                meal = PlanMeal.objects.filter(plan=plan, day_of_week=day_of_week, order_index=order_index).first()
                if not meal:
                    local_errors.append(f"No existe comida para clave ({plan_name}, día={day_of_week}, orden={order_index})")
            recipe = None
            if not local_errors:
                if recipe_name:
                    recipe = Recipe.objects.filter(name=recipe_name).first()
                if not recipe and image_url:
                    recipe = Recipe.objects.filter(image_url=image_url).first()
                if not recipe:
                    if recipe_name and image_url:
                        local_errors.append(f"Receta no existe: nombre={recipe_name} ni imagen_url={image_url}")
                    elif recipe_name:
                        local_errors.append(f"Receta no existe: {recipe_name}")
                    else:
                        local_errors.append(f"Receta no existe para imagen_url: {image_url}")

            if local_errors:
                options_rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'opcion_receta', 'errors': local_errors})
                continue

            option = PlanMealRecipe.objects.filter(meal=meal, recipe=recipe).first()
            if option:
                for k, v in parsed_option.items():
                    setattr(option, k, v)
                option.save()
                options_updated += 1
            else:
                PlanMealRecipe.objects.create(meal=meal, recipe=recipe, **parsed_option)
                options_created += 1

            option_recipes_by_meal.setdefault(meal.id, set()).add(recipe.id)

        # Sync suggested recipes with option rows to keep panel parity
        if option_recipes_by_meal:
            for meal_id, recipe_ids in option_recipes_by_meal.items():
                meal = PlanMeal.objects.filter(id=meal_id).first()
                if meal:
                    meal.suggested_recipes.set(Recipe.objects.filter(id__in=recipe_ids))

        # Pass 4: assignment rows
        for idx, row in enumerate(rows, start=2):
            tipo_fila = str(row.get('tipo_fila', '') or '').strip().lower()
            if tipo_fila != 'asignacion_usuario':
                continue
            plan_name = str(row.get('plan_nombre', '') or '').strip()
            email = str(row.get('usuario_email', '') or '').strip().lower()
            if not plan_name or not email:
                assignments_skipped += 1
                continue
            plan = NutritionPlan.objects.filter(name=plan_name).first()
            if not plan:
                assignments_rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'asignacion_usuario', 'errors': [f"Plan no existe: {plan_name}"]})
                continue
            user = User.objects.filter(email__iexact=email).first()
            if not user:
                assignments_rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'asignacion_usuario', 'errors': [f"Usuario no existe: {email}"]})
                continue
            is_active, err = parse_bool(row.get('asignacion_activa', 'true'))
            if err:
                assignments_rejected += 1
                errors.append({'sheet': 'CSV', 'row': idx, 'type': 'asignacion_usuario', 'errors': [err]})
                continue
            assignment = NutritionPlanAssignment.objects.filter(plan=plan, user=user).first()
            if assignment:
                assignment.is_active = is_active
                assignment.save(update_fields=['is_active'])
                assignments_updated += 1
            else:
                NutritionPlanAssignment.objects.create(plan=plan, user=user, is_active=is_active)
                assignments_created += 1

        # Backward compatibility: if there are no option rows, build defaults from recetas_sugeridas in meal rows
        has_option_rows = any(str((r.get('tipo_fila') or '')).strip().lower() == 'opcion_receta' for r in rows)
        if not has_option_rows:
            for row in rows:
                tipo_fila = str(row.get('tipo_fila', '') or '').strip().lower()
                if tipo_fila != 'comida':
                    continue
                plan_name = str(row.get('plan_nombre', '') or '').strip()
                if not plan_name:
                    continue
                day_of_week, _ = parse_int(row.get('dia_semana', ''), 'dia_semana', minimum=1, maximum=7, allow_blank=True)
                order_index, err = parse_int(row.get('orden_comida', ''), 'orden_comida', minimum=1)
                if err:
                    continue
                plan = NutritionPlan.objects.filter(name=plan_name).first()
                if not plan:
                    continue
                meal = PlanMeal.objects.filter(plan=plan, day_of_week=day_of_week, order_index=order_index).first()
                if not meal:
                    continue
                recetas_str = str(row.get('recetas_sugeridas', '') or '').strip()
                if not recetas_str:
                    continue
                recipe_names = [name.strip() for name in recetas_str.split(';') if name.strip()]
                display = 0
                linked = []
                for recipe_name in recipe_names:
                    recipe = Recipe.objects.filter(name=recipe_name).first()
                    if not recipe:
                        continue
                    PlanMealRecipe.objects.get_or_create(
                        meal=meal,
                        recipe=recipe,
                        defaults={'servings': Decimal('1.0'), 'display_order': display}
                    )
                    linked.append(recipe.id)
                    display += 1
                if linked:
                    meal.suggested_recipes.set(Recipe.objects.filter(id__in=linked))

        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'rejected': rejected,
            'meals_created': meals_created,
            'meals_updated': meals_updated,
            'meals_skipped': meals_skipped,
            'meals_rejected': meals_rejected,
            'options_created': options_created,
            'options_updated': options_updated,
            'options_skipped': options_skipped,
            'options_rejected': options_rejected,
            'assignments_created': assignments_created,
            'assignments_updated': assignments_updated,
            'assignments_skipped': assignments_skipped,
            'assignments_rejected': assignments_rejected,
            'errors': errors,
            'message': (
                f"Importación CSV completada. Planes: {created} creados, {updated} actualizados, {skipped} omitidos, {rejected} rechazados. "
                f"Comidas: {meals_created} creadas, {meals_updated} actualizadas, {meals_skipped} omitidas, {meals_rejected} rechazadas. "
                f"Opciones: {options_created} creadas, {options_updated} actualizadas, {options_skipped} omitidas, {options_rejected} rechazadas. "
                f"Asignaciones: {assignments_created} creadas, {assignments_updated} actualizadas, {assignments_skipped} omitidas, {assignments_rejected} rechazadas."
            ),
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel',
            parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """Importa planes desde Excel con validación estricta por fila.

        Si una fila tiene errores, no se aplica ni crea/modifica nada para esa fila.
        """
        import openpyxl
        import unicodedata
        from decimal import Decimal
        from datetime import time
        from django.core.files.uploadedfile import UploadedFile

        goal_values = {v for v, _ in NutritionPlan.GOAL_CHOICES}
        diet_values = {v for v, _ in NutritionPlan.DIET_TYPE_CHOICES}
        meal_type_values = {v for v, _ in PlanMeal.MEAL_TYPE_CHOICES}

        def normalize_text(value):
            text = str(value or '').strip().lower()
            text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
            return text.replace('-', '_').replace(' ', '_')

        goal_input_map = {
            'lose_weight': 'lose_weight',
            'perder_peso': 'lose_weight',
            'gain_muscle': 'gain_muscle',
            'ganar_musculo': 'gain_muscle',
            'maintain': 'maintain',
            'mantener_peso': 'maintain',
            'body_recomposition': 'body_recomposition',
            'recomposicion_corporal': 'body_recomposition',
            'performance': 'performance',
            'rendimiento_deportivo': 'performance',
        }

        diet_input_map = {
            'normal': 'normal',
            'vegetarian': 'vegetarian',
            'vegetariano': 'vegetarian',
            'vegan': 'vegan',
            'vegano': 'vegan',
            'keto': 'keto',
            'paleo': 'paleo',
            'mediterranean': 'mediterranean',
            'mediterranea': 'mediterranean',
            'low_carb': 'low_carb',
            'bajo_en_carbohidratos': 'low_carb',
            'high_protein': 'high_protein',
            'alto_en_proteinas': 'high_protein',
        }

        meal_type_input_map = {
            'breakfast': 'breakfast',
            'desayuno': 'breakfast',
            'lunch': 'lunch',
            'comida': 'lunch',
            'almuerzo': 'lunch',
            'snack': 'snack',
            'merienda': 'snack',
            'morning_snack': 'snack',
            'afternoon_snack': 'snack',
            'pre_workout': 'snack',
            'post_workout': 'snack',
            'snack_manana': 'snack',
            'snack_tarde': 'snack',
            'media_manana': 'snack',
            'media_tarde': 'snack',
            'dinner': 'dinner',
            'cena': 'dinner',
        }

        plan_aliases = {
            'nombre': 'name', 'name': 'name',
            'plan_nombre': 'name',
            'descripcion': 'description', 'description': 'description',
            'objetivo': 'goal', 'goal': 'goal',
            'tipo_dieta': 'diet_type', 'diet_type': 'diet_type',
            'calorias_diarias': 'daily_calories', 'daily_calories': 'daily_calories',
            'proteinas_g': 'protein_grams', 'protein_grams': 'protein_grams',
            'carbohidratos_g': 'carbs_grams', 'carbs_grams': 'carbs_grams',
            'grasas_g': 'fat_grams', 'fat_grams': 'fat_grams',
            'fibra_g': 'fiber_grams', 'fiber_grams': 'fiber_grams',
            'porcentaje_proteinas': 'protein_percentage', 'protein_percentage': 'protein_percentage',
            'porcentaje_carbohidratos': 'carbs_percentage', 'carbs_percentage': 'carbs_percentage',
            'porcentaje_grasas': 'fat_percentage', 'fat_percentage': 'fat_percentage',
            'comidas_por_dia': 'meals_per_day', 'meals_per_day': 'meals_per_day',
            'duracion_semanas': 'duration_weeks', 'duration_weeks': 'duration_weeks',
            'multiplicador_porcion': 'portion_multiplier', 'portion_multiplier': 'portion_multiplier',
            'es_plantilla': 'is_template', 'is_template': 'is_template',
            'es_sistema': 'is_system', 'is_system': 'is_system',
            'activo': 'is_active', 'is_active': 'is_active',
        }

        def gv(row_dict, canonical, default=''):
            if canonical in row_dict:
                return row_dict[canonical]
            for k, v in plan_aliases.items():
                if v == canonical and k in row_dict:
                    return row_dict[k]
            return default

        def parse_bool(val):
            raw = str(val).strip().lower()
            if raw in ('true', 'si', 'sí', 's', '1', 'yes', 'y'):
                return True, None
            if raw in ('false', 'no', 'n', '0'):
                return False, None
            return None, f"Valor booleano inválido: {val}"

        def parse_int(val, field_name, minimum=None, maximum=None, allow_blank=False):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                if allow_blank:
                    return None, None
                return None, f"{field_name} es obligatorio"
            try:
                parsed = int(float(raw))
            except (TypeError, ValueError):
                return None, f"{field_name} inválido: {val}"
            if minimum is not None and parsed < minimum:
                return None, f"{field_name} debe ser >= {minimum}"
            if maximum is not None and parsed > maximum:
                return None, f"{field_name} debe ser <= {maximum}"
            return parsed, None

        def parse_float(val, field_name, minimum=None, allow_blank=False):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                if allow_blank:
                    return None, None
                return None, f"{field_name} es obligatorio"
            try:
                parsed = float(raw)
            except (TypeError, ValueError):
                return None, f"{field_name} inválido: {val}"
            if minimum is not None and parsed < minimum:
                return None, f"{field_name} debe ser >= {minimum}"
            return parsed, None

        def parse_time_hhmm(val):
            raw = '' if val is None else str(val).strip()
            if raw == '':
                return None, None
            try:
                parts = raw.split(':')
                if len(parts) < 2:
                    raise ValueError()
                return time(int(parts[0]), int(parts[1])), None
            except (TypeError, ValueError):
                return None, f"Hora inválida: {val} (formato esperado HH:MM)"

        def parse_percent(val, field_name):
            if val is None or str(val).strip() == '':
                return None, None
            return parse_int(val, field_name, minimum=0, maximum=100)

        def parse_decimal_or_none(val, field_name):
            if val is None or str(val).strip() == '':
                return None, None
            parsed, err = parse_float(val, field_name, minimum=0)
            if err:
                return None, err
            return Decimal(str(parsed)), None

        def default_if_blank(value, default):
            if value is None:
                return default
            if isinstance(value, str) and value.strip() == '':
                return default
            return value

        def parse_plan_fields(row_dict):
            errors = []
            raw_goal = str(gv(row_dict, 'goal', 'maintain') or 'maintain').strip()
            goal = goal_input_map.get(normalize_text(raw_goal), raw_goal)
            if goal not in goal_values:
                errors.append(f"objetivo inválido: {raw_goal}")
            raw_diet_type = str(gv(row_dict, 'diet_type', 'normal') or 'normal').strip()
            diet_type = diet_input_map.get(normalize_text(raw_diet_type), raw_diet_type)
            if diet_type not in diet_values:
                errors.append(f"tipo_dieta inválido: {raw_diet_type}")

            daily_calories, err = parse_int(default_if_blank(gv(row_dict, 'daily_calories', 2000), 2000), 'calorias_diarias', minimum=0)
            if err:
                errors.append(err)
            protein_grams, err = parse_int(default_if_blank(gv(row_dict, 'protein_grams', 150), 150), 'proteinas_g', minimum=0)
            if err:
                errors.append(err)
            carbs_grams, err = parse_int(default_if_blank(gv(row_dict, 'carbs_grams', 200), 200), 'carbohidratos_g', minimum=0)
            if err:
                errors.append(err)
            fat_grams, err = parse_int(default_if_blank(gv(row_dict, 'fat_grams', 65), 65), 'grasas_g', minimum=0)
            if err:
                errors.append(err)
            fiber_grams, err = parse_int(default_if_blank(gv(row_dict, 'fiber_grams', 25), 25), 'fibra_g', minimum=0)
            if err:
                errors.append(err)

            protein_pct, err = parse_percent(gv(row_dict, 'protein_percentage', ''), 'porcentaje_proteinas')
            if err:
                errors.append(err)
            carbs_pct, err = parse_percent(gv(row_dict, 'carbs_percentage', ''), 'porcentaje_carbohidratos')
            if err:
                errors.append(err)
            fat_pct, err = parse_percent(gv(row_dict, 'fat_percentage', ''), 'porcentaje_grasas')
            if err:
                errors.append(err)

            meals_per_day, err = parse_int(default_if_blank(gv(row_dict, 'meals_per_day', 5), 5), 'comidas_por_dia', minimum=1, maximum=8)
            if err:
                errors.append(err)
            duration_weeks, err = parse_int(default_if_blank(gv(row_dict, 'duration_weeks', 4), 4), 'duracion_semanas', minimum=1)
            if err:
                errors.append(err)
            portion_multiplier, err = parse_float(default_if_blank(gv(row_dict, 'portion_multiplier', 1.0), 1.0), 'multiplicador_porcion', minimum=0.1)
            if err:
                errors.append(err)

            is_template, err = parse_bool(default_if_blank(gv(row_dict, 'is_template', 'false'), 'false'))
            if err:
                errors.append(f"es_plantilla: {err}")
            is_system, err = parse_bool(default_if_blank(gv(row_dict, 'is_system', 'false'), 'false'))
            if err:
                errors.append(f"es_sistema: {err}")
            is_active, err = parse_bool(default_if_blank(gv(row_dict, 'is_active', 'true'), 'true'))
            if err:
                errors.append(f"activo: {err}")

            if errors:
                return None, errors

            return {
                'description': str(gv(row_dict, 'description', '') or ''),
                'goal': goal,
                'diet_type': diet_type,
                'daily_calories': daily_calories,
                'protein_grams': protein_grams,
                'carbs_grams': carbs_grams,
                'fat_grams': fat_grams,
                'fiber_grams': fiber_grams,
                'protein_percentage': protein_pct,
                'carbs_percentage': carbs_pct,
                'fat_percentage': fat_pct,
                'meals_per_day': meals_per_day,
                'duration_weeks': duration_weeks,
                'portion_multiplier': portion_multiplier,
                'is_template': is_template,
                'is_system': is_system,
                'is_active': is_active,
            }, None

        def parse_meal_fields(row_dict):
            errors = []
            raw_meal_type = str(row_dict.get('tipo_comida', '') or 'lunch').strip() or 'lunch'
            meal_type = meal_type_input_map.get(normalize_text(raw_meal_type), raw_meal_type)
            if meal_type not in meal_type_values:
                errors.append(f"tipo_comida inválido: {raw_meal_type}")

            day_of_week, err = parse_int(row_dict.get('dia_semana', ''), 'dia_semana', minimum=1, maximum=7, allow_blank=True)
            if err:
                errors.append(err)
            order_index, err = parse_int(row_dict.get('orden_comida', ''), 'orden_comida', minimum=1)
            if err:
                errors.append(err)
            meal_time, err = parse_time_hhmm(row_dict.get('hora', ''))
            if err:
                errors.append(err)
            calories, err = parse_int(row_dict.get('calorias_comida', ''), 'calorias_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            protein, err = parse_float(row_dict.get('proteinas_comida', ''), 'proteinas_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            carbs, err = parse_float(row_dict.get('carbohidratos_comida', ''), 'carbohidratos_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)
            fat, err = parse_float(row_dict.get('grasas_comida', ''), 'grasas_comida', minimum=0, allow_blank=True)
            if err:
                errors.append(err)

            if errors:
                return None, errors
            return {
                'meal_type': meal_type,
                'day_of_week': day_of_week,
                'order_index': order_index,
                'time': meal_time,
                'calories': calories,
                'protein': Decimal(str(protein)) if protein is not None else None,
                'carbs': Decimal(str(carbs)) if carbs is not None else None,
                'fat': Decimal(str(fat)) if fat is not None else None,
                'name': str(row_dict.get('comida_nombre', '') or '').strip() or f'Comida {order_index}',
                'description': str(row_dict.get('descripcion_comida', '') or ''),
                'recetas_sugeridas': str(row_dict.get('recetas_sugeridas', '') or '').strip(),
            }, None

        def infer_meal_macros_from_suggested_recipes(recetas_sugeridas):
            recipe_names = [name.strip() for name in str(recetas_sugeridas or '').split(';') if name.strip()]
            if not recipe_names:
                return None
            recipes = list(Recipe.objects.filter(name__in=recipe_names))
            if not recipes:
                return None

            def avg_numeric(attr):
                values = [float(getattr(recipe, attr) or 0) for recipe in recipes]
                return (sum(values) / len(values)) if values else 0.0

            return {
                'calories': int(round(avg_numeric('calories'))),
                'protein': Decimal(str(round(avg_numeric('protein'), 2))),
                'carbs': Decimal(str(round(avg_numeric('carbs'), 2))),
                'fat': Decimal(str(round(avg_numeric('fat'), 2))),
            }

        def parse_option_fields(row_dict):
            def optv(data, *keys, default=''):
                for key in keys:
                    if key in data and data.get(key) not in (None, ''):
                        return data.get(key)
                return default

            errors = []
            display_order, err = parse_int(optv(row_dict, 'orden_visualizacion', 'display_order', default=0), 'orden_visualizacion', minimum=0)
            if err:
                errors.append(err)
            servings, err = parse_float(optv(row_dict, 'porciones', 'servings', default=1), 'porciones', minimum=0.01)
            if err:
                errors.append(err)
            custom_calories, err = parse_int(
                optv(row_dict, 'calorias_personalizadas', 'custom_calories', default=''),
                'calorias_personalizadas',
                minimum=0,
                allow_blank=True,
            )
            if err:
                errors.append(err)
            custom_protein, err = parse_decimal_or_none(
                optv(row_dict, 'proteinas_personalizadas', 'custom_protein', default=''),
                'proteinas_personalizadas',
            )
            if err:
                errors.append(err)
            custom_carbs, err = parse_decimal_or_none(
                optv(row_dict, 'carbohidratos_personalizados', 'custom_carbs', default=''),
                'carbohidratos_personalizados',
            )
            if err:
                errors.append(err)
            custom_fat, err = parse_decimal_or_none(
                optv(row_dict, 'grasas_personalizadas', 'custom_fat', default=''),
                'grasas_personalizadas',
            )
            if err:
                errors.append(err)

            if errors:
                return None, errors
            return {
                'display_order': display_order,
                'servings': Decimal(str(servings)),
                'custom_calories': custom_calories,
                'custom_protein': custom_protein,
                'custom_carbs': custom_carbs,
                'custom_fat': custom_fat,
            }, None

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo Excel no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)
        wb = openpyxl.load_workbook(file)
        # Support new multi-sheet format (sheet "Planes") and old single-sheet format
        ws = wb['Planes'] if 'Planes' in wb.sheetnames else wb.active
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
        updated, created, skipped, rejected = 0, 0, 0, 0
        meals_created, meals_updated, meals_skipped, meals_rejected = 0, 0, 0, 0
        options_created, options_updated, options_skipped, options_rejected = 0, 0, 0, 0
        assignments_created, assignments_updated, assignments_skipped, assignments_rejected = 0, 0, 0, 0
        errors = []

        def _get_plan_aliases_for_row(row_dict):
            # Support both old format ('nombre') and new format ('plan_nombre')
            if 'plan_nombre' in row_dict and row_dict['plan_nombre']:
                return str(row_dict['plan_nombre'] or '').strip()
            return str(gv(row_dict, 'name') or '').strip()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = dict(zip(headers, row))
            name = _get_plan_aliases_for_row(row_dict)
            if not name:
                skipped += 1
                continue
            fields, row_errors = parse_plan_fields(row_dict)
            if row_errors:
                rejected += 1
                errors.append({'sheet': 'Planes', 'row': row_idx, 'type': 'plan', 'errors': row_errors})
                continue

            plan = NutritionPlan.objects.filter(name=name).first()
            if plan:
                for k, v in fields.items():
                    setattr(plan, k, v)
                plan.save()
                updated += 1
            else:
                NutritionPlan.objects.create(name=name, **fields)
                created += 1

        # --- Process Comidas sheet if present ---
        if 'Comidas' in wb.sheetnames:
            ws_meals = wb['Comidas']
            meal_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws_meals[1]]
            for row_idx, meal_row in enumerate(ws_meals.iter_rows(min_row=2, values_only=True), start=2):
                mrow = dict(zip(meal_headers, meal_row))
                plan_name = str(mrow.get('plan_nombre', '') or '').strip()
                if not plan_name:
                    meals_skipped += 1
                    continue
                plan = NutritionPlan.objects.filter(name=plan_name).first()
                if not plan:
                    plan = NutritionPlan.objects.create(
                        name=plan_name,
                        description='',
                        goal='maintain',
                        diet_type='normal',
                        daily_calories=2000,
                        protein_grams=150,
                        carbs_grams=200,
                        fat_grams=65,
                        fiber_grams=25,
                        protein_percentage=30,
                        carbs_percentage=40,
                        fat_percentage=30,
                        meals_per_day=5,
                        duration_weeks=4,
                        portion_multiplier=1.0,
                        is_template=True,
                        is_system=False,
                        is_active=True,
                    )
                    created += 1
                parsed_meal, row_errors = parse_meal_fields(mrow)
                if row_errors:
                    meals_rejected += 1
                    errors.append({'sheet': 'Comidas', 'row': row_idx, 'type': 'comida', 'errors': row_errors})
                    continue

                inferred = infer_meal_macros_from_suggested_recipes(parsed_meal['recetas_sugeridas'])
                calories = parsed_meal['calories'] if parsed_meal['calories'] is not None else (inferred['calories'] if inferred else 0)
                protein = parsed_meal['protein'] if parsed_meal['protein'] is not None else (inferred['protein'] if inferred else Decimal('0'))
                carbs = parsed_meal['carbs'] if parsed_meal['carbs'] is not None else (inferred['carbs'] if inferred else Decimal('0'))
                fat = parsed_meal['fat'] if parsed_meal['fat'] is not None else (inferred['fat'] if inferred else Decimal('0'))

                meal = PlanMeal.objects.filter(
                    plan=plan,
                    day_of_week=parsed_meal['day_of_week'],
                    order_index=parsed_meal['order_index'],
                ).first()
                meal_fields = {
                    'name': parsed_meal['name'],
                    'meal_type': parsed_meal['meal_type'],
                    'time': parsed_meal['time'],
                    'calories': calories,
                    'protein': protein,
                    'carbs': carbs,
                    'fat': fat,
                    'description': parsed_meal['description'],
                }
                if meal:
                    for k, v in meal_fields.items():
                        setattr(meal, k, v)
                    meal.save()
                    meals_updated += 1
                else:
                    meal = PlanMeal.objects.create(
                        plan=plan,
                        day_of_week=parsed_meal['day_of_week'],
                        order_index=parsed_meal['order_index'],
                        **meal_fields
                    )
                    meals_created += 1
                recetas_str = parsed_meal['recetas_sugeridas']
                if recetas_str:
                    from .models import Recipe as RecipeModel
                    recipe_names = [r.strip() for r in recetas_str.split(';') if r.strip()]
                    recipes_qs = RecipeModel.objects.filter(name__in=recipe_names)
                    meal.suggested_recipes.set(recipes_qs)

        # --- Process Opciones_Receta_Comida sheet if present ---
        if 'Opciones_Receta_Comida' in wb.sheetnames:
            ws_options = wb['Opciones_Receta_Comida']
            option_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws_options[1]]
            option_recipes_by_meal = {}
            for row_idx, option_row in enumerate(ws_options.iter_rows(min_row=2, values_only=True), start=2):
                orow = dict(zip(option_headers, option_row))
                plan_name = str(orow.get('plan_nombre', '') or '').strip()
                recipe_name = str(orow.get('receta_nombre', '') or '').strip()
                image_url = str((orow.get('imagen_url') or orow.get('image_url') or orow.get('url_imagen') or '')).strip()
                if not plan_name or (not recipe_name and not image_url):
                    options_skipped += 1
                    continue
                plan = NutritionPlan.objects.filter(name=plan_name).first()
                if not plan:
                    plan = NutritionPlan.objects.create(
                        name=plan_name,
                        description='',
                        goal='maintain',
                        diet_type='normal',
                        daily_calories=2000,
                        protein_grams=150,
                        carbs_grams=200,
                        fat_grams=65,
                        fiber_grams=25,
                        protein_percentage=30,
                        carbs_percentage=40,
                        fat_percentage=30,
                        meals_per_day=5,
                        duration_weeks=4,
                        portion_multiplier=1.0,
                        is_template=True,
                        is_system=False,
                        is_active=True,
                    )
                    created += 1
                day_of_week, err = parse_int(orow.get('dia_semana', ''), 'dia_semana', minimum=1, maximum=7, allow_blank=True)
                order_index, err2 = parse_int(orow.get('orden_comida', ''), 'orden_comida', minimum=1)
                local_errors = []
                if err:
                    local_errors.append(err)
                if err2:
                    local_errors.append(err2)
                parsed_option, option_errors = parse_option_fields(orow)
                if option_errors:
                    local_errors.extend(option_errors)
                meal = None
                if not local_errors:
                    meal = PlanMeal.objects.filter(plan=plan, day_of_week=day_of_week, order_index=order_index).first()
                    if not meal:
                        local_errors.append(f"No existe comida para clave ({plan_name}, día={day_of_week}, orden={order_index})")
                recipe = None
                if not local_errors:
                    if recipe_name:
                        recipe = Recipe.objects.filter(name=recipe_name).first()
                    if not recipe and image_url:
                        recipe = Recipe.objects.filter(image_url=image_url).first()
                    if not recipe:
                        if recipe_name and image_url:
                            local_errors.append(f"Receta no existe: nombre={recipe_name} ni imagen_url={image_url}")
                        elif recipe_name:
                            local_errors.append(f"Receta no existe: {recipe_name}")
                        else:
                            local_errors.append(f"Receta no existe para imagen_url: {image_url}")
                if local_errors:
                    options_rejected += 1
                    errors.append({'sheet': 'Opciones_Receta_Comida', 'row': row_idx, 'type': 'opcion_receta', 'errors': local_errors})
                    continue

                option = PlanMealRecipe.objects.filter(meal=meal, recipe=recipe).first()
                if option:
                    for k, v in parsed_option.items():
                        setattr(option, k, v)
                    option.save()
                    options_updated += 1
                else:
                    PlanMealRecipe.objects.create(meal=meal, recipe=recipe, **parsed_option)
                    options_created += 1

                option_recipes_by_meal.setdefault(meal.id, set()).add(recipe.id)

            # Sync suggested_recipes using imported options
            if option_recipes_by_meal:
                for meal_id, recipe_ids in option_recipes_by_meal.items():
                    meal = PlanMeal.objects.filter(id=meal_id).first()
                    if meal:
                        meal.suggested_recipes.set(Recipe.objects.filter(id__in=recipe_ids))

        # --- Process Asignaciones_Usuarios sheet if present ---
        if 'Asignaciones_Usuarios' in wb.sheetnames:
            ws_assign = wb['Asignaciones_Usuarios']
            assign_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws_assign[1]]
            for row_idx, assign_row in enumerate(ws_assign.iter_rows(min_row=2, values_only=True), start=2):
                arow = dict(zip(assign_headers, assign_row))
                plan_name = str(arow.get('plan_nombre', '') or '').strip()
                email = str(arow.get('usuario_email', '') or '').strip().lower()
                if not plan_name or not email:
                    assignments_skipped += 1
                    continue
                plan = NutritionPlan.objects.filter(name=plan_name).first()
                if not plan:
                    assignments_rejected += 1
                    errors.append({'sheet': 'Asignaciones_Usuarios', 'row': row_idx, 'type': 'asignacion_usuario', 'errors': [f"Plan no existe: {plan_name}"]})
                    continue
                user = User.objects.filter(email__iexact=email).first()
                if not user:
                    assignments_rejected += 1
                    errors.append({'sheet': 'Asignaciones_Usuarios', 'row': row_idx, 'type': 'asignacion_usuario', 'errors': [f"Usuario no existe: {email}"]})
                    continue
                is_active, err = parse_bool(arow.get('asignacion_activa', 'true'))
                if err:
                    assignments_rejected += 1
                    errors.append({'sheet': 'Asignaciones_Usuarios', 'row': row_idx, 'type': 'asignacion_usuario', 'errors': [err]})
                    continue
                assignment = NutritionPlanAssignment.objects.filter(plan=plan, user=user).first()
                if assignment:
                    assignment.is_active = is_active
                    assignment.save(update_fields=['is_active'])
                    assignments_updated += 1
                else:
                    NutritionPlanAssignment.objects.create(plan=plan, user=user, is_active=is_active)
                    assignments_created += 1

        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'rejected': rejected,
            'meals_created': meals_created,
            'meals_updated': meals_updated,
            'meals_skipped': meals_skipped,
            'meals_rejected': meals_rejected,
            'options_created': options_created,
            'options_updated': options_updated,
            'options_skipped': options_skipped,
            'options_rejected': options_rejected,
            'assignments_created': assignments_created,
            'assignments_updated': assignments_updated,
            'assignments_skipped': assignments_skipped,
            'assignments_rejected': assignments_rejected,
            'errors': errors,
            'message': (
                f"Importación Excel completada. Planes: {created} creados, {updated} actualizados, {skipped} omitidos, {rejected} rechazados. "
                f"Comidas: {meals_created} creadas, {meals_updated} actualizadas, {meals_skipped} omitidas, {meals_rejected} rechazadas. "
                f"Opciones: {options_created} creadas, {options_updated} actualizadas, {options_skipped} omitidas, {options_rejected} rechazadas. "
                f"Asignaciones: {assignments_created} creadas, {assignments_updated} actualizadas, {assignments_skipped} omitidas, {assignments_rejected} rechazadas."
            ),
        }, status=status.HTTP_200_OK)


class AdminPlanMealViewSet(viewsets.ModelViewSet):
    queryset = PlanMeal.objects.all()
    serializer_class = AdminPlanMealSerializer
    permission_classes = [IsAdminUser]


class AdminFoodViewSet(viewsets.ModelViewSet):
    queryset = Food.objects.all()
    serializer_class = AdminFoodSerializer
    permission_classes = [IsAdminUser]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    ALLERGEN_KEYWORDS = {
        'gluten': ['gluten', 'trigo', 'wheat', 'harina', 'pan', 'pasta', 'cebada', 'barley', 'centeno', 'rye', 'avena', 'oats', 'malta', 'malt', 'semola', 'semolina', 'couscous'],
        'dairy': ['leche', 'milk', 'queso', 'cheese', 'yogur', 'yogurt', 'mantequilla', 'butter', 'nata', 'cream', 'crema', 'lactosa', 'lactose', 'whey', 'caseina', 'casein'],
        'eggs': ['huevo', 'huevos', 'egg', 'eggs', 'mayonesa', 'mayo'],
        'nuts': ['almendra', 'almendras', 'almond', 'nuez', 'nueces', 'walnut', 'avellana', 'hazelnut', 'pistacho', 'pistachio', 'cacahuete', 'cacahuate', 'peanut', 'mani', 'anacardo', 'cashew', 'macadamia', 'pecan'],
        'soy': ['soja', 'soy', 'soya', 'tofu', 'tempeh', 'edamame', 'miso'],
        'fish': ['pescado', 'fish', 'atun', 'tuna', 'salmon', 'bacalao', 'merluza', 'trucha', 'sardina'],
        'shellfish': ['marisco', 'shellfish', 'gamba', 'gambas', 'langostino', 'camaron', 'cangrejo', 'crab', 'langosta', 'lobster', 'mejillon', 'almeja', 'ostra'],
        'sesame': ['sesamo', 'sesame', 'ajonjoli', 'tahini'],
    }

    ALLERGEN_ALIASES = {
        'gluten': 'gluten',
        'dairy': 'dairy',
        'lactose': 'dairy',
        'lactosa': 'dairy',
        'milk': 'dairy',
        'eggs': 'eggs',
        'egg': 'eggs',
        'nuts': 'nuts',
        'nut': 'nuts',
        'tree-nuts': 'nuts',
        'soy': 'soy',
        'soya': 'soy',
        'soja': 'soy',
        'fish': 'fish',
        'shellfish': 'shellfish',
        'sesame': 'sesame',
        'sesamo': 'sesame',
    }

    ALLERGEN_EXPORT_ORDER = ['gluten', 'dairy', 'eggs', 'nuts', 'soy', 'fish', 'shellfish', 'sesame']

    def _normalize_text(self, value):
        text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii').lower()
        return ' '.join(text.replace('-', ' ').replace('_', ' ').split())

    def _normalize_allergen(self, value):
        key = self._normalize_text(value).replace(' ', '-')
        if not key:
            return ''
        return self.ALLERGEN_ALIASES.get(key, key)

    def _parse_allergens_cell(self, value):
        if value is None:
            return []

        if isinstance(value, list):
            parts = value
        else:
            text = str(value).strip()
            if not text:
                return []
            cleaned = text.strip('[]')
            parts = re.split(r'[|,;/]+', cleaned)

        normalized = []
        for item in parts:
            allergen = self._normalize_allergen(item)
            if allergen and allergen not in normalized:
                normalized.append(allergen)

        return normalized

    def _infer_allergens_from_text(self, *values):
        source = self._normalize_text(' '.join(str(v or '') for v in values))
        if not source:
            return []

        inferred = []
        for allergen, keywords in self.ALLERGEN_KEYWORDS.items():
            if any(keyword in source for keyword in keywords):
                inferred.append(allergen)
        return inferred

    def _resolve_allergens_for_food(self, *, name, brand, category, provided):
        parsed = self._parse_allergens_cell(provided)
        inferred = self._infer_allergens_from_text(name, brand, category)

        merged = []
        for allergen in self.ALLERGEN_EXPORT_ORDER:
            if allergen in parsed or allergen in inferred:
                merged.append(allergen)

        for allergen in parsed + inferred:
            if allergen not in merged:
                merged.append(allergen)

        return merged

    def _allergens_to_export_value(self, allergens):
        if not isinstance(allergens, list):
            return ''
        ordered = [allergen for allergen in self.ALLERGEN_EXPORT_ORDER if allergen in allergens]
        extras = [allergen for allergen in allergens if allergen not in ordered]
        return ','.join(ordered + extras)

    @action(detail=False, methods=['get'], url_path='list-for-recipes')
    def list_for_recipes(self, request):
        """Devuelve lista simplificada de alimentos para seleccionar en recetas"""
        from .serializers import FoodMinimalSerializer
        foods = self.get_queryset().order_by('name')
        serializer = FoodMinimalSerializer(foods, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """Exporta todos los alimentos en formato CSV"""
        import csv
        from django.http import HttpResponse
        foods = self.get_queryset()
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="alimentos_exportacion.csv"'
        fieldnames = [
            'nombre', 'marca', 'calorías', 'proteínas', 'carbohidratos', 'grasas',
            'fibra', 'azúcar', 'sodio', 'tamaño_porcion', 'unidad_porcion', 'categoría', 'equivalencia', 'tienda', 'alergenos', 'verificado'
        ]
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        for food in foods:
            writer.writerow({
                'nombre': food.name,
                'marca': food.brand or '',
                'calorías': food.calories or 0,
                'proteínas': food.protein or 0,
                'carbohidratos': food.carbs or 0,
                'grasas': food.fat or 0,
                'fibra': food.fiber or 0,
                'azúcar': food.sugar or 0,
                'sodio': food.sodium or 0,
                'tamaño_porcion': food.serving_size or 0,
                'unidad_porcion': food.serving_unit or '',
                'categoría': food.category or '',
                'equivalencia': food.equivalence_category or '',
                'tienda': food.store or '',
                'alergenos': self._allergens_to_export_value(food.allergens),
                'verificado': 'Sí' if food.is_verified else 'No',
            })
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporta todos los alimentos en formato Excel (XLSX)"""
        import io
        import xlsxwriter
        from django.http import HttpResponse
        foods = self.get_queryset()
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'strings_to_urls': False})
        worksheet = workbook.add_worksheet('Alimentos')
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        headers = [
            'nombre', 'marca', 'calorías', 'proteínas', 'carbohidratos', 'grasas',
            'fibra', 'azúcar', 'sodio', 'tamaño_porcion', 'unidad_porcion', 'categoría', 'equivalencia', 'tienda', 'alergenos', 'verificado'
        ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        for row_idx, food in enumerate(foods, start=1):
            worksheet.write(row_idx, 0, food.name)
            worksheet.write(row_idx, 1, food.brand or '')
            worksheet.write(row_idx, 2, food.calories or 0)
            worksheet.write(row_idx, 3, float(food.protein or 0))
            worksheet.write(row_idx, 4, float(food.carbs or 0))
            worksheet.write(row_idx, 5, float(food.fat or 0))
            worksheet.write(row_idx, 6, float(food.fiber or 0))
            worksheet.write(row_idx, 7, float(food.sugar or 0))
            worksheet.write(row_idx, 8, float(food.sodium or 0))
            worksheet.write(row_idx, 9, food.serving_size or 0)
            worksheet.write(row_idx, 10, food.serving_unit or '')
            worksheet.write(row_idx, 11, food.category or '')
            worksheet.write(row_idx, 12, food.equivalence_category or '')
            worksheet.write(row_idx, 13, food.store or '')
            worksheet.write(row_idx, 14, self._allergens_to_export_value(food.allergens))
            worksheet.write(row_idx, 15, 'Sí' if food.is_verified else 'No')
        worksheet.set_column('A:A', 38)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('N:N', 26)
        workbook.close()
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="alimentos_exportacion.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """Descarga una plantilla Excel vacia para importacion de alimentos."""
        import io
        import xlsxwriter
        from django.http import HttpResponse

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True, 'strings_to_urls': False})
        worksheet = workbook.add_worksheet('Plantilla Alimentos')

        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        headers = [
            'nombre', 'marca', 'calorías', 'proteínas', 'carbohidratos', 'grasas',
            'fibra', 'azúcar', 'sodio', 'tamaño_porcion', 'unidad_porcion', 'categoría', 'equivalencia', 'tienda', 'alergenos', 'verificado'
        ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Fila de ejemplo para guiar el formato esperado.
        sample = [
            'Arroz cocido', 'Genérico', 130, 2.7, 28.2, 0.3,
            0.4, 0.1, 1.0, 100, 'g', 'Cereales', 'arroz_cereales', 'Mercadona', 'gluten', 'No'
        ]
        for col, value in enumerate(sample):
            worksheet.write(1, col, value)

        worksheet.set_column('A:A', 28)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:N', 16)

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_alimentos.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import-csv')
    def import_csv(self, request):
        """Importa alimentos desde un archivo CSV. Acepta columnas en español o inglés. Añade o modifica, nunca elimina."""
        import csv
        from django.core.files.uploadedfile import UploadedFile

        # Aliases: columna exportada → campo interno
        food_aliases = {
            'nombre': 'name', 'name': 'name',
            'marca': 'brand', 'brand': 'brand',
            'calorías': 'calories', 'calorias': 'calories', 'calories': 'calories',
            'proteínas': 'protein', 'proteinas': 'protein', 'protein': 'protein',
            'carbohidratos': 'carbs', 'carbs': 'carbs',
            'grasas': 'fat', 'fat': 'fat',
            'fibra': 'fiber', 'fiber': 'fiber',
            'azúcar': 'sugar', 'azucar': 'sugar', 'sugar': 'sugar',
            'sodio': 'sodium', 'sodium': 'sodium',
            'tamaño_porcion': 'serving_size', 'tamano_porcion': 'serving_size', 'serving_size': 'serving_size',
            'unidad_porcion': 'serving_unit', 'serving_unit': 'serving_unit',
            'categoría': 'category', 'categoria': 'category', 'category': 'category',
            'equivalencia': 'equivalence_category', 'grupo_equivalencia': 'equivalence_category',
            'equivalence': 'equivalence_category', 'equivalence_category': 'equivalence_category',
            'tienda': 'store', 'store': 'store',
            'alergenos': 'allergens', 'alérgenos': 'allergens', 'allergens': 'allergens',
            'verificado': 'is_verified', 'is_verified': 'is_verified',
        }

        def gv(row, canonical, default=''):
            if canonical in row:
                return row[canonical]
            for k, v in food_aliases.items():
                if v == canonical and k in row:
                    return row[k]
            return default

        def to_bool(val):
            return str(val).lower() in ('true', 'sí', 'si', 's', '1', 'yes')

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo CSV no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)
        decoded = file.read().decode('utf-8')
        reader = csv.DictReader(decoded.splitlines())
        updated, created, skipped = 0, 0, 0
        for row in reader:
            name = gv(row, 'name', '').strip()
            if not name:
                skipped += 1
                continue
            food = Food.objects.filter(name=name).first()
            fields = {
                'brand': gv(row, 'brand', ''),
                'calories': float(gv(row, 'calories', 0) or 0),
                'protein': float(gv(row, 'protein', 0) or 0),
                'carbs': float(gv(row, 'carbs', 0) or 0),
                'fat': float(gv(row, 'fat', 0) or 0),
                'fiber': float(gv(row, 'fiber', 0) or 0),
                'sugar': float(gv(row, 'sugar', 0) or 0),
                'sodium': float(gv(row, 'sodium', 0) or 0),
                'serving_size': float(gv(row, 'serving_size', 0) or 0),
                'serving_unit': gv(row, 'serving_unit', ''),
                'category': gv(row, 'category', ''),
                'equivalence_category': gv(row, 'equivalence_category', ''),
                'store': gv(row, 'store', ''),
                'allergens': self._resolve_allergens_for_food(
                    name=name,
                    brand=gv(row, 'brand', ''),
                    category=gv(row, 'category', ''),
                    provided=gv(row, 'allergens', ''),
                ),
                'is_verified': to_bool(gv(row, 'is_verified', 'false')),
            }
            if food:
                for k, v in fields.items():
                    setattr(food, k, v)
                food.save()
                updated += 1
            else:
                Food.objects.create(name=name, **fields)
                created += 1
        return Response({
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'message': f"Se subió el archivo correctamente. {created} alimentos añadidos, {updated} modificados. Los alimentos no presentes en el archivo no se eliminaron."
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """Importa alimentos desde un archivo Excel. Acepta columnas en español o inglés. Añade o modifica, nunca elimina."""
        import openpyxl
        from openpyxl.utils.exceptions import InvalidFileException
        from django.core.files.uploadedfile import UploadedFile

        food_aliases = {
            'nombre': 'name', 'name': 'name',
            'marca': 'brand', 'brand': 'brand',
            'calorías': 'calories', 'calorias': 'calories', 'calories': 'calories',
            'proteínas': 'protein', 'proteinas': 'protein', 'protein': 'protein',
            'carbohidratos': 'carbs', 'carbs': 'carbs',
            'grasas': 'fat', 'fat': 'fat',
            'fibra': 'fiber', 'fiber': 'fiber',
            'azúcar': 'sugar', 'azucar': 'sugar', 'sugar': 'sugar',
            'sodio': 'sodium', 'sodium': 'sodium',
            'tamaño_porcion': 'serving_size', 'tamano_porcion': 'serving_size', 'serving_size': 'serving_size',
            'unidad_porcion': 'serving_unit', 'serving_unit': 'serving_unit',
            'categoría': 'category', 'categoria': 'category', 'category': 'category',
            'equivalencia': 'equivalence_category', 'grupo_equivalencia': 'equivalence_category',
            'equivalence': 'equivalence_category', 'equivalence_category': 'equivalence_category',
            'tienda': 'store', 'store': 'store',
            'alergenos': 'allergens', 'alérgenos': 'allergens', 'allergens': 'allergens',
            'verificado': 'is_verified', 'is_verified': 'is_verified',
        }

        def gv(row_dict, canonical, default=''):
            if canonical in row_dict:
                return row_dict[canonical]
            for k, v in food_aliases.items():
                if v == canonical and k in row_dict:
                    return row_dict[k]
            return default

        def to_bool(val):
            return str(val).lower() in ('true', 'sí', 'si', 's', '1', 'yes')

        recovered_date_values = []

        def to_float(value, field_name, row_name=''):
            if value is None:
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)

            # openpyxl puede devolver datetime/date si la celda tiene formato de fecha.
            # En algunos editores/locales, decimales con punto (8.5, 7.1) se convierten
            # automáticamente a fecha (8/5, 7/1). Recuperamos esos casos con heurística dd.mm.
            if hasattr(value, 'isoformat') and not isinstance(value, str):
                try:
                    day = int(getattr(value, 'day'))
                    month = int(getattr(value, 'month'))
                    recovered = float(f"{day}.{month}")
                    label = row_name or 'fila sin nombre'
                    recovered_date_values.append(
                        f"{label}: campo '{field_name}' convertido de fecha {value} a {recovered}"
                    )
                    return recovered
                except Exception:
                    raise ValueError(
                        f"campo '{field_name}' esperaba un número y recibió una fecha: {value}"
                    )

            normalized = str(value).strip().replace(',', '.')
            if normalized == '':
                return 0.0
            try:
                return float(normalized)
            except (TypeError, ValueError):
                raise ValueError(
                    f"campo '{field_name}' esperaba un número y recibió: '{value}'"
                )

        def can_parse_float(value):
            if value is None:
                return True
            if isinstance(value, (int, float)):
                return True
            if hasattr(value, 'isoformat') and not isinstance(value, str):
                return False
            normalized = str(value).strip().replace(',', '.')
            if normalized == '':
                return True
            try:
                float(normalized)
                return True
            except (TypeError, ValueError):
                return False

        file = request.FILES.get('file')
        if not file or not isinstance(file, UploadedFile):
            return Response({'error': 'Archivo Excel no proporcionado'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
        except InvalidFileException:
            return Response({'error': 'El archivo no es un Excel válido (.xlsx).'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response({'error': f'No se pudo leer el archivo Excel: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]

        # Normalizar cabeceras para evitar desajustes por mayúsculas/espacios
        # y para detectar columnas duplicadas que pisan valores al hacer zip.
        normalized_headers = [h.lower() for h in raw_headers]
        canonical_headers = []
        canonical_seen = set()
        duplicate_headers = set()
        canonical_indexes = {}
        for header in normalized_headers:
            canonical = food_aliases.get(header, header)
            canonical_headers.append(canonical)
            if canonical:
                if canonical in canonical_seen:
                    duplicate_headers.add(canonical)
                else:
                    canonical_seen.add(canonical)

        for idx, canonical in enumerate(canonical_headers):
            if not canonical:
                continue
            canonical_indexes.setdefault(canonical, []).append(idx)

        updated, created, skipped = 0, 0, 0
        errors = []
        skipped_details = []
        row_number = 1  # 1-indexed desde la primera fila de datos (fila 2 de Excel)
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_number += 1
            row_dict = {}
            for idx, canonical in enumerate(canonical_headers):
                if not canonical:
                    continue
                if canonical in row_dict:
                    # Si hay columnas repetidas, usamos la primera ocurrencia.
                    continue
                row_dict[canonical] = row[idx] if idx < len(row) else None

            def get_best_value(canonical, default='', numeric=False):
                if canonical not in canonical_indexes:
                    return default

                candidates = []
                for idx in canonical_indexes[canonical]:
                    value = row[idx] if idx < len(row) else None
                    candidates.append(value)

                if not numeric:
                    for value in candidates:
                        if value is not None and str(value).strip() != '':
                            return value
                    return default

                # Para numéricos, si hay columnas duplicadas, elegir la primera
                # que realmente sea parseable como número.
                for value in candidates:
                    if can_parse_float(value):
                        return value

                # Si ninguna es válida, devolvemos la primera no vacía para que
                # to_float genere un error claro al usuario.
                for value in candidates:
                    if value is not None and str(value).strip() != '':
                        return value
                return default

            name = str(gv(row_dict, 'name') or '').strip()
            if not name:
                # Si toda la fila está vacía, ignorarla silenciosamente
                if all(v is None or str(v).strip() == '' for v in row):
                    continue
                skipped += 1
                skipped_details.append({'row': row_number, 'name': None, 'reason': 'Nombre vacío'})
                continue

            try:
                food = Food.objects.filter(name=name).first()
                fields = {
                    'brand': get_best_value('brand', '') or '',
                    'calories': to_float(get_best_value('calories', 0, numeric=True), 'calories', name),
                    'protein': to_float(get_best_value('protein', 0, numeric=True), 'protein', name),
                    'carbs': to_float(get_best_value('carbs', 0, numeric=True), 'carbs', name),
                    'fat': to_float(get_best_value('fat', 0, numeric=True), 'fat', name),
                    'fiber': to_float(get_best_value('fiber', 0, numeric=True), 'fiber', name),
                    'sugar': to_float(get_best_value('sugar', 0, numeric=True), 'sugar', name),
                    'sodium': to_float(get_best_value('sodium', 0, numeric=True), 'sodium', name),
                    'serving_size': to_float(get_best_value('serving_size', 0, numeric=True), 'serving_size', name),
                    'serving_unit': get_best_value('serving_unit', '') or '',
                    'category': get_best_value('category', '') or '',
                    'equivalence_category': get_best_value('equivalence_category', '') or '',
                    'store': get_best_value('store', '') or '',
                    'allergens': self._resolve_allergens_for_food(
                        name=name,
                        brand=get_best_value('brand', '') or '',
                        category=get_best_value('category', '') or '',
                        provided=get_best_value('allergens', ''),
                    ),
                    'is_verified': to_bool(get_best_value('is_verified', 'false')),
                }
                if food:
                    for k, v in fields.items():
                        setattr(food, k, v)
                    food.save()
                    updated += 1
                else:
                    Food.objects.create(name=name, **fields)
                    created += 1
            except Exception as row_error:
                skipped += 1
                error_msg = str(row_error)
                errors.append({'name': name, 'error': error_msg})
                skipped_details.append({'row': row_number, 'name': name, 'reason': error_msg})

        response_payload = {
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors,
            'skipped_details': skipped_details,
            'message': f"Se subió el archivo correctamente. {created} alimentos añadidos, {updated} modificados. Los alimentos no presentes en el archivo no se eliminaron."
        }

        if duplicate_headers:
            response_payload['warnings'] = [
                'Cabeceras repetidas detectadas en Excel: '
                + ', '.join(sorted(duplicate_headers))
                + '. Se usó la primera columna de cada campo.'
            ]

        if recovered_date_values:
            existing_warnings = response_payload.get('warnings', [])
            existing_warnings.append(
                f"Se recuperaron {len(recovered_date_values)} valores numéricos que venían como fecha (p. ej. 8.5 -> 8/5)."
            )
            existing_warnings.extend(recovered_date_values[:30])
            response_payload['warnings'] = existing_warnings

        return Response(response_payload, status=status.HTTP_200_OK)


class AdminPlanMealRecipeViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar recetas sugeridas con cantidades personalizadas"""
    queryset = PlanMealRecipe.objects.all().select_related('meal', 'recipe')
    serializer_class = PlanMealRecipeSerializer
    permission_classes = [IsAdminUser]
    
    def get_queryset(self):
        """Filtrar por meal_id si se proporciona"""
        queryset = super().get_queryset()
        meal_id = self.request.query_params.get('meal_id')
        if meal_id:
            queryset = queryset.filter(meal_id=meal_id)
        return queryset


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_default_plans(request):
    """
    Lista de planes de nutrición predeterminados/plantillas
    """
    # Planes que son plantillas del sistema (sin usuario asignado o marcados como template)
    plans = NutritionPlan.objects.filter(
        user__isnull=True
    ).prefetch_related('meals__suggested_recipes') | NutritionPlan.objects.filter(
        is_template=True
    ).prefetch_related('meals__suggested_recipes')
    
    plans = plans.distinct()
    
    # Paginación
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))
    
    total = plans.count()
    start = (page - 1) * page_size
    end = start + page_size
    
    plans_list = plans[start:end]
    
    results = []
    for plan in plans_list:
        # Calcular porcentajes de macros
        protein_pct = 30  # valor por defecto
        carbs_pct = 40
        fat_pct = 30
        
        if plan.daily_calories and plan.protein_grams:
            protein_pct = round((plan.protein_grams * 4) / plan.daily_calories * 100, 1)
        if plan.daily_calories and plan.carbs_grams:
            carbs_pct = round((plan.carbs_grams * 4) / plan.daily_calories * 100, 1)
        if plan.daily_calories and plan.fat_grams:
            fat_pct = round((plan.fat_grams * 9) / plan.daily_calories * 100, 1)
        
        results.append({
            'id': str(plan.id),
            'name': plan.name,
            'description': plan.description,
            'goal': plan.goal,
            'diet_type': plan.diet_type,
            'daily_calories': plan.daily_calories,
            'protein_grams': float(plan.protein_grams) if plan.protein_grams else 0,
            'carbs_grams': float(plan.carbs_grams) if plan.carbs_grams else 0,
            'fat_grams': float(plan.fat_grams) if plan.fat_grams else 0,
            'protein_percentage': protein_pct,
            'carbs_percentage': carbs_pct,
            'fat_percentage': fat_pct,
            'meals_per_day': plan.meals_per_day,
            'duration_weeks': plan.duration_weeks,
            'is_active': plan.is_active,
            'is_template': plan.is_template,
            'is_system': plan.is_system,
            'is_default': getattr(plan, 'is_default', False),
            'meals_count': plan.meals.count(),
            'created_at': plan.created_at.isoformat() if plan.created_at else None,
            'updated_at': plan.updated_at.isoformat() if plan.updated_at else None,
        })
    
    return Response({
        'results': results,
        'count': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size if total > 0 else 0,
    })


@api_view(['POST'])
@perm_classes([IsAdminUser])
def admin_change_user_plan(request):
    """
    Asigna un plan base (template/sistema) a un usuario concreto,
    creando una copia personalizada según su perfil.
    """
    user_id = request.data.get('user_id')
    default_plan_id = request.data.get('default_plan_id')

    if not user_id or not default_plan_id:
        return Response(
            {'error': 'user_id y default_plan_id son requeridos'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        target_user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    try:
        default_plan = NutritionPlan.objects.prefetch_related('meals__suggested_recipes').get(pk=default_plan_id)
    except NutritionPlan.DoesNotExist:
        return Response({'error': 'Plan base no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if not default_plan.is_active:
        return Response({'error': 'El plan base está inactivo'}, status=status.HTTP_400_BAD_REQUEST)

    # Solo permitir plantillas/sistema como origen para evitar asignar planes de otros usuarios.
    if default_plan.user_id and not default_plan.is_template and not default_plan.is_system:
        return Response(
            {'error': 'Solo se pueden asignar planes plantilla o del sistema'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        service = PersonalizedNutritionService(target_user)
        new_plan = service.assign_plan_from_default(
            default_plan,
            changed_by=request.user,
            reason='admin_manual_change',
            notes=f'Plan asignado por admin ({request.user.email}) desde plantilla {default_plan.name}',
        )

        if not new_plan:
            return Response({'error': 'No se pudo asignar el plan al usuario'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = AdminNutritionPlanSerializer(new_plan)
        return Response(
            {
                'message': 'Plan asignado correctamente',
                'plan': serializer.data,
                'user_id': target_user.id,
                'default_plan_id': str(default_plan.id),
            },
            status=status.HTTP_200_OK,
        )
    except Exception as exc:
        return Response({'error': f'Error asignando plan: {str(exc)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@perm_classes([IsAdminUser])
def admin_bulk_change_plans(request):
    """
    Asignación masiva de un plan base a múltiples usuarios.
    Soporta:
      - user_ids: [1,2,3]
      - change_all: true + filter { role?, main_goal? }
    """
    default_plan_id = request.data.get('default_plan_id')
    if not default_plan_id:
        return Response({'error': 'default_plan_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        default_plan = NutritionPlan.objects.prefetch_related('meals__suggested_recipes').get(pk=default_plan_id)
    except NutritionPlan.DoesNotExist:
        return Response({'error': 'Plan base no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if not default_plan.is_active:
        return Response({'error': 'El plan base está inactivo'}, status=status.HTTP_400_BAD_REQUEST)

    if default_plan.user_id and not default_plan.is_template and not default_plan.is_system:
        return Response(
            {'error': 'Solo se pueden asignar planes plantilla o del sistema'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    change_all = bool(request.data.get('change_all', False))
    raw_user_ids = request.data.get('user_ids', [])
    filters_payload = request.data.get('filter', {}) or {}

    users_qs = User.objects.all()
    if change_all:
        role_value = filters_payload.get('role')
        goal_value = filters_payload.get('main_goal')
        if role_value:
            users_qs = users_qs.filter(role=role_value)
        if goal_value:
            users_qs = users_qs.filter(main_goal=goal_value)
    else:
        if not isinstance(raw_user_ids, list) or not raw_user_ids:
            return Response(
                {'error': 'Debes enviar user_ids o activar change_all'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        users_qs = users_qs.filter(id__in=raw_user_ids)

    users = list(users_qs)
    if not users:
        return Response(
            {
                'message': 'No hay usuarios para procesar con los filtros actuales',
                'success_count': 0,
                'error_count': 0,
                'errors': [],
            },
            status=status.HTTP_200_OK,
        )

    success_count = 0
    errors = []

    for target_user in users:
        try:
            service = PersonalizedNutritionService(target_user)
            assigned = service.assign_plan_from_default(
                default_plan,
                changed_by=request.user,
                reason='admin_bulk_change',
                notes=f'Asignación masiva por admin ({request.user.email}) desde plantilla {default_plan.name}',
            )
            if assigned:
                success_count += 1
            else:
                errors.append({'user_id': target_user.id, 'error': 'No se pudo asignar el plan'})
        except Exception as exc:
            errors.append({'user_id': target_user.id, 'error': str(exc)})

    return Response(
        {
            'message': f'Plan cambiado para {success_count} usuarios',
            'success_count': success_count,
            'error_count': len(errors),
            'errors': errors[:50],
        },
        status=status.HTTP_200_OK,
    )


@api_view(['GET', 'PUT', 'PATCH'])
@perm_classes([IsAdminUser])
def admin_user_plan(request, user_id: int):
    """
    Resumen del plan activo del usuario y consumo reciente de macros.
    Query params opcionales:
      - days: ventana en días para sumar macros (por defecto 7)
      - end_date: fecha fin YYYY-MM-DD (por defecto hoy)
    """
    user = get_object_or_404(User, pk=user_id)

    plan = NutritionPlan.objects.filter(user=user).prefetch_related('meals__suggested_recipes').order_by('-is_active', '-created_at').first()

    if request.method in ['PUT', 'PATCH']:
        if not plan:
            return Response({'error': 'El usuario no tiene un plan nutricional activo'}, status=status.HTTP_404_NOT_FOUND)

        try:
            new_daily_calories = int(float(request.data.get('daily_calories')))
        except (TypeError, ValueError):
            return Response({'error': 'daily_calories es requerido y debe ser numérico'}, status=status.HTTP_400_BAD_REQUEST)

        if new_daily_calories <= 0:
            return Response({'error': 'daily_calories debe ser mayor que 0'}, status=status.HTTP_400_BAD_REQUEST)

        old_daily_calories = int(plan.daily_calories or 0)
        service = PersonalizedNutritionService(user)
        adjusted = service.adjust_plan_calories(
            plan,
            new_daily_calories - old_daily_calories,
            reason='admin_manual_update',
            notes=f'Ajuste manual desde endpoint usuario-plan: {old_daily_calories} -> {new_daily_calories} kcal',
        )

        protein_grams = request.data.get('protein_grams', None)
        carbs_grams = request.data.get('carbs_grams', None)
        fat_grams = request.data.get('fat_grams', None)
        update_fields = []
        for field_name, raw_value in [
            ('protein_grams', protein_grams),
            ('carbs_grams', carbs_grams),
            ('fat_grams', fat_grams),
        ]:
            if raw_value is None:
                continue
            try:
                setattr(adjusted, field_name, int(float(raw_value)))
                update_fields.append(field_name)
            except (TypeError, ValueError):
                return Response({'error': f'{field_name} debe ser numérico'}, status=status.HTTP_400_BAD_REQUEST)

        if update_fields:
            adjusted.save(update_fields=update_fields)

        plan = NutritionPlan.objects.filter(pk=adjusted.pk).prefetch_related('meals__suggested_recipes').first()
    plan_data = AdminNutritionPlanSerializer(plan).data if plan else None

    days = max(int(request.query_params.get('days', 7)), 1)
    end_date_param = request.query_params.get('end_date')
    if end_date_param:
        try:
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
        except ValueError:
            end_date = timezone.localdate()
    else:
        end_date = timezone.localdate()
    start_date = end_date - timedelta(days=days - 1)

    meal_logs = MealLog.objects.filter(user=user, date__range=(start_date, end_date))

    aggregates = meal_logs.aggregate(
        calories=Sum('calories'),
        protein=Sum('protein'),
        carbs=Sum('carbs'),
        fat=Sum('fat'),
    )

    per_day = list(
        meal_logs.values('date').annotate(
            calories=Sum('calories'),
            protein=Sum('protein'),
            carbs=Sum('carbs'),
            fat=Sum('fat'),
        ).order_by('-date')
    )

    target_calories = None
    if plan and plan.daily_calories:
        target_calories = plan.daily_calories
    elif hasattr(user, 'daily_calories_target') and user.daily_calories_target:
        target_calories = user.daily_calories_target
    else:
        # Fallback inteligente: estimación personalizada por perfil (Harris-Benedict + objetivo)
        try:
            target_calories = PersonalizedNutritionService(user).calculate_daily_calories()
        except Exception:
            target_calories = None

    macros_target = {
        'calories': target_calories,
        'protein': float(plan.protein_grams) if plan and plan.protein_grams else None,
        'carbs': float(plan.carbs_grams) if plan and plan.carbs_grams else None,
        'fat': float(plan.fat_grams) if plan and plan.fat_grams else None,
    }

    return Response({
        'plan': plan_data,
        'user_id': user.id,
        'period': {
            'start_date': start_date,
            'end_date': end_date,
            'days': days,
        },
        'macros_target': macros_target,
        'macro_intake': {
            'totals': {k: (float(v) if v is not None else 0) for k, v in aggregates.items()},
            'per_day': per_day,
        },
        'logs_count': meal_logs.count(),
    })


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_user_plan_history(request, user_id: int):
    """
    Historial de cambios de planes nutricionales de un usuario (hasta 100 registros más recientes).
    """
    user = get_object_or_404(User, pk=user_id)
    history_qs = user.nutrition_plan_history.select_related('changed_by').order_by('-created_at')[:100]
    serializer = NutritionPlanHistorySerializer(history_qs, many=True)
    return Response({
        'user_id': user.id,
        'count': len(serializer.data),
        'history': serializer.data,
    })


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_user_meal_logs(request, user_id: int):
    """
    Logs de comidas de un usuario para revisión/admin.
    Query params:
      - start_date / end_date (YYYY-MM-DD) para filtrar rango
      - limit (por defecto 50, máx 200)
    """
    user = get_object_or_404(User, pk=user_id)

    start_param = request.query_params.get('start_date')
    end_param = request.query_params.get('end_date')

    start_date = None
    end_date = None
    if start_param:
        try:
            start_date = datetime.strptime(start_param, '%Y-%m-%d').date()
        except ValueError:
            start_date = None
    if end_param:
        try:
            end_date = datetime.strptime(end_param, '%Y-%m-%d').date()
        except ValueError:
            end_date = None

    logs_qs = MealLog.objects.filter(user=user)
    if start_date:
        logs_qs = logs_qs.filter(date__gte=start_date)
    if end_date:
        logs_qs = logs_qs.filter(date__lte=end_date)

    logs_qs = logs_qs.order_by('-date', '-created_at')
    limit = min(int(request.query_params.get('limit', 50)), 200)
    logs_qs = logs_qs[:limit]

    serializer = MealLogSerializer(logs_qs, many=True)

    aggregates = logs_qs.aggregate(
        calories=Sum('calories'),
        protein=Sum('protein'),
        carbs=Sum('carbs'),
        fat=Sum('fat'),
    )

    return Response({
        'user_id': user.id,
        'count': len(serializer.data),
        'totals': {k: (float(v) if v is not None else 0) for k, v in aggregates.items()},
        'logs': serializer.data,
    })


@api_view(['PATCH', 'DELETE'])
@perm_classes([IsAdminUser])
def admin_user_meal_log_detail(request, user_id: int, log_id):
    """
    Editar o eliminar un log de comida del usuario (uso admin/staff).
    """
    user = get_object_or_404(User, pk=user_id)
    log = get_object_or_404(MealLog, pk=log_id, user=user)

    if request.method == 'DELETE':
        log.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = MealLogSerializer(log, data=request.data, partial=True, context={'request': request})
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_user_plans_stats(request):
    """
    Estadísticas generales de planes nutricionales de usuarios
    """
    # Estadísticas básicas
    total_user_plans = NutritionPlan.objects.filter(user__isnull=False).count()
    active_user_plans = NutritionPlan.objects.filter(user__isnull=False, is_active=True).count()
    inactive_user_plans = NutritionPlan.objects.filter(user__isnull=False, is_active=False).count()
    
    # Usuarios con planes activos
    users_with_active_plans = User.objects.filter(
        nutrition_plans__is_active=True
    ).distinct().count()
    
    # Planes recientes
    now = timezone.now()
    recent_plans_7_days = NutritionPlan.objects.filter(
        user__isnull=False,
        created_at__gte=now - timedelta(days=7)
    ).count()
    recent_plans_30_days = NutritionPlan.objects.filter(
        user__isnull=False,
        created_at__gte=now - timedelta(days=30)
    ).count()
    
    # Estadísticas de calorías
    calories_stats = NutritionPlan.objects.filter(
        user__isnull=False,
        daily_calories__isnull=False
    ).aggregate(
        average=Avg('daily_calories'),
        min=Count('daily_calories', filter=Q(daily_calories__isnull=False)),
        max=Count('daily_calories', filter=Q(daily_calories__isnull=False))
    )
    
    # Obtener min y max reales
    plans_with_calories = NutritionPlan.objects.filter(
        user__isnull=False,
        daily_calories__isnull=False
    )
    if plans_with_calories.exists():
        calories_stats['min'] = plans_with_calories.aggregate(Min('daily_calories'))['daily_calories__min']
        calories_stats['max'] = plans_with_calories.aggregate(Max('daily_calories'))['daily_calories__max']
    else:
        calories_stats['min'] = None
        calories_stats['max'] = None
    
    # Planes más populares (por nombre de plan)
    popular_plans = NutritionPlan.objects.filter(
        user__isnull=False
    ).values('name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    most_popular_plans = [
        {'name': item['name'], 'count': item['count']}
        for item in popular_plans
    ]
    
    # Timeline de creación
    creation_timeline = {
        'last_24_hours': NutritionPlan.objects.filter(
            user__isnull=False,
            created_at__gte=now - timedelta(hours=24)
        ).count(),
        'last_7_days': recent_plans_7_days,
        'last_30_days': recent_plans_30_days,
        'last_90_days': NutritionPlan.objects.filter(
            user__isnull=False,
            created_at__gte=now - timedelta(days=90)
        ).count(),
    }
    
    # Top usuarios por número de planes
    top_users = User.objects.filter(
        nutrition_plans__isnull=False
    ).annotate(
        plan_count=Count('nutrition_plans')
    ).order_by('-plan_count')[:10]
    
    top_users_by_plans = [
        {
            'email': user.email,
            'name': f"{user.first_name} {user.last_name}".strip() or user.email,
            'plan_count': user.plan_count
        }
        for user in top_users
    ]
    
    # Estadísticas de cambios de planes
    total_changes = NutritionPlanHistory.objects.count()
    changes_by_admins = NutritionPlanHistory.objects.filter(
        changed_by__is_staff=True
    ).count()
    changes_by_users = NutritionPlanHistory.objects.filter(
        Q(changed_by__is_staff=False) | Q(changed_by__isnull=True)
    ).count()
    changes_last_30_days = NutritionPlanHistory.objects.filter(
        created_at__gte=now - timedelta(days=30)
    ).count()
    
    # Planes más cambiados
    most_changed_plans = NutritionPlanHistory.objects.filter(
        new_plan__isnull=False
    ).values('new_plan__name').annotate(
        change_count=Count('id')
    ).order_by('-change_count')[:10]
    
    most_changed_plans_list = [
        {
            'plan_name': item['new_plan__name'] or 'Plan eliminado',
            'change_count': item['change_count']
        }
        for item in most_changed_plans
    ]
    
    # Razones de cambio
    change_reasons = NutritionPlanHistory.objects.values('reason').annotate(
        count=Count('id')
    )
    
    reason_display_map = {
        'user_request': 'Solicitud del usuario',
        'admin_change': 'Cambio por administrador',
        'auto_assigned': 'Asignación automática',
        'goal_change': 'Cambio de objetivo',
        'upgrade': 'Actualización de plan',
        'other': 'Otro',
    }
    
    change_reasons_list = [
        {
            'reason': item['reason'],
            'reason_display': reason_display_map.get(item['reason'], item['reason']),
            'count': item['count']
        }
        for item in change_reasons
    ]
    
    # Distribución de calorías
    calorie_distribution = {
        'low': NutritionPlan.objects.filter(
            user__isnull=False,
            daily_calories__lt=1500
        ).count(),
        'moderate': NutritionPlan.objects.filter(
            user__isnull=False,
            daily_calories__gte=1500,
            daily_calories__lt=2000
        ).count(),
        'high': NutritionPlan.objects.filter(
            user__isnull=False,
            daily_calories__gte=2000,
            daily_calories__lt=2500
        ).count(),
        'very_high': NutritionPlan.objects.filter(
            user__isnull=False,
            daily_calories__gte=2500
        ).count(),
    }
    
    return Response({
        'total_user_plans': total_user_plans,
        'active_user_plans': active_user_plans,
        'inactive_user_plans': inactive_user_plans,
        'users_with_active_plans': users_with_active_plans,
        'recent_plans_7_days': recent_plans_7_days,
        'recent_plans_30_days': recent_plans_30_days,
        'calories_stats': {
            'average': round(calories_stats['average'] or 0, 1),
            'min': calories_stats['min'],
            'max': calories_stats['max'],
        },
        'most_popular_plans': most_popular_plans,
        'creation_timeline': creation_timeline,
        'top_users_by_plans': top_users_by_plans,
        'plan_changes': {
            'total_changes': total_changes,
            'changes_by_admins': changes_by_admins,
            'changes_by_users': changes_by_users,
            'changes_last_30_days': changes_last_30_days,
            'most_changed_plans': most_changed_plans_list,
            'change_reasons': change_reasons_list,
        },
        'calorie_distribution': calorie_distribution,
    })


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_user_plans_usage_stats(request):
    """
    Estadísticas de uso de planes por defecto
    """
    # Planes por defecto (sin usuario asignado o marcados como template)
    default_plans = NutritionPlan.objects.filter(
        Q(user__isnull=True) | Q(is_template=True)
    ).distinct()
    
    total_default_plans = default_plans.count()
    active_default_plans = default_plans.filter(is_active=True).count()
    
    # Uso de cada plan por defecto
    plan_usage = []
    for plan in default_plans:
        users_count = NutritionPlan.objects.filter(
            name=plan.name,
            user__isnull=False
        ).values('user').distinct().count()
        
        plan_usage.append({
            'plan_id': str(plan.id),
            'plan_name': plan.name,
            'daily_calories': plan.daily_calories or 0,
            'target_audience': getattr(plan, 'target_audience', None),
            'min_role_required': getattr(plan, 'min_role_required', None),
            'users_count': users_count,
            'is_default': getattr(plan, 'is_default', False),
        })
    
    # Ordenar por número de usuarios
    plan_usage.sort(key=lambda x: x['users_count'], reverse=True)
    
    return Response({
        'total_default_plans': total_default_plans,
        'active_default_plans': active_default_plans,
        'plan_usage': plan_usage,
    })


@api_view(['GET'])
@perm_classes([IsAdminUser])
def admin_user_plans_history(request):
    """
    Historial general de cambios de planes nutricionales
    """
    limit = min(int(request.query_params.get('limit', 100)), 500)
    user_id = request.query_params.get('user_id')
    
    history_qs = NutritionPlanHistory.objects.select_related(
        'user', 'old_plan', 'new_plan', 'changed_by'
    ).order_by('-created_at')
    
    if user_id and user_id != 'all':
        history_qs = history_qs.filter(user_id=user_id)
    
    history_qs = history_qs[:limit]
    
    serializer = NutritionPlanHistorySerializer(history_qs, many=True)
    
    return Response({
        'count': len(serializer.data),
        'history': serializer.data,
    })


# =============================================================================
# CATEGORÍAS DE EQUIVALENCIA (admin CRUD)
# =============================================================================

class AdminEquivalenceCategoryViewSet(viewsets.ModelViewSet):
    """
    CRUD de categorías de equivalencia de alimentos.
    Las categorías del sistema (is_system=True) solo pueden editarse (no eliminarse).
    """
    queryset = EquivalenceCategory.objects.all()
    serializer_class = EquivalenceCategorySerializer
    permission_classes = [IsAdminUser]
    lookup_field = 'slug'

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.is_system:
            return Response(
                {'error': 'No se pueden eliminar las categorías del sistema.'},
                status=status.HTTP_403_FORBIDDEN
            )
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
