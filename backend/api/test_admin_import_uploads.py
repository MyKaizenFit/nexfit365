import csv
import io
import unicodedata
import uuid

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from nutrition.models import Food, Recipe, RecipeIngredient
from workouts.models import Exercise, WorkoutDay, WorkoutDayExercise, WorkoutProgram


class AdminImportUploadSimulationTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_superuser(
            email='qa-import-admin@test.com',
            password='Admin123456!',
            first_name='QA',
            last_name='Import'
        )
        self.client.force_authenticate(user=self.admin_user)

        self.seed_exercise = Exercise.objects.create(
            name='QA Seed Exercise',
            description='Ejercicio semilla QA',
            instructions='Instrucciones semilla',
            category='strength',
            difficulty='beginner',
            muscle_groups=['chest'],
            equipment=['dumbbells'],
            is_system=False,
            is_active=True,
            created_by=self.admin_user,
        )

        self.seed_plan = WorkoutProgram.objects.create(
            name='QA Seed Plan',
            description='Plan semilla QA',
            difficulty='beginner',
            goal='general_fitness',
            location='any',
            duration_weeks=4,
            days_per_week=3,
            estimated_duration_minutes=60,
            is_template=True,
            is_active=True,
            is_system=False,
        )
        seed_day = WorkoutDay.objects.create(
            program=self.seed_plan,
            name='Día 1',
            day_number=1,
            order_index=1,
            is_rest_day=False,
        )
        WorkoutDayExercise.objects.create(
            workout_day=seed_day,
            exercise=self.seed_exercise,
            sets=3,
            reps='10',
            rest_seconds=60,
            order_index=1,
        )

        self.seed_food = Food.objects.create(
            name='QA Seed Food',
            brand='Marca QA',
            calories=120,
            protein=10,
            carbs=12,
            fat=4,
            category='Proteínas',
            serving_size=100,
            serving_unit='g',
            created_by=self.admin_user,
        )

        self.seed_recipe = Recipe.objects.create(
            name='QA Seed Recipe',
            description='Receta semilla QA',
            category='Almuerzo',
            difficulty='Fácil',
            prep_time_minutes=15,
            servings=1,
            instructions='Mezclar y servir',
            is_system=False,
            is_active=True,
            created_by=self.admin_user,
        )
        RecipeIngredient.objects.create(
            recipe=self.seed_recipe,
            food=self.seed_food,
            quantity=100,
            unit='g',
            order=1,
        )
        self.seed_recipe.refresh_from_db()

    @staticmethod
    def _normalize(value):
        normalized = unicodedata.normalize('NFKD', str(value or '').strip().lower())
        return ''.join(ch for ch in normalized if not unicodedata.combining(ch))

    def _find_key(self, keys, candidates):
        normalized_map = {self._normalize(key): key for key in keys}
        for candidate in candidates:
            resolved = normalized_map.get(self._normalize(candidate))
            if resolved:
                return resolved
        return None

    def _assert_import_response(self, payload, stats_key=None):
        if stats_key:
            stats = payload.get('stats', {})
            section = stats.get(stats_key, {})
            self.assertGreaterEqual(section.get('updated', 0), 1)
            self.assertGreaterEqual(section.get('created', 0), 1)
            return

        self.assertGreaterEqual(payload.get('updated', 0), 1)
        self.assertGreaterEqual(payload.get('created', 0), 1)

    def _run_csv_import_simulation(
        self,
        export_url,
        import_url,
        name_candidates,
        update_candidates,
        new_name,
        model,
        stats_key=None,
    ):
        export_response = self.client.get(export_url)
        self.assertEqual(export_response.status_code, 200)

        content = export_response.content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        self.assertGreaterEqual(len(rows), 1)

        name_key = self._find_key(reader.fieldnames or [], name_candidates)
        self.assertIsNotNone(name_key)

        update_key = self._find_key(reader.fieldnames or [], update_candidates)

        rows[0][name_key] = rows[0].get(name_key) or 'Registro Semilla'
        if update_key:
            rows[0][update_key] = 'Actualizado por prueba automática CSV'

        new_row = dict(rows[0])
        new_row[name_key] = new_name
        if update_key:
            new_row[update_key] = 'Creado por prueba automática CSV'
        rows.append(new_row)

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=reader.fieldnames)
        writer.writeheader()
        writer.writerows(rows)

        upload = SimpleUploadedFile('simulated_import.csv', output.getvalue().encode('utf-8'))
        import_response = self.client.post(import_url, {'file': upload}, format='multipart')
        self.assertEqual(import_response.status_code, 200)

        payload = import_response.json()
        self._assert_import_response(payload, stats_key=stats_key)
        self.assertTrue(model.objects.filter(name=new_name).exists())

    def _run_excel_import_simulation(
        self,
        export_url,
        import_url,
        name_candidates,
        update_candidates,
        new_name,
        model,
        sheet_name=None,
        stats_key=None,
    ):
        export_response = self.client.get(export_url)
        self.assertEqual(export_response.status_code, 200)

        workbook = load_workbook(io.BytesIO(export_response.content))
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        self.assertGreaterEqual(worksheet.max_row, 2)

        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in worksheet[1]]
        name_key = self._find_key(headers, name_candidates)
        self.assertIsNotNone(name_key)

        update_key = self._find_key(headers, update_candidates)

        normalized_headers = {self._normalize(header): index for index, header in enumerate(headers)}
        name_index = normalized_headers[self._normalize(name_key)]
        update_index = normalized_headers[self._normalize(update_key)] if update_key else None

        source_row_values = [worksheet.cell(row=2, column=index + 1).value for index in range(len(headers))]

        if update_index is not None:
            worksheet.cell(row=2, column=update_index + 1).value = 'Actualizado por prueba automática Excel'
            source_row_values[update_index] = 'Creado por prueba automática Excel'

        source_row_values[name_index] = new_name
        worksheet.append(source_row_values)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        upload = SimpleUploadedFile(
            'simulated_import.xlsx',
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        import_response = self.client.post(import_url, {'file': upload}, format='multipart')
        self.assertEqual(import_response.status_code, 200)

        payload = import_response.json()
        self._assert_import_response(payload, stats_key=stats_key)
        self.assertTrue(model.objects.filter(name=new_name).exists())

    def test_exercises_csv_upload_simulation(self):
        new_name = f'QA CSV Exercise {uuid.uuid4().hex[:8]}'
        self._run_csv_import_simulation(
            export_url='/api/admin/exercises/export-csv/',
            import_url='/api/admin/exercises/import-csv/',
            name_candidates=['nombre', 'name'],
            update_candidates=['descripción', 'descripcion', 'description'],
            new_name=new_name,
            model=Exercise,
        )

    def test_exercises_excel_upload_simulation(self):
        new_name = f'QA XLSX Exercise {uuid.uuid4().hex[:8]}'
        self._run_excel_import_simulation(
            export_url='/api/admin/exercises/export-excel/',
            import_url='/api/admin/exercises/import-excel/',
            name_candidates=['nombre', 'name'],
            update_candidates=['descripción', 'descripcion', 'description'],
            new_name=new_name,
            model=Exercise,
            sheet_name='Ejercicios',
        )

    def test_workout_plans_csv_upload_simulation(self):
        new_name = f'QA CSV Plan {uuid.uuid4().hex[:8]}'
        self._run_csv_import_simulation(
            export_url='/api/admin/workouts/workouts/export_csv/',
            import_url='/api/admin/workouts/workouts/import_csv/',
            name_candidates=['plan_name', 'nombre_plan', 'name', 'nombre'],
            update_candidates=['plan_description', 'descripcion_plan', 'descripción_plan', 'description'],
            new_name=new_name,
            model=WorkoutProgram,
            stats_key='plans',
        )

    def test_workout_plans_excel_upload_simulation(self):
        new_name = f'QA XLSX Plan {uuid.uuid4().hex[:8]}'
        self._run_excel_import_simulation(
            export_url='/api/admin/workouts/workouts/export_excel/',
            import_url='/api/admin/workouts/workouts/import_excel/',
            name_candidates=['Nombre', 'name', 'nombre'],
            update_candidates=['Descripción', 'descripcion', 'description'],
            new_name=new_name,
            model=WorkoutProgram,
            sheet_name='Planes',
            stats_key='plans',
        )

    def test_foods_csv_upload_simulation(self):
        new_name = f'QA CSV Food {uuid.uuid4().hex[:8]}'
        self._run_csv_import_simulation(
            export_url='/api/admin/nutrition/foods/export-csv/',
            import_url='/api/admin/nutrition/foods/import-csv/',
            name_candidates=['nombre', 'name'],
            update_candidates=['marca', 'brand'],
            new_name=new_name,
            model=Food,
        )

    def test_foods_excel_upload_simulation(self):
        new_name = f'QA XLSX Food {uuid.uuid4().hex[:8]}'
        self._run_excel_import_simulation(
            export_url='/api/admin/nutrition/foods/export-excel/',
            import_url='/api/admin/nutrition/foods/import-excel/',
            name_candidates=['nombre', 'name'],
            update_candidates=['marca', 'brand'],
            new_name=new_name,
            model=Food,
            sheet_name='Alimentos',
        )

    def test_recipes_csv_upload_simulation(self):
        new_name = f'QA CSV Recipe {uuid.uuid4().hex[:8]}'
        self._run_csv_import_simulation(
            export_url='/api/admin/nutrition/recipes/export-csv/',
            import_url='/api/admin/nutrition/recipes/import-csv/',
            name_candidates=['nombre', 'name'],
            update_candidates=['descripción', 'descripcion', 'description'],
            new_name=new_name,
            model=Recipe,
        )

    def test_recipes_excel_upload_simulation(self):
        new_name = f'QA XLSX Recipe {uuid.uuid4().hex[:8]}'
        self._run_excel_import_simulation(
            export_url='/api/admin/nutrition/recipes/export-excel/',
            import_url='/api/admin/nutrition/recipes/import-excel/',
            name_candidates=['nombre', 'name'],
            update_candidates=['descripción', 'descripcion', 'description'],
            new_name=new_name,
            model=Recipe,
            sheet_name='Recetas',
        )